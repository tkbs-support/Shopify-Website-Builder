"""RMD-specific narrative edits from the adversarial-review gate (Step 4.6).
Edits cells by CONTENT match (robust to row shifts). Re-run safe (idempotent-ish).
Generalizable bugs were already fixed in the skill scripts; this is the client-specific layer.
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NAVY = "1F3864"; GRAY = "808080"
F = lambda **kw: Font(name="Arial", **kw)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SEV_FILL = {"CRITICAL": "F8CBAD", "WARNING": "FFE699", "OK": "C6EFCE", "VERIFY": "FFE699"}
SEV_FONTC = {"CRITICAL": "9C0006", "WARNING": "9C6500", "OK": "276221", "VERIFY": "9C6500"}

SEO = "Rachel Marie Designs - SEO & AEO Plan (June 2026).xlsx"
INV = "Rachel Marie Designs - Inventory Recovery Plan (June 2026).xlsx"


def find_row(ws, col, text, maxr=200):
    for r in range(1, min(ws.max_row, maxr) + 1):
        v = ws.cell(row=r, column=col).value
        if v and str(v).strip().lower() == text.strip().lower():
            return r
    return None


def find_row_contains(ws, col, text, maxr=200):
    for r in range(1, min(ws.max_row, maxr) + 1):
        v = ws.cell(row=r, column=col).value
        if v and text.lower() in str(v).lower():
            return r
    return None


def set_sev(ws, r, c, sev):
    cell = ws.cell(row=r, column=c, value=sev)
    cell.fill = PatternFill("solid", start_color=SEV_FILL[sev])
    cell.font = F(size=10, bold=(sev == "CRITICAL"), color=SEV_FONTC[sev])
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER


# ================= SEO WORKBOOK =================
wb = load_workbook(SEO)

# NOTE: collection-meta boilerplate (CRITICAL), blog-freshness recency, the agentic-commerce
# strength row, and product/home microdata are now handled NATIVELY by the skill build
# (analyze-audit.py + build-reports.py). Only genuinely client-specific narrative remains below.

# --- SEO Fix Roadmap: name the real crawl sinks in the indexation row ---
ws = wb["Fix Roadmap"]
r = find_row_contains(ws, 2, "Indexation hygiene")
if r:
    ws.cell(row=r, column=3, value="incl. smart collections quick-order (907) + all-items (872) as crawl sinks; duplicate collections bracelets-1/earrings-1/rings-1 + misspelled “jewlery”; ~24 ‘-copy’ products")

# --- Collection SEO: keyworded proposed titles for top navigable collections ---
ws = wb["Collection SEO"]
TITLES = {
    "earrings": "Handmade Crystal Earrings | Rachel Marie Designs",
    "bracelets": "Handmade Crystal Bracelets | Rachel Marie Designs",
    "necklaces": "Handmade Crystal Necklaces | Rachel Marie Designs",
    "rings": "Handmade Crystal Rings | Rachel Marie Designs",
    "jewelry": "Handmade Swarovski Crystal Jewelry | Rachel Marie Designs",
    "new-arrivals": "New Crystal Jewelry Arrivals | Rachel Marie Designs",
    "long-necklaces": "Long Crystal Necklaces | Rachel Marie Designs",
    "short-necklaces": "Short Crystal Necklaces | Rachel Marie Designs",
    "pendant": "Crystal Pendant Necklaces | Rachel Marie Designs",
    "dangle": "Crystal Dangle Earrings | Rachel Marie Designs",
    "essentials": "Everyday Crystal Jewelry Essentials | Rachel Marie Designs",
}
# find the URL column (header "URL") and the proposed-title column (header contains "Proposed")
hdr_row = find_row(ws, 1, "Category") or 4
url_col = prop_col = None
for c in range(1, 12):
    h = ws.cell(row=hdr_row, column=c).value or ""
    if str(h).strip().lower() == "url":
        url_col = c
    if "proposed" in str(h).lower():
        prop_col = c
edits = 0
if url_col and prop_col:
    for r in range(hdr_row + 1, ws.max_row + 1):
        u = ws.cell(row=r, column=url_col).value or ""
        handle = str(u).rsplit("/", 1)[-1].strip().lower()
        if handle in TITLES:
            ws.cell(row=r, column=prop_col, value=TITLES[handle])
            edits += 1
wb.save(SEO)
print(f"SEO workbook updated (collection titles set: {edits})")

# ================= INVENTORY WORKBOOK =================
# Size framing is now handled type-aware INSIDE the skill (REAL SIZE GAPS KPI + Sizing Readiness
# by Type section + split Phase 4a/b/c roadmap), so no inventory post-edits are needed here.
print("Inventory workbook: no post-edits (type-aware sizing handled by the skill build)")
