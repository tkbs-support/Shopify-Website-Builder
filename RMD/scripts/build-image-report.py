"""Build the Image Quality Audit workbook + AI Visualizer manifest/ZIP from
data/image-analysis.json. Run from the RMD folder after analyze-images.py.

Outputs:
  - Rachel Marie Designs - Image Quality Audit (June 2026).xlsx
  - data/visualizer-manifest.json   (matches visualizer Image classification vocab)
  - data/visualizer-import.zip      (folder-per-product, variant-named images)
"""
import datetime
import json
import os
import re
import zipfile

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DATA = "data"
A = json.load(open(os.path.join(DATA, "image-analysis.json"), encoding="utf-8"))
SHOP = json.load(open(os.path.join(DATA, "shop.json"), encoding="utf-8"))
CLIENT = SHOP["name"]
TODAY = datetime.date.today()
AUDIT_DATE = TODAY.strftime("%B %d, %Y")
MONTH_YEAR = TODAY.strftime("%B %Y")

NAVY = "1F3864"
TEAL = "2E9BA6"
RED = "C00000"
ORANGE = "ED7D31"
GREEN = "548235"
GRAY = "808080"
SECTION_FILL = "D9E2F3"
F = lambda **kw: Font(name="Arial", **kw)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

T = A["totals"]
N = A["scope"]["in_stock_products"]
N_IMG = A["scope"]["images_analyzed"]

ISSUE_LABEL = {
    "NO_IMAGES": "No images at all",
    "VARIANTS_UNLINKED": "Styles/variants not linked to any image",
    "ONE_IMAGE_MANY_COLORS": "One photo shared by 3+ color variants",
    "STOCK_DUPLICATE": "Same photo reused on other products (stock)",
    "STOCK_SWATCH_NAME": "Image file named after crystal color (stock swatch)",
    "LOW_RES": "Low resolution (<800px)",
    "BLURRY_REVIEW": "Possible blur/grain — review",
}
ACTION = {
    "NO_IMAGES": "Generate or photograph — top priority",
    "VARIANTS_UNLINKED": "Link existing images to variants (free fix)",
    "ONE_IMAGE_MANY_COLORS": "Generate per-color images (AI Visualizer)",
    "STOCK_DUPLICATE": "Replace with product-specific image (AI Visualizer)",
    "STOCK_SWATCH_NAME": "Replace swatch with product shot (AI Visualizer)",
    "LOW_RES": "Re-shoot or upscale",
    "BLURRY_REVIEW": "Eyeball — replace if grainy",
}


def title_block(ws, title, subtitle, span=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = F(size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = F(size=10, italic=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="2F5496")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def kpi(ws, row, col, label, value, sub, color, fmt=None):
    for r in range(row, row + 3):
        for cc in range(col, col + 2):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", start_color=color)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    c = ws.cell(row=row, column=col, value=label)
    c.font = F(size=9, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    c = ws.cell(row=row + 1, column=col, value=value)
    c.font = F(size=20, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        c.number_format = fmt
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
    c = ws.cell(row=row + 2, column=col, value=sub)
    c.font = F(size=8, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)


def section(ws, row, text, span=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F(size=11, bold=True, color=NAVY)
    c.fill = PatternFill("solid", start_color=SECTION_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def thead(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def trow(ws, row, values, fmts=None, start_col=1):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = F(size=10)
        c.border = BORDER
        if fmts and fmts[i]:
            c.number_format = fmts[i]


def labels(show_val=False, show_percent=False):
    d = DataLabelList()
    d.showVal = show_val
    d.showPercent = show_percent
    d.showSerName = False
    d.showCatName = False
    d.showLegendKey = False
    d.showBubbleSize = False
    return d


def color_points(series, colors):
    for i, col in enumerate(colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = col
        series.dPt.append(pt)


wb = Workbook()
ws = wb.active
ws.title = "Summary"
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKLM", [42, 12, 12, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11]):
    ws.column_dimensions[col].width = w

clean = N - T.get("products_with_any_issue", 0)
title_block(ws, f"{CLIENT} — Product Image Quality Audit",
            f"Audit date: {AUDIT_DATE}  |  {SHOP['domain']}  |  Scope: {N} in-stock products, {N_IMG} images analyzed  |  "
            "Method: resolution + variant linkage + duplicate detection + blur scoring", span=13)
kpi(ws, 4, 1, "PRODUCTS WITH IMAGE ISSUES", T.get("products_with_any_issue", 0) / N,
    f"{T.get('products_with_any_issue', 0)} of {N} in-stock products", RED, "0.0%")
kpi(ws, 4, 4, "REUSED / STOCK PHOTOS", T.get("images_dup", 0),
    "same image on 2+ products", RED, "#,##0")
kpi(ws, 4, 7, "VARIANTS WITH NO IMAGE", len(A.get("variant_rows", [])),
    "styles not linked to any photo", ORANGE, "#,##0")
kpi(ws, 4, 10, "LOW-RES IMAGES", T.get("images_low_res", 0),
    "below 800px — too small to zoom", ORANGE, "#,##0")
ws.row_dimensions[4].height = 26
ws.row_dimensions[5].height = 30

r = 8
hs = r
section(ws, r, "IMAGE HEALTH AT A GLANCE — in-stock products", span=5)
r += 1
thead(ws, r, ["Status", "Products", "% of in-stock"])
hh = r
for i, (label, n) in enumerate([("Clean — no image issues found", clean),
                                ("Has at least one image issue", T.get("products_with_any_issue", 0))]):
    rr = r + 1 + i
    trow(ws, rr, [label, n, f"=B{rr}/{N}"], fmts=[None, "#,##0", "0.0%"])
d = DoughnutChart()
d.title = "In-stock products: image health"
d.add_data(Reference(ws, min_col=2, min_row=hh + 1, max_row=hh + 2), titles_from_data=False)
d.set_categories(Reference(ws, min_col=1, min_row=hh + 1, max_row=hh + 2))
d.height = 7
d.width = 11
d.dataLabels = labels(show_percent=True)
color_points(d.series[0], [GREEN, RED])
ws.add_chart(d, "F" + str(hs))
r = max(r + 4, hs + 16)

bs = r
section(ws, r, "ISSUES BY TYPE — products affected", span=5)
r += 1
thead(ws, r, ["Issue", "Products", "% of in-stock"])
bh = r
issue_counts = [
    ("STOCK_DUPLICATE", T.get("products_stock_dup", 0)),
    ("VARIANTS_UNLINKED", T.get("products_variants_unlinked", 0)),
    ("ONE_IMAGE_MANY_COLORS", T.get("products_one_image_many_colors", 0)),
    ("NO_IMAGES", T.get("products_no_images", 0)),
    ("LOW_RES", T.get("products_low_res", 0)),
    ("STOCK_SWATCH_NAME", T.get("products_swatch_name", 0)),
    ("BLURRY_REVIEW", T.get("products_blurry", 0)),
]
issue_counts.sort(key=lambda kv: -kv[1])
for i, (key, n) in enumerate(issue_counts):
    rr = r + 1 + i
    trow(ws, rr, [ISSUE_LABEL[key], n, f"=B{rr}/{N}"], fmts=[None, "#,##0", "0.0%"])
ch = BarChart()
ch.type = "bar"
ch.title = "In-stock products affected per issue"
ch.add_data(Reference(ws, min_col=2, min_row=bh + 1, max_row=bh + len(issue_counts)), titles_from_data=False)
ch.set_categories(Reference(ws, min_col=1, min_row=bh + 1, max_row=bh + len(issue_counts)))
ch.legend = None
ch.height = 8.5
ch.width = 16
ch.gapWidth = 60
ch.series[0].graphicalProperties.solidFill = RED
ch.dataLabels = labels(show_val=True)
ws.add_chart(ch, "F" + str(bs))
r = max(r + len(issue_counts) + 2, bs + 18)

ts_ = r
section(ws, r, "BIGGEST STOCK-PHOTO CLUSTERS — one photo, many products", span=5)
r += 1
thead(ws, r, ["Products sharing the photo", "Copies", "Example products"])
multi = [c for c in A["clusters"] if c["products"] > 1][:8]
for i, c in enumerate(multi):
    rr = r + 1 + i
    trow(ws, rr, [c["products"], c["images"], ", ".join(c["product_titles"][:4])],
         fmts=["#,##0", "#,##0", None])
note = ws.cell(row=r + len(multi) + 1, column=1,
               value="Each row is ONE identical image found on multiple products — shoppers see the same crystal stock photo regardless of which product they open.")
note.font = F(size=9, italic=True, color=GRAY)
note.alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r + len(multi) + 1, start_column=1, end_row=r + len(multi) + 2, end_column=5)
r = r + len(multi) + 4

section(ws, r, "METHOD NOTES", span=5)
for i, txt in enumerate([
    f"Duplicates: perceptual hash match across {N_IMG} downloaded images (exact visual copies).",
    f"Blur flag: bottom of sharpness distribution (cutoff {A['scope']['blur_cutoff']}) — a REVIEW list, not a verdict.",
    "Stock-swatch flag: image filename matches a crystal color name rather than a product.",
    "Variant linkage read directly from Shopify (variant has no image assigned).",
]):
    c = ws.cell(row=r + 1 + i, column=1, value=txt)
    c.font = F(size=9, italic=True, color=GRAY)

# ---------- Product Worklist ----------
ws = wb.create_sheet("Product Worklist")
ws.sheet_view.showGridLines = False
title_block(ws, "Product Worklist — what to fix, product by product",
            "Sorted worst-first. Actions reference the AI Visualizer export (data/visualizer-import.zip + manifest).", span=9)
hdr = 4
thead(ws, hdr, ["Product", "Type", "Images", "Variants", "Colors", "Unlinked Variants", "Issues", "Recommended Action", "Store Preview"])
rows = [p for p in A["products"] if p["issues"]]
for i, p in enumerate(rows):
    rr = hdr + 1 + i
    acts = "; ".join(dict.fromkeys(ACTION[k] for k in p["issues"]))
    trow(ws, rr, [p["product"], p["type"], p["images"], p["variants"], p["colors"],
                  p["unlinked_variants"], ", ".join(ISSUE_LABEL[k] for k in p["issues"]),
                  acts, p["preview"]],
         fmts=[None, None, "#,##0", "#,##0", "#,##0", "#,##0", None, None, None])
    ws.cell(row=rr, column=7).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=rr, column=8).alignment = Alignment(wrap_text=True, vertical="top")
for col, w in zip("ABCDEFGHI", [34, 13, 8, 9, 8, 10, 42, 46, 40]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{hdr}:I{hdr + len(rows)}"

# ---------- Image Issues ----------
ws = wb.create_sheet("Image Issues")
ws.sheet_view.showGridLines = False
title_block(ws, "Individual Image Issues",
            "Every analyzed image with at least one flag. Width/height are the original upload dimensions.", span=6)
hdr = 4
thead(ws, hdr, ["Product", "Issues", "Width", "Height", "Sharpness", "Image URL"])
for i, rec in enumerate(A["image_rows"]):
    rr = hdr + 1 + i
    trow(ws, rr, [rec["product"], ", ".join(ISSUE_LABEL.get(k, k) for k in rec["issues"]),
                  rec["width"], rec["height"], rec.get("blur"), rec["image_url"]],
         fmts=[None, None, "#,##0", "#,##0", "#,##0", None])
for col, w in zip("ABCDEF", [34, 44, 9, 9, 11, 70]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{hdr}:F{hdr + len(A['image_rows'])}"

# ---------- Unlinked Variants ----------
ws = wb.create_sheet("Unlinked Variants")
ws.sheet_view.showGridLines = False
title_block(ws, "Variants (Styles) Not Linked To Any Image",
            "These styles show the product's default image regardless of which color the shopper picks.", span=4)
hdr = 4
thead(ws, hdr, ["Product", "Variant / Color", "In Stock?", "Issue"])
for i, v in enumerate(A["variant_rows"]):
    rr = hdr + 1 + i
    trow(ws, rr, [v["product"], v["variant"], "Yes" if v["in_stock"] else "No", v["issue"]])
for col, w in zip("ABCD", [36, 30, 10, 40]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{hdr}:D{hdr + len(A['variant_rows'])}"

# ---------- Stock Photo Clusters ----------
ws = wb.create_sheet("Stock Photo Clusters")
ws.sheet_view.showGridLines = False
title_block(ws, "Identical Images Used Across Products",
            "Each row = one unique image and every product it appears on.", span=4)
hdr = 4
thead(ws, hdr, ["# Products", "# Copies", "Products", "Sample Image URL"])
for i, c in enumerate([c for c in A["clusters"] if c["products"] > 1]):
    rr = hdr + 1 + i
    trow(ws, rr, [c["products"], c["images"], ", ".join(c["product_titles"]), c["sample_url"]],
         fmts=["#,##0", "#,##0", None, None])
    ws.cell(row=rr, column=3).alignment = Alignment(wrap_text=True, vertical="top")
for col, w in zip("ABCD", [11, 10, 80, 60]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------- Fix Roadmap ----------
ws = wb.create_sheet("Fix Roadmap")
ws.sheet_properties.tabColor = GREEN
ws.sheet_view.showGridLines = False
title_block(ws, "Image Recovery — Fix Roadmap", "Sequenced cheapest-first.", span=5)
hdr = 4
thead(ws, hdr, ["Phase", "Action", "Scope", "Effort", "Notes"])
road = [
    ("Phase 1", "Link existing images to their variants in Shopify admin (no new photos needed).",
     f"{len(A['variant_rows'])} variants", "Hours", "Free fix — images already exist, just unassigned"),
    ("Phase 2", "Generate product-specific images with the AI Visualizer for stock/duplicate/swatch-only products. Import file provided.",
     f"{T.get('products_stock_dup', 0) + T.get('products_one_image_many_colors', 0)} products (see Worklist)",
     "Batch generation + review", "visualizer-import.zip + visualizer-manifest.json are ready"),
    ("Phase 3", "Photograph or generate images for products with none at all.",
     f"{T.get('products_no_images', 0)} products", "Per product", "Highest visibility impact"),
    ("Phase 4", "Replace low-resolution and confirmed-grainy images.",
     f"{T.get('images_low_res', 0)} low-res + {T.get('images_blurry', 0)} flagged for review", "Per image", "Review the Blurry list by eye first"),
]
for i, row in enumerate(road):
    rr = hdr + 1 + i
    trow(ws, rr, list(row))
    ws.cell(row=rr, column=1).font = F(size=10, bold=True, color=NAVY)
    for cc in range(1, 6):
        ws.cell(row=rr, column=cc).alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[rr].height = 40
for col, w in zip("ABCDE", [10, 64, 30, 26, 44]):
    ws.column_dimensions[col].width = w

out_xlsx = f"{CLIENT} - Image Quality Audit ({MONTH_YEAR}).xlsx"
wb.save(out_xlsx)
print("saved", out_xlsx)

# =====================================================================
# AI VISUALIZER EXPORT — manifest + ZIP
# =====================================================================
products_full = {}
with open(os.path.join(DATA, "images.jsonl"), encoding="utf-8") as f:
    for line in f:
        n = json.loads(line)
        nid = n.get("id", "")
        pid = n.get("__parentId")
        if pid is None and "/Product/" in nid:
            n["_images"] = []
            n["_variants"] = []
            products_full[nid] = n
        elif pid and "/ProductVariant/" in nid:
            products_full[pid]["_variants"].append(n)
        elif pid:
            products_full[pid]["_images"].append(n)

by_handle = {p["handle"]: p for p in A["products"]}
img_issue_by_url = {}
for rec in A["image_rows"]:
    img_issue_by_url[rec["image_url"]] = rec["issues"]


def classify(url, width, height):
    issues = img_issue_by_url.get(url, [])
    if "STOCK_SWATCH_NAME" in issues:
        return "swatch", "filename matches crystal color name"
    if "DUPLICATE_ACROSS_PRODUCTS" in issues:
        return "duplicate", "identical image found on other products"
    if "LOW_RES" in issues:
        return "too_small", f"{width}x{height} below 800px"
    if "BLURRY_REVIEW" in issues:
        return "unknown", "low sharpness score - review"
    return "good", ""


def match_variant(filename, alt, colors):
    blob = re.sub(r"[^A-Z0-9]+", " ", (filename + " " + alt).upper())
    best = ""
    for cv in sorted(colors, key=len, reverse=True):
        if cv.upper() in blob:
            return cv
    return best


manifest = {"client": CLIENT, "store": SHOP["myshopify_domain"], "generated": str(TODAY),
            "source": "shopify-image-audit", "products": []}
zip_items = []  # (zip_path, cache_path)

for gid, p in products_full.items():
    if (p.get("totalInventory") or 0) <= 0:
        continue
    wl = by_handle.get(p["handle"])
    if not wl or not wl["issues"]:
        continue
    color_opt = next((o for o in p.get("options") or [] if (o["name"] or "").strip().lower() in ("color", "colors")), None)
    colors = (color_opt or {}).get("values") or []
    images = []
    needs_generation = set(colors)
    for img in p["_images"]:
        fname = img["url"].split("/")[-1].split("?")[0]
        cls, reason = classify(img["url"], img.get("width") or 0, img.get("height") or 0)
        vmatch = match_variant(fname, img.get("altText") or "", colors)
        if cls == "good" and vmatch:
            needs_generation.discard(vmatch)
        images.append({"filename": fname, "variant_name": vmatch, "source_url": img["url"],
                       "width": img.get("width") or 0, "height": img.get("height") or 0,
                       "classification": cls, "classification_reason": reason})
        cache = os.path.join(DATA, "imgcache", img["id"].rsplit("/", 1)[-1] + ".img")
        if os.path.exists(cache):
            safe_prod = re.sub(r'[<>:"/\\|?*]', "_", p["title"])[:60].strip()
            safe_name = re.sub(r'[<>:"/\\|?*]', "_", (vmatch or os.path.splitext(fname)[0]))[:50].strip() or "image"
            zip_items.append((f"{safe_prod}/{safe_name}.png", cache))
    manifest["products"].append({
        "name": p["title"], "slug": p["handle"],
        "category": p.get("productType") or "", "source_url": p.get("onlineStorePreviewUrl") or "",
        "variant_names": colors, "structure_description": "",
        "issues": wl["issues"],
        "generate_variants": sorted(needs_generation),
        "unlinked_variant_count": wl["unlinked_variants"],
        "images": images,
    })

with open(os.path.join(DATA, "visualizer-manifest.json"), "w", encoding="utf-8") as f:
    json.dump(manifest, f, indent=1)
print(f"manifest: {len(manifest['products'])} products -> data/visualizer-manifest.json")

from PIL import Image as PImage
zip_path = os.path.join(DATA, "visualizer-import.zip")
seen = set()
with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
    for zpath, cache in zip_items:
        final = zpath
        n = 2
        while final in seen:
            base, ext = os.path.splitext(zpath)
            final = f"{base}-{n}{ext}"
            n += 1
        seen.add(final)
        try:
            with PImage.open(cache) as im:
                buf = __import__("io").BytesIO()
                im.convert("RGB").save(buf, "JPEG", quality=82, optimize=True)
                encoded = buf.getvalue()
            raw = open(cache, "rb").read()
            use_raw = len(raw) < len(encoded) and raw[:3] == b"\xff\xd8\xff"
            z.writestr(os.path.splitext(final)[0] + ".jpg", raw if use_raw else encoded)
        except Exception:
            pass
print(f"zip: {len(seen)} images -> {zip_path} ({os.path.getsize(zip_path)//1024//1024} MB)")
