"""Dump Fix Roadmap rows + header formatting for both June reports (safe encoding)."""
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

wb = load_workbook("The Clothing Cove - Inventory Recovery Plan (June 2026).xlsx", read_only=True)
ws = wb["Fix Roadmap"]
print("===== Inventory Fix Roadmap")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    vals = [str(v)[:95] if v is not None else "" for v in row]
    if any(vals):
        print(i, "|", " | ".join(vals))
wb.close()

wb = load_workbook("The Clothing Cove - SEO & AEO Plan (June 2026).xlsx")
for name in ["Fix Roadmap", "Technical SEO", "Local SEO & Entity", "AEO Readiness"]:
    ws = wb[name]
    print(f"\n===== {name}: dims={ws.dimensions}, merged={list(ws.merged_cells.ranges)[:6]}")
    # row numbers of each data row (first cell value) for targeting
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        first = str(row[0])[:40] if row[0] is not None else ""
        second = str(row[1])[:50] if len(row) > 1 and row[1] is not None else ""
        if first or second:
            print(f"  r{i}: {first} | {second}")
    # header style sample
    hdr_row = 3
    c = ws.cell(row=hdr_row, column=1)
    print(f"  header r3 style: font={c.font.name},{c.font.size},bold={c.font.bold},color={c.font.color.rgb if c.font.color else None} fill={c.fill.start_color.rgb if c.fill and c.fill.start_color else None}")
    d = ws.cell(row=4, column=1)
    print(f"  data r4 style: font={d.font.name},{d.font.size} fill={d.fill.start_color.rgb if d.fill and d.fill.start_color else None}")
wb.close()
