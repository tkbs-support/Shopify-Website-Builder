# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font
F="The Clothing Cove - SEO & AEO Plan.xlsx"
OKf=PatternFill("solid",start_color="C6EFCE"); OKfont=Font(name="Arial",color="276221")
wb=openpyxl.load_workbook(F)
def ok(ws,cell):
    ws[cell]="OK"; ws[cell].fill=OKf; ws[cell].font=OKfont

# --- Executive Summary condensed snapshot ---
ex=wb["Executive Summary"]
ex["A44"]="TECHNICAL & AEO SNAPSHOT (live site — re-crawled June 17, 2026)"
ex["B46"]="Keyworded + location title now LIVE (new theme)"; ok(ex,"E46")
ex["A47"]="Theme SEO work (TKBS theme)"; ex["B47"]="PUBLISHED & live — title, H1, meta & schema all live"; ok(ex,"E47")

# --- Technical SEO detail sheet ---
ts=wb["Technical SEO"]
ts["A2"]="Live theme: Prestige Updates TKBS - 05-15-2026 (id 145504469107) — PUBLISHED June 2026"
ts["B5"]="PUBLISHED"; ok(ts,"C5"); ts["D5"]="The TKBS theme carrying all the SEO work (keyworded title/H1/meta, JSON-LD, OG) is now PUBLISHED and live."
ts["B6"]="Optimized"; ok(ts,"C6"); ts["D6"]="Live title: “The Clothing Cove | Women's Clothing, Dresses & Brighton Jewelry | Milford, MI”."
ts["B8"]="Present"; ok(ts,"C8"); ts["D8"]="Keyworded H1 live: “The Clothing Cove — Women's Clothing, Dresses & Brighton Jewelry…”."

# --- AEO Readiness entity row ---
aeo=wb["AEO Readiness"]
if isinstance(aeo["B11"].value,str) and "Generic homepage title" in aeo["B11"].value:
    aeo["B11"]="Keyworded homepage title live; ClothingStore schema present (minor: unify Organization/ClothingStore)"
    ok(aeo,"C11")

wb.save(F)
print("stale snapshot rows corrected. charts:", len(ex._charts))
for c in ("A44","B46","E46","B47","E47"): print(f"  Exec {c}: {ex[c].value}")
for c in ("A2","B5","C5","B6","C6","B8","C8"): print(f"  Tech {c}: {ts[c].value}")
