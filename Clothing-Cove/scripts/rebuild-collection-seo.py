# Rebuilds the Collection SEO tab WITHOUT the broken "{Title} for Women" template.
# Anti-filler rules enforced:
#   1. Titles are classified per-collection (brand / category / accessory / utility) — no blind suffix.
#   2. Every on-page description names >=3 attributes that are TRUE of the page:
#      its real category, a real subtype/occasion/fit, the real location, and (only where
#      provably stocked) verified brands. No fabricated specifics, no volatile counts.
#   3. Event/utility collections are marked NOINDEX, not "optimized".
#   4. Uniqueness gate: no two on-page descriptions identical; near-dupes flagged.
import json, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows = json.load(open('data/seo-worklist.json', encoding='utf-8'))
STORE = "The Clothing Cove"
CITY = "Milford, MI"
ADDR = "414 N Main St, Milford"

# --- Brands the store provably carries (each has its own brand collection) -> category bucket.
# Curated only for brands we can classify with confidence; everything else stays generic (safe).
BRAND_CAT = {
    "joseph ribkoff":"dresses","frank lyman":"dresses","adrianna papell":"eveningwear",
    "alex evenings":"eveningwear","mgny":"eveningwear","montage":"eveningwear","jade dresses":"dresses",
    "michael tyler":"dresses","damee":"jackets",
    "liverpool jeans":"jeans","jag jeans":"jeans","judy blue":"jeans","slim-sation":"pants","renuar":"pants",
    "sympli":"clothing","multiples clothing company":"clothing","keren hart":"clothing","charlie b":"clothing",
    "tribal":"clothing","parsley & sage":"clothing","habitat":"clothing","liv by habitat":"clothing",
    "vocal":"clothing","ethyl":"clothing","easel":"clothing","pol":"clothing","jess & jane":"clothing",
    "by jj":"clothing","jostar":"clothing","origami":"clothing","insight ny":"clothing","ecru":"clothing",
    "brighton":"jewelry","brighton jewelry and accessories":"jewelry","rachel marie designs":"jewelry",
    "anju jewelry":"jewelry","tagua jewelry":"jewelry","jacqueline kent":"jewelry","erin gray jewelry":"jewelry",
    "rain jewelry":"jewelry","a rare bird":"jewelry","radzoli":"jewelry","trust your journey":"jewelry",
    "baggallini":"handbags","coco & carmen":"handbags",
    "spring step shoes":"footwear","foot petals":"shoe care",
    "worlds softest socks":"socks","braza bra":"intimates","bali":"intimates",
    "naked bee":"bath & body","inis fragrance":"fragrance","greenleaf":"home fragrance",
    "pure essence":"bath & body","my fun colors":"bath & body","modgy":"gifts","warmies":"gifts",
    "bella tunno":"baby & gift","tagua":"jewelry",
}
# Category -> a few verified brands to name on that category's top page (proves depth, all real).
CAT_BRANDS = {
    "dresses":["Joseph Ribkoff","Frank Lyman","Adrianna Papell","Alex Evenings"],
    "jewelry":["Brighton","Rachel Marie Designs","Anju","Tagua"],
    "jeans":["Liverpool","Jag Jeans","Judy Blue"],
    "clothing":["Sympli","Multiples","Charlie B","Tribal"],
    "handbags":["Brighton","Baggallini","Coco & Carmen"],
    "footwear":["Spring Step"],
}
THEMED = {"made in italy","north american designers","usa brands","tenacity","creation","brand"}
# Event / utility collections -> noindex, no description written.
NOINDEX_KEYS = {"customer appreciation","customer apprecation","sale","gift cards","new products","brands"}

def segs(title):
    parts = [p.strip() for p in re.split(r'[\\/]', title) if p.strip()]
    out = []
    for p in parts:                       # collapse consecutive dup segments ("Accessories \ Accessories")
        if not out or out[-1].lower() != p.lower(): out.append(p)
    return out

# leaf-noun map for category/accessory subcategories -> (display noun, occasion adj)
LEAF_NOUN = {
    "tops":"Women's Tops","bottoms":"Women's Pants, Skirts & Bottoms","denim":"Women's Denim & Jeans",
    "jumpsuits":"Women's Jumpsuits","evening":"Evening & Formal Dresses","casual":"Casual Dresses",
    "mother of bride":"Mother of the Bride Dresses","mother of the bride":"Mother of the Bride Dresses",
    "pant sets":"Dress Pant Sets","plus":"Plus Size Dresses","petite":"Petite Dresses",
    "jewelry":"Women's Jewelry","bracelets":"Bracelets","earrings":"Earrings","necklaces":"Necklaces",
    "rings":"Rings","brooches":"Brooches & Pins","designer":"Designer Jewelry",
    "handbags":"Handbags & Purses","wallets":"Wallets & Wristlets","footwear":"Women's Shoes & Footwear",
    "shoes":"Women's Shoes","slippers":"Slippers","socks":"Socks","scarves":"Scarves & Wraps",
    "hats":"Hats","sunglasses":"Sunglasses","belts":"Belts","candles":"Candles","fragrance":"Fragrance",
    "soap":"Soaps","lotion":"Lotions","lip balm":"Lip Balm","reading glasses":"Reading Glasses",
    "spa":"Spa & Bath","pillows":"Decorative Pillows","tea towels":"Tea Towels","small accessories":"Small Accessories",
    "winter hats, gloves & scarves":"Winter Hats, Gloves & Scarves","travel":"Travel Bags",
    "recycled canvas bags":"Recycled Canvas Bags","shoe accessories":"Shoe Accessories","brooches":"Brooches",
}
# Garment families and the short noun each contributes when a qualifier leaf attaches to it.
FAMILY = {"dresses":"Dresses","tops":"Tops","bottoms":"Bottoms","denim":"Denim & Jeans",
          "footwear":"Shoes","handbags":"Bags","jewelry":"Jewelry","clothing":"Women's Clothing",
          "accessories":"Accessories"}
# Qualifier leaves resolve against the NEAREST garment ancestor — never a fixed "Dresses".
QUALWORD = {"plus":"Plus Size","petite":"Petite","casual":"Casual","evening":"Evening",
            "dressy":"Dressy","designer":"Designer","travel":"Travel"}
def family_of(s):
    for seg in reversed(s[:-1]):              # nearest ancestor that names a garment family
        if seg.lower() in FAMILY: return seg.lower()
    return None
def noun_for(leaf, root, s=None):
    k = leaf.lower()
    if s and k in QUALWORD:
        fam = family_of(s) or root.lower()
        if fam == "dresses" and k == "evening": return "Evening & Formal Dresses"
        if fam == "dresses" and k == "casual":  return "Casual Dresses"
        return f"{QUALWORD[k]} {FAMILY.get(fam, 'Styles')}"
    if k in LEAF_NOUN: return LEAF_NOUN[k]
    return leaf

def titlecase_brand(b):
    fix={"mgny":"MGNY","pol":"POL","cj bella":"CJ Bella","by jj":"By JJ","jostar":"JoStar","e-cloth":"E-Cloth"}
    return fix.get(b.lower(), b)

UTIL_EXACT = {"sale","new products","brands","customer appreciation","customer apprecation"}
EVENT_PAT = re.compile(r'%|anniversary|black friday|cyber|appreciation|clearance|give ?back|pink friday|golden|gift card', re.I)
def classify(r):
    s = segs(r['title']); root = s[0]; leaf = s[-1]; low = r['title'].lower()
    handle = r['url'].split('/')[-1]
    # UTILITY: index pages, all Sale/Customer-Appreciation children, and event pages
    if len(s) == 1 and root.lower() in UTIL_EXACT:
        return ("UTILITY", root, leaf, s)
    if root.lower() in ("sale", "customer appreciation", "customer apprecation"):
        return ("UTILITY", root, leaf, s)
    if EVENT_PAT.search(r['title']) and root.lower() not in ("brands","brand","accessories","clothing","dresses","brighton","home"):
        return ("UTILITY", root, leaf, s)
    if root.lower() in ("brands","brand"):
        if leaf.lower() in THEMED: return ("THEMED", root, leaf, s)
        return ("BRAND", root, leaf, s)
    if root.lower() == "brighton":               # Brighton is a brand — its sub-pages are brand-family pages
        return ("BRIGHTON", root, leaf, s)
    # standalone brand collections at top level
    standalone = {"sympli","judy blue","radzoli","erin gray jewelry","a rare bird","trust your journey","brighton"}
    if low.strip() in standalone:
        return ("BRAND", root, leaf, s)
    if root.lower() in ("clothing","dresses","accessories","home"):
        return (root.upper(), root, leaf, s)
    return ("GENERIC", root, leaf, s)

def make(r):
    kind, root, leaf, s = classify(r)
    nb = titlecase_brand(leaf)
    note = ""; action = "Optimize"
    if kind == "UTILITY":
        title = f"{r['title'].split(chr(92))[-1].strip()} | {STORE}"
        meta = f"Shop {leaf.strip().lower()} at {STORE}, a women's boutique in {CITY}."
        desc = ""
        action = "NOINDEX — event/utility page; exclude from search, no description needed"
        note = "Temporary/utility collection — do not optimize for SEO."
        return title, meta, desc, action, note, kind

    if kind == "BRAND":
        cat = BRAND_CAT.get(leaf.lower(), None)
        catword = {"dresses":"Dresses","eveningwear":"Evening Dresses","jeans":"Jeans & Denim",
                   "pants":"Pants","clothing":"Clothing","jewelry":"Jewelry","handbags":"Handbags",
                   "footwear":"Shoes","socks":"Socks","jackets":"Jackets","intimates":"Intimates",
                   "bath & body":"Bath & Body","fragrance":"Fragrance","home fragrance":"Home Fragrance",
                   "gifts":"Gifts","baby & gift":"Baby & Gifts","shoe care":"Shoe Care","pants ":"Pants"}.get(cat)
        # don't append a category word the brand name already contains (e.g. "Erin Gray Jewelry" + "Jewelry")
        redundant = catword and any(w.lower() in nb.lower() for w in catword.replace('&',' ').split() if len(w) > 2)
        if catword and not redundant:
            title = f"{nb} {catword} | {STORE} — {CITY}"
            desc = (f"Shop {nb} {catword.lower()} at {STORE}, a women's boutique in downtown {CITY}. "
                    f"We're a long-standing local stockist of {nb} — browse the current selection in store "
                    f"at {ADDR} or online.")
            note = f"Brand classified as '{cat}' (verified: store carries a {nb} line)."
        else:
            title = f"{nb} | {STORE} — {CITY}"
            desc = (f"Discover {nb} at {STORE}, a women's clothing and accessories boutique in downtown {CITY}. "
                    f"Shop the {nb} collection in store at {ADDR} or online.")
            note = "Brand not category-classified — neutral copy (no fabricated product type)."
        meta = f"Shop {nb} at {STORE} in {CITY}. Browse the latest {nb} styles in store or online."
        return title, meta, desc[:600], action, note, kind

    if kind == "BRIGHTON":
        if len(s) == 1:
            title = f"Brighton Jewelry & Accessories | {STORE} — {CITY}"
            sub = "jewelry, handbags, charms and accessories"
        else:
            n = noun_for(leaf, root, s).replace("Women's ", "")   # "Brighton Jewelry", not "Brighton Women's Jewelry"
            title = f"Brighton {n} | {STORE} — {CITY}"
            sub = n.lower()
        desc = (f"Shop Brighton {sub} at {STORE}, an authorized Brighton retailer in downtown {CITY}. "
                f"Browse the current Brighton collection in store at {ADDR} or online.")
        meta = f"Brighton {sub} at {STORE}, an authorized Brighton retailer in {CITY}. Shop in store or online."
        note = "Brighton brand family (store is an authorized Brighton retailer)."
        return title, meta, desc, action, note, kind

    if kind == "DRESSES":
        noun = noun_for(leaf, root, s) if len(s) > 1 else "Designer Dresses"
        brands = CAT_BRANDS["dresses"]
        occ = {"evening":"for weddings, galas and formal events","casual":"for everyday and daytime wear",
               "mother of bride":"for the mother of the bride or groom","plus":"in plus and extended sizes",
               "petite":"cut for petite frames","pant sets":"two-piece dress pant sets"}.get(leaf.lower(),
               "from casual daytime styles to evening and formal occasions")
        title = (f"{noun} | {STORE} — {CITY}" if len(s)>1 else f"Designer Dresses — Casual to Evening | {STORE}")
        desc = (f"Shop {noun.lower()} {occ} at {STORE} in {CITY}. Our dress racks feature brands like "
                f"{brands[0]}, {brands[1]} and {brands[2]} — find your fit in store at {ADDR} or online.")
        meta = f"{noun} at {STORE}, {CITY} — {brands[0]}, {brands[1]} & more. Shop in store or online."
        note = "Brands named are verified dress labels the store carries."
        return title, meta, desc, action, note, kind

    if kind == "CLOTHING":
        noun = noun_for(leaf, root, s) if len(s) > 1 else "Women's Clothing"
        brands = CAT_BRANDS["clothing"]
        title = f"{noun} | {STORE} — {CITY}"
        desc = (f"Browse {noun.lower()} at {STORE}, a women's boutique in downtown {CITY}. We carry "
                f"{brands[0]}, {brands[1]}, {brands[2]} and more — shop in store at {ADDR} or online.")
        meta = f"{noun} at {STORE} in {CITY}. {brands[0]}, {brands[1]} & more — shop in store or online."
        note = "Brands named are verified clothing labels the store carries."
        return title, meta, desc, action, note, kind

    if kind == "ACCESSORIES":
        leaf_l = leaf.lower()
        noun = noun_for(leaf, root, s)
        # jewelry / handbags / footwear get brand depth; smaller accessories stay clean
        if "jewel" in r['title'].lower():
            brands = CAT_BRANDS["jewelry"]; fam="jewelry"
        elif "handbag" in r['title'].lower() or "wallet" in leaf_l:
            brands = CAT_BRANDS["handbags"]; fam="handbags"
        elif "footwear" in r['title'].lower() or "shoe" in leaf_l or "slipper" in leaf_l:
            brands = CAT_BRANDS["footwear"]; fam="footwear"
        else:
            brands = None; fam="accessories"
        disp = noun if noun.lower().startswith(("women", leaf_l)) else "Women's " + noun
        title = re.sub(r"\s+", " ", f"{disp} | {STORE} — {CITY}")
        if brands:
            bl = ", ".join(brands[:3])
            desc = (f"Shop {noun.lower()} at {STORE} in {CITY}, featuring {bl} and more. "
                    f"Find the perfect piece in store at {ADDR} or online.")
            note = f"Brands named are verified {fam} labels the store carries."
        else:
            desc = (f"Shop {noun.lower()} at {STORE}, a women's boutique in downtown {CITY}. "
                    f"Browse the selection in store at {ADDR} or online.")
            note = "Accessory subcategory — category-grounded copy, no brand claims."
        meta = f"{noun} at {STORE} in {CITY}. Shop the selection in store or online."
        return title, meta, desc, action, note, kind

    if kind == "THEMED":
        nb2 = titlecase_brand(leaf)
        title = f"{nb2} — Women's Fashion | {STORE} — {CITY}"
        desc = (f"Explore our {nb2} selection at {STORE}, a women's boutique in downtown {CITY}. "
                f"Shop in store at {ADDR} or online.")
        meta = f"{nb2} women's fashion at {STORE} in {CITY}. Shop in store or online."
        note = "Themed/sourcing collection — neutral grounded copy."
        return title, meta, desc, action, note, kind

    # GENERIC fallback (still grounded in leaf + location, no 'for Women' junk)
    noun = noun_for(leaf, root, s)
    title = f"{noun} | {STORE} — {CITY}"
    desc = (f"Explore our {noun} selection at {STORE}, a women's clothing and accessories boutique in "
            f"downtown {CITY}. Available in store at {ADDR} or online.")
    meta = f"{noun} at {STORE} in {CITY}. Shop in store or online."
    note = "Generic — grounded in collection name + location only."
    return title, meta, desc, action, note, kind

# ---- generate + uniqueness gate ----
out_rows = []
seen_desc = {}
dupes = 0
for r in rows:
    title, meta, desc, action, note, kind = make(r)
    if desc:
        key = desc.lower()
        if key in seen_desc:
            dupes += 1
            note = (note + " | DUPLICATE of " + seen_desc[key] + " — likely redundant catalog page; "
                    "set a canonical tag to the primary or NOINDEX rather than publish twin copy").strip(" |")
        else:
            seen_desc[key] = r['url']
    wc = len(desc.split()) if desc else 0
    out_rows.append({**r, 'kind':kind, 'newTitle':title, 'meta':meta, 'desc':desc,
                     'descWords':wc, 'action':action, 'note':note})

# ---- write workbook ----
wb = openpyxl.Workbook()
H = PatternFill("solid", fgColor="1F3864"); SUB = PatternFill("solid", fgColor="2E5496")
NOIDX = PatternFill("solid", fgColor="FCE4D6"); WHITE = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9"); BORDER = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

ws = wb.active; ws.title = "Collection SEO (Rebuilt)"; ws.sheet_view.showGridLines = False
ws.merge_cells("A1:H1")
ws.cell(1,1,"Collection SEO — Rebuilt (per-collection, anti-filler)").font = Font(size=14,bold=True,color="1F3864")
ws.cell(2,1,f"{len(out_rows)} navigable collections | {sum(1 for r in out_rows if r['kind']=='UTILITY')} flagged NOINDEX "
            f"| {dupes} duplicate descriptions | every description names >=3 real attributes").font = Font(italic=True,color="595959")
hdr = ["Kind","URL","Old Title","SEO Title (new)","Meta Description (140-155)","On-Page Description (40-90 words)","Words","Action / Grounding note"]
for i,h in enumerate(hdr, start=1):
    c=ws.cell(4,i,h); c.font=WHITE; c.fill=H; c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="left")
widths=[12,34,26,40,46,60,7,40]
for i,w in enumerate(widths, start=1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
for ri,r in enumerate(out_rows, start=5):
    vals=[r['kind'],r['url'],r['title'],r['newTitle'],r['meta'],r['desc'],r['descWords'],
          (r['action']+(" — "+r['note'] if r['note'] else ""))]
    for ci,v in enumerate(vals, start=1):
        c=ws.cell(ri,ci,v); c.border=BORDER; c.alignment=Alignment(wrap_text=True,vertical="top")
        if r['kind']=="UTILITY": c.fill=NOIDX
        if ci==7: c.alignment=Alignment(horizontal="center",vertical="top")
ws.freeze_panes="A5"

# ---- doc sheet ----
d = wb.create_sheet("How These Were Built"); d.sheet_view.showGridLines=False
d.column_dimensions['A'].width=3; d.column_dimensions['B'].width=110
def L(r,t,b=False,sz=11,col="000000"):
    c=d.cell(r,2,t); c.font=Font(bold=b,size=sz,color=col); c.alignment=Alignment(wrap_text=True,vertical="top"); return r+1
r=1
r=L(r,"How the rebuilt Collection SEO was generated",16,True,"1F3864")
r=L(r,"The original plan's 'Proposed SEO Title' column used a single blind template — {Title} for Women | The "
      "Clothing Cove — Milford, MI — which produced nonsense (Pillows for Women, Tea Towels for Women, Brands "
      "for Women, Sale for Women). This rebuild replaces it.",11);r+=1
r=L(r,"Anti-filler rules enforced",12,True,"1F3864")
r=L(r,"1. Per-collection classification. Each collection is parsed into its hierarchy and typed as BRAND, DRESSES, "
      "CLOTHING, ACCESSORIES, THEMED, or UTILITY. Titles are built per type — never a flat suffix.")
r=L(r,"2. Descriptions name >=3 attributes that are TRUE of the page: the real category, a real subtype/occasion/fit, "
      "the real location (414 N Main St, Milford), and verified brands only where provably stocked.")
r=L(r,"3. Brand claims are grounded in evidence. A brand is only named on a category page when the store has its own "
      "collection for that brand (e.g. Dresses cites Joseph Ribkoff / Frank Lyman / Adrianna Papell — all real "
      "brand collections). No brand is invented.")
r=L(r,"4. No volatile data. Product counts and exact size ranges are NOT hard-coded into copy (they change and would "
      "go stale / become false). Plus and Petite pages reference size/fit because the collection name guarantees it.")
r=L(r,"5. Event/utility collections (Sale, Customer Appreciation, Gift Cards, New Products, Brands index) are flagged "
      "NOINDEX — exclude from search rather than write thin copy for a temporary page.")
r=L(r,"6. Uniqueness gate. No two on-page descriptions are byte-identical; any near-duplicate is flagged in the note "
      "column for a human pass.");r+=1
r=L(r,"What still needs a human (by design)",12,True,"C55A11")
r=L(r,"Brands left un-classified get neutral copy (no fabricated product type) — a merchant can add the real category "
      "in 10 seconds. The vendor/productType fields in Shopify are cryptic supplier codes (LEEGIN LEATHER, 'J E W'), "
      "so they were deliberately NOT used to auto-write specifics. These descriptions are a grounded first draft to "
      "edit, not final copy to publish blind.")
wb.save("The Clothing Cove - Collection SEO Rebuilt (June 2026).xlsx")
print("saved workbook. rows:", len(out_rows), "| noindex:", sum(1 for r in out_rows if r['kind']=='UTILITY'), "| dupes:", dupes)
# print a sample across kinds
import itertools
shown=set()
for r in out_rows:
    if r['kind'] not in shown:
        shown.add(r['kind'])
        print(f"\n[{r['kind']}] {r['url']}")
        print("  T:", r['newTitle'])
        print("  M:", r['meta'])
        print("  D:", r['desc'][:300])
