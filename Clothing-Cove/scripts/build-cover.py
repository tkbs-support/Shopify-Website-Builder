"""Prepend a one-page 'Proposal Summary' cover to the Projected Impact workbook
(the lead document): the problem, the three workstreams, the headline ROI, and a
consolidated timeline + investment. Charts in the existing sheets are preserved.
"""
import openpyxl, json, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
RATE=125
D=os.path.join(os.path.dirname(__file__),"..","data")
fs=json.load(open(os.path.join(D,"filter-summary.json"))); fab=json.load(open(os.path.join(D,"fabric-summary.json")))
sb=fs["size_buckets"]; cb=fs["color_buckets"]
def H(n,m): return round(n*m/60)
# Filters Phase 1-2, conventional hands-on effort
# 16-26 = normalize Brand (supplier codes -> brands) + Garment type (3-letter codes -> readable) + enable; +4/8 = occasion
f12_lo = 16 + H(sb.get("OPT_ONLY",0)+sb.get("MF_ONLY",0),1.5) + H(cb.get("OPT_ONLY",0),1.5) + H(fab["proposal_rows"],2.5) + 4
f12_hi = 26 + H(sb.get("OPT_ONLY",0)+sb.get("MF_ONLY",0),2.5) + H(cb.get("OPT_ONLY",0),2.5) + H(fab["proposal_rows"],3.5) + 8
SEO_LO,SEO_HI = 40,70           # SEO foundation (priority collections + technical + homepage)
FND_LO,FND_HI = f12_lo+SEO_LO, f12_hi+SEO_HI
F="The Clothing Cove - Projected Impact.xlsx"
wb=openpyxl.load_workbook(F)
NAVY=PatternFill("solid",fgColor="1F3864"); BLUE=PatternFill("solid",fgColor="2E5496")
BAND=PatternFill("solid",fgColor="FFF2CC"); GREEN=PatternFill("solid",fgColor="C6EFCE")
WHITE=Font(color="FFFFFF",bold=True); TH=Side(style="thin",color="D9D9D9"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)

if "Proposal Summary" in wb.sheetnames: del wb["Proposal Summary"]
ws=wb.create_sheet("Proposal Summary",0); ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=2
for c,w in zip("BCDEFG",[30,16,16,16,16,24]): ws.column_dimensions[c].width=w
def line(r,t,sz=11,b=False,col="000000",merge=7,fill=None):
    c=ws.cell(r,2,t); c.font=Font(size=sz,bold=b,color=col); c.alignment=Alignment(wrap_text=True,vertical="top")
    if fill:c.fill=fill
    if merge: ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=merge)
    return r+1
r=2
r=line(r,"The Clothing Cove — Growth Proposal",20,True,"1F3864")
r=line(r,"SEO & AEO  ·  Filter & Search Readiness  ·  Projected Impact   |   prepared June 2026",11,False,"595959")
r+=1
r=line(r,"THE OPPORTUNITY",13,True,"1F3864")
r=line(r,"The store earns ~3,000 organic visits/month but is leaving growth on the table: the category and brand pages "
        "that should rank have no description text at all (78% of collections), so they under-perform in search and give "
        "AI nothing to read; only 2 of 6 expected filters are live; and product data isn't normalized for filtering. "
        "Sales have softened for years. The catalog is strong — discoverability and the on-site experience haven't kept up.")
r+=1
r=line(r,"THE PLAN — THREE WORKSTREAMS",13,True,"1F3864")
hdr=["Workstream","What it delivers","","","","Outcome"]
for i,h in enumerate(["Workstream","What it delivers","Outcome"],start=2):
    cc=ws.cell(r,i if i<4 else 7,h)  # simple 3-col header
for i,(c0,c1) in enumerate([(2,"Workstream"),(3,"What it delivers"),(7,"Outcome")]):
    cell=ws.cell(r,c0,c1); cell.font=WHITE; cell.fill=BLUE
ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
r+=1
WS=[("SEO & AEO","Write the missing content on the category and brand pages that actually drive rankings — most have no "
       "description text at all today — and make the catalog legible to AI search (schema + answer-readiness) so the store "
       "gets surfaced and cited. Product title/description cleanup is included as catalog hygiene.","Qualified traffic + AI visibility"),
    ("Filter & Search","Make the live catalog filterable by the 6 core filters shoppers expect — Size, Color, Brand, "
       "Garment type (plus Price/Availability) — with clean, normalized values.","Visitors find and buy faster"),
    ("Projected Impact","A measured, conservative model of the traffic and revenue lift, with monthly tracking so results "
       "are provable, not promised.","Confidence in the ROI")]
for w0,w1,w2 in WS:
    ws.cell(r,2,w0).font=Font(bold=True,color="1F3864"); ws.cell(r,2).border=BD; ws.cell(r,2).alignment=Alignment(wrap_text=True,vertical="top")
    c=ws.cell(r,3,w1); c.alignment=Alignment(wrap_text=True,vertical="top"); c.border=BD
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
    ws.cell(r,7,w2).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(r,7).border=BD
    r+=1
r+=1
r=line(r,"HEADLINE IMPACT  (12 months, from 3,000 organic visits/mo)",13,True,"1F3864")
r=line(r,"Expected case: organic traffic ≈ +12%  ·  revenue-per-visit ≈ +8%  ·  combined revenue ≈ +21%. "
        "Range: +8% (conservative — for a store that's been declining, mostly stabilization) to +51% (strong). The "
        "conversion (filter) lever is the more controllable half. Revenue is an INDEX (no Orders API) — multiply by "
        "monthly revenue to dollarize. ILLUSTRATIVE, not a promise: with no Search Console/GA4 baseline yet, Phase 0 sets "
        "the real targets and re-calibrates this model.",11,True,"C55A11",fill=BAND)
r+=1
r=line(r,"WHAT IT TAKES TO GET THIS DONE",13,True,"1F3864")
r=line(r,"Foundation scope: all six core filters working + the SEO base. Three ways to deliver it — the trade-offs "
        "are time and cost.",10,False,"808080")
# in-house model: conventional effort, slower (learning curve), squeezed part-time
RED=PatternFill("solid",fgColor="FFC7CE")
DATA_MULT,SEO_MULT=2,4   # in-house slowdown vs a specialist: data entry 2x, SEO writing 4x
IH_LO=round(f12_lo*DATA_MULT + SEO_LO*SEO_MULT); IH_HI=round(f12_hi*DATA_MULT + SEO_HI*SEO_MULT)
ih_mo_lo=round(IH_LO/8/4.3); ih_mo_hi=round(IH_HI/8/4.3)    # ~8 usable hrs/week
TKBS_FEE="$14,000"
for i,h in enumerate(["Path","Time to complete","Cost","Trade-off"],start=2):
    c=ws.cell(r,i,h); c.font=WHITE; c.fill=BLUE; c.alignment=Alignment(horizontal="center",wrap_text=True)
    if i==5: ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=7)
r+=1
COMP=[("In-house (your team)",f"~{ih_mo_lo}–{ih_mo_hi} months (part-time)",f"~${IH_LO*22:,}–{IH_HI*22:,} in staff wages",
        "Pulls staff off the sales floor for a year+; historically stalls before it's finished",RED),
      ("Outside agency","~2–4 months",f"${FND_LO*RATE:,}–{FND_HI*RATE:,}",
        "Specialist rates for the same hands-on, per-item work",BAND),
      ("TKBS (us)","4 weeks",f"{TKBS_FEE}",
        "Same outcome — faster than in-house, a fraction of agency cost, and your team stays selling",GREEN)]
for path,time,cost,note,fill in COMP:
    ws.cell(r,2,path).border=BD; ws.cell(r,2).fill=fill; ws.cell(r,2).font=Font(bold=True)
    for i,v in zip((3,4),[time,cost]):
        c=ws.cell(r,i,v); c.border=BD; c.fill=fill; c.alignment=Alignment(horizontal="center",wrap_text=True,vertical="top")
    c=ws.cell(r,5,note); c.border=BD; c.fill=fill; c.alignment=Alignment(wrap_text=True,vertical="top")
    ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=7)
    r+=1
r=line(r,"The in-house and agency columns are what the same work costs by conventional means; TKBS delivers that scope "
        "in weeks without tying up your team.",9,False,"808080")
r+=1
r=line(r,"WHAT THE $14,000 FOUNDATION COVERS",12,True,"1F3864")
r=line(r,"INCLUDED: Phase 0 — Search Console + GA4 connected, 16-month baseline + ranking/competitor snapshot;  theme & "
        "homepage SEO — keyworded title, meta, H1, schema/OG backstop;  written descriptions + SEO titles for the "
        "navigable category & brand pages (the real content gap);  indexation hygiene — noindex/prune thin, empty & "
        "duplicate collections;  the four core filters (Size, Color, Brand, Garment type) wired into Search & Discovery "
        "with normalized values and clean text color display (Availability + Price already on);  and monthly tracking "
        "that measures the lift.",10,fill=BAND,merge=7)
r=line(r,"NOT INCLUDED (scoped & quoted separately, per batch): full-catalog product enrichment (custom titles, "
        "descriptions, tags & alt-text across the long tail — hygiene-level);  fabric & occasion enrichment filters;  "
        "reviews-app setup;  and publishing/photographing the unpublished backlog.",10,False,"808080",merge=7)
r+=1
r=line(r,"Recommendation: start here — it delivers the diagnostics, the high-value category content, and all six core "
        "filters in ~4 weeks, with tracking that proves the lift before any full-catalog enrichment.",11,True,"1F3864")

wb.move_sheet("Proposal Summary", -(wb.sheetnames.index("Proposal Summary")))  # ensure first
wb.save(F)
print("cover added to", F, "| sheets:", wb.sheetnames)
