import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open(r"a:\TKBS Marketing - Git\Shopify-Website-Builder\collection-analysis.json") as f:
    collections = json.load(f)

def categorize(handle):
    top_level = ["clothing", "dresses", "brighton", "accessories", "new-arrivals", "sale",
                 "shop-with-purpose", "brands", "sympli"]
    if handle in top_level:
        return "Top-Level"

    subcats = ["clothing-tops", "clothing-bottoms", "clothing-denim", "clothing-jumpsuits",
               "dresses-evening", "dresses-casual", "dresses-mother-of-bride", "dresses-pant-sets",
               "dresses-plus", "dresses-petite",
               "brighton-jewelry", "brighton-handbags", "brighton-collections", "brighton-charms",
               "brighton-accessories", "brighton-eyewear",
               "accessories-jewelry", "accessories-handbags", "accessories-footwear", "accessories-accessories",
               "shop-with-purpose-clothing", "shop-with-purpose-accessories", "shop-with-purpose-home",
               "shop-with-purpose-brands", "shop-with-purpose-unique-gifts",
               "sale-accessories", "sale-bottoms", "sale-dresses", "sale-handbags", "sale-tops",
               "sympli-best-sellers", "sympli-bottoms", "sympli-dresses", "sympli-layering",
               "sympli-sale", "sympli-tops"]
    if handle in subcats:
        return "Subcategory"

    if handle.startswith("brands-"):
        return "Brand"

    depth = handle.count("-")
    if depth >= 3:
        return "Deep Sub"
    return "Sub-Sub"

wb = Workbook()
ws = wb.active
ws.title = "All Navigable Collections"

hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
hfill = PatternFill("solid", fgColor="1B2838")
halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
bfont = Font(name="Arial", size=10)
border = Border(
    left=Side(style="thin", color="D9E0E6"), right=Side(style="thin", color="D9E0E6"),
    top=Side(style="thin", color="D9E0E6"), bottom=Side(style="thin", color="D9E0E6"),
)

yes_fill = PatternFill("solid", fgColor="C8E6C9")
no_fill = PatternFill("solid", fgColor="FFCDD2")
na_fill = PatternFill("solid", fgColor="E0E0E0")

cat_fills = {
    "Top-Level": PatternFill("solid", fgColor="E8F5E9"),
    "Subcategory": PatternFill("solid", fgColor="FFF8E1"),
    "Brand": PatternFill("solid", fgColor="E3F2FD"),
    "Sub-Sub": PatternFill("solid", fgColor="F3E5F5"),
    "Deep Sub": PatternFill("solid", fgColor="FFF3E0"),
}

headers = ["Category", "Collection URL", "Current Title", "Has SEO Text", "Has Image",
           "Text Preview", "Proposed SEO Title", "Needs Work", "Status"]
widths = [13, 50, 40, 13, 12, 55, 62, 12, 12]

for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = halign
    cell.border = border
    ws.column_dimensions[chr(64 + col) if col <= 26 else ""].width = w

ws.column_dimensions["A"].width = 13
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 13
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 55
ws.column_dimensions["G"].width = 62
ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12

sorted_collections = sorted(collections, key=lambda c: (
    {"Top-Level": 0, "Subcategory": 1, "Brand": 2, "Sub-Sub": 3, "Deep Sub": 4}[categorize(c["handle"])],
    c["handle"]
))

proposed_titles = {
    "clothing": "Women's Clothing | The Clothing Cove - Milford, MI",
    "dresses": "Designer Dresses - Casual to Evening | The Clothing Cove",
    "brighton": "Brighton Jewelry & Accessories | The Clothing Cove",
    "accessories": "Women's Accessories - Jewelry, Handbags & More | The Clothing Cove",
    "new-arrivals": "New Arrivals - Latest Women's Fashion | The Clothing Cove",
    "sale": "Sale - Women's Clothing & Accessories | The Clothing Cove",
    "shop-with-purpose": "Shop With Purpose - Ethically Made Fashion | The Clothing Cove",
    "brands": "Designer Brands | The Clothing Cove - Milford, MI",
    "sympli": "Sympli Clothing | The Clothing Cove - Milford, MI",
    "clothing-tops": "Women's Tops - Casual, Dressy & Layering | The Clothing Cove",
    "clothing-bottoms": "Women's Pants, Skirts & Bottoms | The Clothing Cove",
    "clothing-denim": "Women's Denim - Jeans, Jackets & Jeggings | The Clothing Cove",
    "clothing-jumpsuits": "Women's Jumpsuits | The Clothing Cove",
    "dresses-evening": "Evening & Formal Dresses | The Clothing Cove",
    "dresses-casual": "Casual Dresses | The Clothing Cove",
    "dresses-mother-of-bride": "Mother of the Bride Dresses | The Clothing Cove",
    "dresses-pant-sets": "Dress Pant Sets | The Clothing Cove",
    "dresses-plus": "Plus Size Dresses | The Clothing Cove",
    "dresses-petite": "Petite Dresses | The Clothing Cove",
    "brighton-jewelry": "Brighton Jewelry - Necklaces, Bracelets & Rings | The Clothing Cove",
    "brighton-handbags": "Brighton Handbags & Wallets | The Clothing Cove",
    "brighton-collections": "Brighton Collections | The Clothing Cove",
    "brighton-charms": "Brighton Charms & Beads | The Clothing Cove",
    "brighton-accessories": "Brighton Accessories | The Clothing Cove",
    "brighton-eyewear": "Brighton Eyewear - Sunglasses & Readers | The Clothing Cove",
    "accessories-jewelry": "Women's Jewelry - Fashion & Designer Pieces | The Clothing Cove",
    "accessories-handbags": "Women's Handbags & Purses | The Clothing Cove",
    "accessories-footwear": "Women's Shoes & Footwear | The Clothing Cove",
    "accessories-accessories": "Women's Fashion Accessories | The Clothing Cove",
    "shop-with-purpose-clothing": "Ethically Made Women's Clothing | The Clothing Cove",
    "shop-with-purpose-accessories": "Ethically Made Accessories | The Clothing Cove",
    "shop-with-purpose-home": "Ethically Made Home Goods | The Clothing Cove",
    "shop-with-purpose-brands": "Ethical & Fair Trade Brands | The Clothing Cove",
    "shop-with-purpose-unique-gifts": "Unique Gifts - Shop With Purpose | The Clothing Cove",
    "sale-accessories": "Accessories on Sale | The Clothing Cove",
    "sale-bottoms": "Women's Bottoms on Sale | The Clothing Cove",
    "sale-dresses": "Dresses on Sale | The Clothing Cove",
    "sale-handbags": "Handbags on Sale | The Clothing Cove",
    "sale-tops": "Women's Tops on Sale | The Clothing Cove",
    "sympli-best-sellers": "Sympli Best Sellers | The Clothing Cove",
    "sympli-bottoms": "Sympli Bottoms & Pants | The Clothing Cove",
    "sympli-dresses": "Sympli Dresses | The Clothing Cove",
    "sympli-layering": "Sympli Layering Pieces | The Clothing Cove",
    "sympli-sale": "Sympli Sale | The Clothing Cove",
    "sympli-tops": "Sympli Tops | The Clothing Cove",
    "brands-sympli": "Sympli Clothing | The Clothing Cove - Milford, MI",
    "brands-brighton-jewelry-and-accessories": "Brighton Jewelry & Accessories | The Clothing Cove",
    "brands-joseph-ribkoff": "Joseph Ribkoff Dresses & Fashion | The Clothing Cove",
    "brands-frank-lyman": "Frank Lyman Dresses & Separates | The Clothing Cove",
    "brands-renuar": "Renuar Perfect Fit Pants & Clothing | The Clothing Cove",
    "brands-liverpool-jeans": "Liverpool Jeans & Denim | The Clothing Cove",
    "brands-rachel-marie-designs": "Rachel Marie Designs Jewelry | The Clothing Cove - Michigan Made",
    "brands-charlie-b": "Charlie B Clothing | The Clothing Cove",
    "brands-tribal": "Tribal Sportswear | The Clothing Cove",
    "brands-democracy": "Democracy Clothing - Ab-Solution Jeans & More | The Clothing Cove",
    "brands-liv-by-habitat": "Liv by Habitat - Sustainable Women's Fashion | The Clothing Cove",
    "brands-damee": "Damee Jackets & Wearable Art | The Clothing Cove",
    "brands-alex-evenings": "Alex Evenings Dresses | The Clothing Cove",
    "brands-adrianna-papell": "Adrianna Papell Dresses | The Clothing Cove",
    "brands-baggallini": "Baggallini Bags & Travel | The Clothing Cove",
    "brands-brighton-meridian": "Brighton Meridian Collection | The Clothing Cove",
    "brands-jag-jeans": "JAG Jeans | The Clothing Cove",
    "brands-jess-jane": "Jess & Jane Tops | The Clothing Cove",
    "brands-michael-tyler": "Michael Tyler Fashion | The Clothing Cove",
    "brands-montage": "Montage Evening Wear | The Clothing Cove",
    "brands-mgny": "MGNY by Mori Lee Gowns | The Clothing Cove",
    "brands-slim-sation": "Slim-Sation Pants | The Clothing Cove",
    "brands-spring-step-shoes": "Spring Step Shoes | The Clothing Cove",
    "brands-tagua-jewelry": "Tagua Eco-Friendly Jewelry | The Clothing Cove",
}

for row_idx, c in enumerate(sorted_collections, 2):
    cat = categorize(c["handle"])
    has_text = "Yes" if c["has_text"] else "No"
    has_image = "Yes" if c["has_image"] else "No"
    needs_work = "No" if c["has_text"] else "Yes"
    preview = c["text_preview"] if c["has_text"] else ""
    proposed = proposed_titles.get(c["handle"], "")

    row_data = [cat, f"/{c['handle']}", c["title"], has_text, has_image, preview, proposed, needs_work, "Pending"]
    row_fill = cat_fills.get(cat, PatternFill())

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = bfont
        cell.border = border
        cell.fill = row_fill

        if col == 4:
            cell.fill = yes_fill if val == "Yes" else no_fill
        elif col == 5:
            cell.fill = yes_fill if val == "Yes" else no_fill
        elif col == 8:
            cell.fill = no_fill if val == "Yes" else yes_fill

ws.auto_filter.ref = f"A1:I{len(sorted_collections)+1}"
ws.freeze_panes = "A2"

# --- Summary sheet ---
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 45
ws2.column_dimensions["B"].width = 20

cats = {}
for c in sorted_collections:
    cat = categorize(c["handle"])
    if cat not in cats:
        cats[cat] = {"total": 0, "has_text": 0, "has_image": 0, "needs_work": 0}
    cats[cat]["total"] += 1
    if c["has_text"]:
        cats[cat]["has_text"] += 1
    if c["has_image"]:
        cats[cat]["has_image"] += 1
    if not c["has_text"]:
        cats[cat]["needs_work"] += 1

summary = [
    ("Collection SEO Audit -The Clothing Cove", "", True, 14),
    ("", "", False, 11),
    ("Total Navigable Collections", len(sorted_collections), True, 12),
    ("With SEO Text Description", sum(1 for c in collections if c["has_text"]), False, 11),
    ("With Image Only (no text)", sum(1 for c in collections if c["has_image"] and not c["has_text"]), False, 11),
    ("No Description At All", sum(1 for c in collections if not c["has_text"] and not c["has_image"]), False, 11),
    ("Collections Needing Work", sum(1 for c in collections if not c["has_text"]), True, 12),
    ("", "", False, 11),
    ("Breakdown by Category:", "", True, 12),
]

for cat_name in ["Top-Level", "Subcategory", "Brand", "Sub-Sub", "Deep Sub"]:
    if cat_name in cats:
        d = cats[cat_name]
        summary.append((f"  {cat_name} -{d['total']} total, {d['needs_work']} need work", f"{d['has_text']} have text", False, 11))

summary += [
    ("", "", False, 11),
    ("What We Fix Per Collection:", "", True, 12),
    ("1. SEO Title", "Keyword-rich, 50-65 chars", False, 11),
    ("2. Meta Description", "140-155 chars with location & CTA", False, 11),
    ("3. Collection Description", "2-3 sentences with product attributes", False, 11),
    ("", "", False, 11),
    ("IMPORTANT:", "", True, 12),
    ("Collection changes are store-wide", "NOT theme-specific", False, 11),
    ("Changes go live immediately", "Owner approval required", False, 11),
    ("", "", False, 11),
    ("Color Key:", "", True, 12),
    ("  Green rows = Top-Level Categories", "Highest priority", False, 11),
    ("  Yellow rows = Subcategories", "High priority", False, 11),
    ("  Blue rows = Brand Pages", "Medium priority", False, 11),
    ("  Purple rows = Sub-Subcategories", "Lower priority", False, 11),
    ("  Orange rows = Deep Subcategories", "Lowest priority", False, 11),
]

for row_idx, (label, val, bold, size) in enumerate(summary, 1):
    c1 = ws2.cell(row=row_idx, column=1, value=label)
    c2 = ws2.cell(row=row_idx, column=2, value=val)
    c1.font = Font(name="Arial", size=size, bold=bold)
    c2.font = Font(name="Arial", size=11)
    if "IMPORTANT" in label:
        c1.font = Font(name="Arial", size=12, bold=True, color="B42318")

out = r"a:\TKBS Marketing - Git\Shopify-Website-Builder\The Clothing Cove - Collection SEO Plan.xlsx"
wb.save(out)

print(f"Saved: {out}")
print(f"Total rows: {len(sorted_collections)}")
for cat_name in ["Top-Level", "Subcategory", "Brand", "Sub-Sub", "Deep Sub"]:
    if cat_name in cats:
        d = cats[cat_name]
        print(f"  {cat_name}: {d['total']} total, {d['needs_work']} need work, {d['has_text']} have text")
