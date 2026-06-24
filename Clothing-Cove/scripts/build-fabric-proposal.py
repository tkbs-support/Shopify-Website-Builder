"""Fabric backfill PROPOSAL (Phase 2). Extracts fabric keywords from product descriptions,
maps them to shopify.fabric taxonomy metaobjects using a mapping LEARNED from products your
team already tagged (single-fabric cases). Text extraction => review-first, not blind import.
Outputs workbook + data/fabric-summary.json.
"""
import json, os, sys, re
from collections import Counter, defaultdict
DATA=os.path.join(os.path.dirname(__file__),"..","data")

# canonical fabric vocabulary (variant -> canonical)
CANON={"elastane":"spandex","lycra":"spandex","spandex":"spandex","viscose":"viscose","rayon":"rayon",
       "cotton":"cotton","polyester":"polyester","nylon":"nylon","linen":"linen","silk":"silk","wool":"wool",
       "cashmere":"cashmere","modal":"modal","bamboo":"bamboo","denim":"denim","leather":"leather",
       "acrylic":"acrylic","lyocell":"tencel","tencel":"tencel","polyamide":"nylon"}
FAB=re.compile(r'\b('+'|'.join(sorted(CANON,key=len,reverse=True))+r')\b', re.I)
TAG=re.compile(r'<[^>]+>')
def fabrics_in(desc):
    txt=TAG.sub(' ', desc or '')
    found=[CANON[m.lower()] for m in FAB.findall(txt)]
    return list(dict.fromkeys(found))      # unique, preserve order
def mfvals(mf):
    if not mf: return []
    v=mf.get("value") if isinstance(mf,dict) else mf
    try: return json.loads(v)
    except Exception: return [v] if v else []

# join: descriptions from products.jsonl, fabric metafield from size-color.jsonl
desc_by={}
for line in open(os.path.join(DATA,"products.jsonl"),encoding="utf-8"):
    p=json.loads(line)
    if p.get("handle"): desc_by[p["handle"]]=p.get("descriptionHtml") or ""
sc=[json.loads(l) for l in open(os.path.join(DATA,"size-color.jsonl"),encoding="utf-8")]
instock=[p for p in sc if p.get("status")=="ACTIVE" and (p.get("totalInventory") or 0)>0]
APPAREL_TYPES={"D R C","D R E","S O M","S O C","T O P","T O C","T O B","P A M","P A J","K N W",
               "J A K","S K S","T A N","L E G","K N C","P A C","P A P","T O L","S L P","B R A"}
apparel=[p for p in instock if (p.get("productType") or "").strip() in APPAREL_TYPES]

# ---- learn fabric keyword -> GID from products already tagged (single fabric <-> single GID) ----
votes=defaultdict(Counter)
for p in apparel:
    gids=list(dict.fromkeys(mfvals(p.get("fabric"))))
    fabs=fabrics_in(desc_by.get(p["handle"],""))
    if len(gids)==1 and len(fabs)==1:
        votes[fabs[0]][gids[0]]+=1
fab_map={}; fab_amb={}
for fab,c in votes.items():
    top,cnt=c.most_common(1)[0]; tot=sum(c.values())
    if cnt>=2 and cnt/tot>=0.70: fab_map[fab]=(top,cnt,tot)
    else: fab_amb[fab]=(c.most_common(3),tot)
print("learned fabrics:", {f:(g,n) for f,(g,n,t) in fab_map.items()})
print("ambiguous fabrics:", {f:v[0] for f,v in fab_amb.items()})

# label GIDs by the canonical fabric(s) that map to them (no read_metaobjects scope)
gid_label={g:f for f,(g,_,_) in fab_map.items()}

# ---- apply: apparel WITH a fabric in description but NO fabric metafield ----
imp=[]; review=Counter(); have_mf=0; no_keyword=0
for p in apparel:
    if mfvals(p.get("fabric")): have_mf+=1; continue
    fabs=fabrics_in(desc_by.get(p["handle"],""))
    if not fabs: no_keyword+=1; continue
    mapped=[fab_map[f][0] for f in fabs if f in fab_map]
    unmapped=[f for f in fabs if f not in fab_map]
    for u in unmapped: review[u]+=1
    if mapped:
        gids=list(dict.fromkeys(mapped))
        imp.append({"id":p["id"],"handle":p["handle"],"title":p["title"],
                    "fabs":", ".join(fabs),"gids":gids,
                    "names":", ".join(gid_label.get(g,g) for g in gids),
                    "unmapped":", ".join(unmapped)})
print(f"apparel in-stock: {len(apparel)} | already tagged: {have_mf} | no fabric keyword in desc: {no_keyword}")
print(f"fabric PROPOSAL rows: {len(imp)}")

summary={"apparel_instock":len(apparel),"already_tagged":have_mf,"no_keyword":no_keyword,
         "proposal_rows":len(imp),"learned":list(fab_map.keys()),"review":dict(review.most_common())}
json.dump(summary, open(os.path.join(DATA,"fabric-summary.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)

# ---- workbook ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
H=PatternFill("solid",fgColor="1F3864"); SUB=PatternFill("solid",fgColor="2E5496")
WHITE=Font(color="FFFFFF",bold=True); TH=Side(style="thin",color="D9D9D9"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
wb=openpyxl.Workbook()
rm=wb.active; rm.title="README"; rm.sheet_view.showGridLines=False
rm.column_dimensions['A'].width=2; rm.column_dimensions['B'].width=108
def L(r,t,b=False,sz=11,col="000000"):
    c=rm.cell(r,2,t); c.font=Font(bold=b,size=sz,color=col); c.alignment=Alignment(wrap_text=True,vertical="top"); return r+1
r=1
r=L(r,"Fabric backfill — PROPOSAL (review before import)",16,True,"1F3864")
r=L(r,f"Proposes shopify.fabric for {len(imp)} in-stock apparel products whose DESCRIPTION names a fabric but which "
      f"have no fabric metafield. Values map to taxonomy entries learned from {have_mf} products your team already tagged.",11); r+=1
r=L(r,"Why this is review-first (unlike size/color)",12,True,"C55A11")
r=L(r,"Size/color came from structured variant options. Fabric is parsed from free text, so it carries more risk: "
      "a 'polyester lining' mention can look like the main fabric, and blends list several fibers. Default here = "
      "capture EVERY fabric named (so a 'Cotton' filter catches cotton blends). Spot-check the Proposal tab before importing.")
r=L(r,"The 'Fabric Mapping' tab shows every keyword->taxonomy decision; 'Needs Review' lists fabrics we couldn't map.")
order=["README","Fabric Proposal","Fabric Mapping","Fabric Needs Review"]

ws=wb.create_sheet("Fabric Proposal")
cols=["ID","Handle","Title","Metafield: shopify.fabric [list.metaobject_reference]","(ref) Fabrics found","(ref) Mapped to"]
for i,h in enumerate(cols,start=1):
    c=ws.cell(1,i,h); c.font=WHITE; c.fill=H; c.alignment=Alignment(wrap_text=True)
for ri,row in enumerate(imp,start=2):
    ws.cell(ri,1,row["id"]); ws.cell(ri,2,row["handle"]); ws.cell(ri,3,row["title"][:70])
    ws.cell(ri,4,"\n".join(row["gids"])); ws.cell(ri,5,row["fabs"]); ws.cell(ri,6,row["names"])
    for ci in range(1,7): ws.cell(ri,ci).alignment=Alignment(wrap_text=True,vertical="top")
for col,w in zip("ABCDEF",[26,26,40,46,24,22]): ws.column_dimensions[col].width=w
ws.freeze_panes="A2"

ws=wb.create_sheet("Fabric Mapping")
for i,h in enumerate(["Fabric keyword","Metaobject GID","Learned from # products"],start=1):
    c=ws.cell(1,i,h); c.font=WHITE; c.fill=SUB
for ri,(f,(g,cnt,tot)) in enumerate(sorted(fab_map.items()),start=2):
    ws.cell(ri,1,f).border=BD; ws.cell(ri,2,g).border=BD; ws.cell(ri,3,cnt).border=BD
for col,w in zip("ABC",[20,40,18]): ws.column_dimensions[col].width=w

ws=wb.create_sheet("Fabric Needs Review")
ws.cell(1,1,"Fabrics found in descriptions we could not map confidently — assign a taxonomy value by hand").font=Font(bold=True,color="C55A11")
for i,h in enumerate(["Fabric keyword","# products affected"],start=1):
    c=ws.cell(3,i,h); c.font=WHITE; c.fill=SUB
for ri,(f,n) in enumerate(review.most_common(),start=4):
    ws.cell(ri,1,f).border=BD; ws.cell(ri,2,n).border=BD
for col,w in zip("AB",[20,18]): ws.column_dimensions[col].width=w

wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
out="The Clothing Cove - Fabric Backfill Proposal (June 2026).xlsx"
wb.save(out)
print("saved:", out)
