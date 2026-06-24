"""Corrected size/color filter-readiness analysis (dual standard: option AND metafield).
Replaces the options-only 'missing size/color' metric that mislabeled false negatives,
metafield-only products, and genuine one-size items. In-stock scoped.
Outputs: The Clothing Cove - Filter Readiness CORRECTED (June 2026).xlsx
"""
import json, os, sys, re
from collections import Counter, defaultdict
sys.path.insert(0, os.path.dirname(__file__))
from type_expectations import expectation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

SRC = os.path.join(os.path.dirname(__file__), "..", "data", "size-color.jsonl")

def norm(n): return (n or "").strip().lower()
SIZE_OPT_NAMES = {"size", "shoe size", "ring size"}
ONE_SIZE_CUES = re.compile(r"\b(poncho|ruana|wrap|shawl|cape|kimono|one[ -]?size|free[ -]?size|o/s|adjustable|scarf)\b", re.I)

rows = [json.loads(l) for l in open(SRC, encoding="utf-8")]
instock = [p for p in rows if p.get("status") == "ACTIVE" and (p.get("totalInventory") or 0) > 0]

# per-type option coverage (for empirical expectation fallback on unmapped codes)
by_type = defaultdict(list)
for p in instock: by_type[(p.get("productType") or "").strip()].append(p)
type_cov = {}
for code, ps in by_type.items():
    n = len(ps)
    c = sum(1 for p in ps if any(norm(o["name"]) in ("color","colors") for o in p.get("options") or []))
    s = sum(1 for p in ps if any(norm(o["name"]) in SIZE_OPT_NAMES for o in p.get("options") or []))
    type_cov[code] = (c/n if n else 0, s/n if n else 0)

def has_size_opt(p): return any(norm(o["name"]) in SIZE_OPT_NAMES for o in p.get("options") or [])
def has_size_mf(p):  return bool(p.get("sizeStd") or p.get("shoeSize") or p.get("ringSize") or p.get("accSize"))
def has_color_opt(p):return any(norm(o["name"]) in ("color","colors") for o in p.get("options") or [])
def has_color_mf(p): return bool(p.get("colorPat"))
def only_default(p):
    opts = p.get("options") or []
    return len(opts) == 1 and norm(opts[0]["name"]) == "title"

def classify(opt, mf):
    if opt and mf: return "READY"
    if opt:        return "OPT_ONLY"
    if mf:         return "MF_ONLY"
    return "GAP"

size_rows, color_rows = [], []
size_buckets = Counter(); color_buckets = Counter()
size_scope = color_scope = 0
extras = {"fabric":0,"dressOcc":0,"fit":0,"length":0}; extras_scope = 0

for p in instock:
    code = (p.get("productType") or "").strip()
    cc, sc = type_cov.get(code, (0,0))
    label, cat, ec, es, basis, review = expectation(code, cc, sc)

    if es:
        size_scope += 1
        st = classify(has_size_opt(p), has_size_mf(p))
        sub = ""
        if st == "GAP":
            one = only_default(p) and ONE_SIZE_CUES.search(p.get("title") or "")
            sub = "one_size_likely" if one else "real_gap"
        size_buckets[st if st != "GAP" else f"GAP:{sub}"] += 1
        if st != "READY":
            action = {"OPT_ONLY":"Populate shopify.size metafield (option already present)",
                      "MF_ONLY":"Add a Size variant option (metafield already present)",
                      }.get(st)
            if st == "GAP":
                action = ("Set 'One Size' in shopify.size + add Size variant option" if sub=="one_size_likely"
                          else "Add size run: Size variant option + shopify.size metafield")
            size_rows.append([cat, label, p["handle"], p["title"][:60],
                              "Y" if has_size_opt(p) else "", "Y" if has_size_mf(p) else "",
                              st if st!="GAP" else ("GAP (one-size?)" if sub=="one_size_likely" else "GAP"), action])
    if ec:
        color_scope += 1
        ct = classify(has_color_opt(p), has_color_mf(p))
        color_buckets[ct] += 1
        if ct != "READY":
            action = {"OPT_ONLY":"Populate shopify.color-pattern metafield (enables swatch filter)",
                      "MF_ONLY":"Add a Color variant option",
                      "GAP":"Add Color variant option + shopify.color-pattern metafield"}[ct]
            color_rows.append([cat, label, p["handle"], p["title"][:60],
                               "Y" if has_color_opt(p) else "", "Y" if has_color_mf(p) else "", ct, action])
    # curated extras population among apparel/dresses
    if cat == "Apparel":
        extras_scope += 1
        if p.get("fabric"): extras["fabric"] += 1
        if p.get("dressOcc"): extras["dressOcc"] += 1
        if p.get("fit"): extras["fit"] += 1
        if p.get("topLen") or p.get("skirtLen"): extras["length"] += 1

print(f"in-stock: {len(instock)}")
print(f"SIZE scope {size_scope}: {dict(size_buckets)}")
print(f"COLOR scope {color_scope}: {dict(color_buckets)}")
print(f"extras scope (apparel) {extras_scope}: {extras}")
json.dump({"instock":len(instock),
           "size_scope":size_scope,"size_buckets":dict(size_buckets),
           "color_scope":color_scope,"color_buckets":dict(color_buckets),
           "extras_scope":extras_scope,"extras":extras,
           "size_worklist":len(size_rows),"color_worklist":len(color_rows)},
          open(os.path.join(os.path.dirname(__file__),"..","data","filter-summary.json"),"w"), indent=1)

# ---------------- workbook ----------------
H=PatternFill("solid",fgColor="1F3864"); SUB=PatternFill("solid",fgColor="2E5496")
OK=PatternFill("solid",fgColor="C6EFCE"); WARN=PatternFill("solid",fgColor="FFEB9C"); BAD=PatternFill("solid",fgColor="FFC7CE")
WHITE=Font(color="FFFFFF",bold=True); TH=Side(style="thin",color="D9D9D9"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
wb=openpyxl.Workbook()

def sheet(title):
    ws=wb.create_sheet(title); ws.sheet_view.showGridLines=False; return ws

# Summary
ws=wb.active; ws.title="Corrected Summary"; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=2
for col,w in zip("BCDEF",[40,14,14,14,46]): ws.column_dimensions[col].width=w
r=2
ws.cell(r,2,"Filter Readiness — Corrected (Dual Standard)").font=Font(size=15,bold=True,color="1F3864"); r+=1
ws.cell(r,2,f"In-stock scoped ({len(instock):,} products). READY = variant option AND taxonomy metafield. "
            f"Re-fetched live incl. shopify.size / shoe-size / ring-size / color-pattern.").font=Font(italic=True,color="595959"); r+=2

def block(ws, r, title, scope, buckets, mapping):
    ws.cell(r,2,title).font=Font(size=12,bold=True,color="1F3864"); r+=1
    for h,c in zip(["Bucket","Count","% of scope","What it means"],[2,3,4,5]):
        cell=ws.cell(r,c,h); cell.font=WHITE; cell.fill=SUB
    r+=1
    for key,lbl,meaning,fill in mapping:
        cnt=buckets.get(key,0)
        ws.cell(r,2,lbl).border=BD
        ws.cell(r,3,cnt).border=BD
        pc=ws.cell(r,4,(cnt/scope if scope else 0)); pc.number_format="0.0%"; pc.border=BD
        m=ws.cell(r,5,meaning); m.border=BD; m.alignment=Alignment(wrap_text=True)
        for cc in (2,3,4,5): ws.cell(r,cc).fill=fill
        r+=1
    tot=ws.cell(r,2,"Scope total"); tot.font=Font(bold=True); ws.cell(r,3,scope).font=Font(bold=True)
    return r+2

r=block(ws,r,"SIZE  (types that expect size)",size_scope,size_buckets,[
    ("READY","✅ Ready (option + metafield)","Selectable on PDP and powers a normalized size filter.",OK),
    ("OPT_ONLY","◐ Option only","Selectable + filterable on raw values; needs shopify.size to normalize.",WARN),
    ("MF_ONLY","◐ Metafield only","Filterable but NOT selectable on the PDP; needs a Size variant option.",WARN),
    ("GAP:real_gap","✗ Real gap","No size anywhere — a sized item missing its sizes.",BAD),
    ("GAP:one_size_likely","? Likely one-size","Ponchos/wraps/scarves — confirm, then set 'One Size' in both places.",WARN),
])
r=block(ws,r,"COLOR  (types that expect color)",color_scope,color_buckets,[
    ("READY","✅ Ready (option + metafield)","Selectable + drives the normalized color-swatch filter.",OK),
    ("OPT_ONLY","◐ Option only","Filterable on raw values (BLPK, GREY…); needs color-pattern for swatches.",WARN),
    ("MF_ONLY","◐ Metafield only","Swatch data present but no Color option to select on PDP.",WARN),
    ("GAP","✗ Gap","No color anywhere.",BAD),
])

# Reconciliation vs old report
ws.cell(r,2,"Why this differs from the original 'missing size/color' counts").font=Font(size=12,bold=True,color="C55A11"); r+=1
for line in [
  "Original metric = 'no variant option literally named Size' — options only, one metafield (color-pattern) fetched.",
  "That jammed four different situations into one number:",
  "  • Shoe size / Ring size options were not counted as size (false negatives).",
  "  • Products carrying size only in a metafield were counted missing.",
  "  • Genuinely one-size items (ponchos) were counted missing though they need no size run.",
  "  • Real gaps — the only ones that need sizes added.",
  "Corrected metric fetches the taxonomy metafields and applies the dual standard you set.",
]:
    c=ws.cell(r,2,line); c.alignment=Alignment(wrap_text=True); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5); r+=1

# chart on summary (size buckets)
cs=wb.create_sheet("_chartdata"); cs.sheet_view.showGridLines=False
order=[("Ready","READY"),("Option only","OPT_ONLY"),("Metafield only","MF_ONLY"),("Real gap","GAP:real_gap"),("One-size?","GAP:one_size_likely")]
cs.cell(1,1,"Bucket"); cs.cell(1,2,"Size")
for i,(lbl,k) in enumerate(order,start=2):
    cs.cell(i,1,lbl); cs.cell(i,2,size_buckets.get(k,0))
ch=BarChart(); ch.title="Size readiness (in-stock, size-expecting)"; ch.height=7; ch.width=14
data=Reference(cs,min_col=2,min_row=1,max_row=1+len(order)); cats=Reference(cs,min_col=1,min_row=2,max_row=1+len(order))
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.legend=None
ws.add_chart(ch, f"B{r+1}")

def worklist(name, header, data_rows):
    ws=sheet(name)
    ws.cell(1,1,f"{name} — {len(data_rows):,} products need action").font=Font(size=13,bold=True,color="1F3864")
    for i,h in enumerate(header,start=1):
        c=ws.cell(3,i,h); c.font=WHITE; c.fill=H; c.alignment=Alignment(wrap_text=True)
    for ri,row in enumerate(data_rows,start=4):
        for ci,v in enumerate(row,start=1):
            c=ws.cell(ri,ci,v); c.border=BD; c.alignment=Alignment(wrap_text=True,vertical="top")
    widths=[14,22,30,46,8,9,16,48]
    for i,w in enumerate(widths,start=1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes="A4"

HDR=["Category","Type","Handle","Title","Opt?","MF?","Status","Recommended action"]
worklist("Size Worklist", HDR, sorted(size_rows, key=lambda x:(x[6],x[0])))
worklist("Color Worklist", HDR, sorted(color_rows, key=lambda x:(x[6],x[0])))

# extras opportunities
ws=sheet("Filter Opportunities")
ws.cell(1,1,"Richer filters within reach (taxonomy already defined — measuring population)").font=Font(size=12,bold=True,color="1F3864")
ws.cell(2,1,f"Scope: {extras_scope:,} in-stock apparel products").font=Font(italic=True,color="595959")
for i,h in enumerate(["Attribute (metafield)","Populated","% of apparel","Note"],start=1):
    c=ws.cell(4,i,h); c.font=WHITE; c.fill=SUB
opp=[("shopify.fabric","fabric","Fabric filter — popular for apparel"),
     ("shopify.dress-occasion","dressOcc","Occasion filter (wedding, work, casual)"),
     ("shopify.fit","fit","Fit filter (relaxed, slim, true-to-size)"),
     ("shopify.*-length-type","length","Length filter (mini/midi/maxi, top length)")]
for i,(mf,key,note) in enumerate(opp,start=5):
    ws.cell(i,1,mf).border=BD; ws.cell(i,2,extras[key]).border=BD
    pc=ws.cell(i,3,(extras[key]/extras_scope if extras_scope else 0)); pc.number_format="0.0%"; pc.border=BD
    ws.cell(i,4,note).border=BD
for col,w in zip("ABCD",[26,12,14,46]): ws.column_dimensions[col].width=w

wb._sheets.remove(cs); wb._sheets.append(cs)  # keep chartdata last
out="The Clothing Cove - Filter Readiness CORRECTED (June 2026).xlsx"
wb.save(out)
print("saved:", out)
