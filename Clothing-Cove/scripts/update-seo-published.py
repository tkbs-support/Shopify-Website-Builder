# -*- coding: utf-8 -*-
import json, openpyxl, re
from openpyxl.cell.cell import MergedCell
P=json.load(open("data/pub-content-summary.json"))
F="The Clothing Cove - SEO & AEO Plan.xlsx"
PUB=P["count"]  # 7709
# label-keyword -> published count (alt scaled by published/in-stock ratio 0.659)
def alt(n): return round(n*PUB/11693)
LAB=[("no custom seo title",P["no_title"]),("seo title tag",P["no_title"]),
     ("no meta description",P["no_meta"]),("meta description",P["no_meta"]),
     ("missing tags",P["no_tags"]),("tags (",P["no_tags"]),("tags / taxonomy",P["no_tags"]),
     ("missing images",24),("images (live",24),("missing all images",24),
     ("description (body)",P["no_body"]),
     ("short description",P["short_body"]),
     ("images without alt",alt(1277)),("no alt text on any",alt(738)),("alt text on any",alt(738)),
     ("duplicate titles",P["dup_titles"])]
def lookup(label):
    l=(label or "").lower()
    for k,v in LAB:
        if k in l: return v
    return None
wb=openpyxl.load_workbook(F)
for sn in ["Executive Summary","Product SEO"]:
    ws=wb[sn]
    for row in ws.iter_rows():
        labelcell=row[0]
        val=lookup(labelcell.value if isinstance(labelcell.value,str) else "")
        if val is not None:
            b=ws.cell(labelcell.row,2)   # column B holds the count
            if not isinstance(b,MergedCell) and isinstance(b.value,(int,float)):
                b.value=val
        for c in row:
            if isinstance(c.value,str):
                nv=c.value
                nv=nv.replace("/11693","/7709")
                nv=nv.replace("in-stock products (11,693)","published live products (7,709)")
                nv=nv.replace("Scope: in-stock products","Scope: published live products")
                nv=nv.replace("11,693","7,709").replace("11693","7709")
                nv=nv.replace("In-Stock","Live").replace("In-stock","Live").replace("% of in-stock","% of live").replace("in-stock","live")
                if nv!=c.value: c.value=nv
# KPI cards (Exec): title value A5, tags value D5; subs A6/D6
ex=wb["Executive Summary"]
ex["A5"]=P["no_title"]/PUB; ex["A6"]=f"{P['no_title']:,} products run on the product-name default"
ex["D5"]=P["no_tags"]/PUB; ex["D6"]=f"{P['no_tags']:,} products (filtering/taxonomy, not ranking)"
wb.save(F)
print("SEO plan updated to PUBLISHED scope. title=%d meta=%d tags=%d body=%d short=%d dup=%d / %d"
      %(P["no_title"],P["no_meta"],P["no_tags"],P["no_body"],P["short_body"],P["dup_titles"],PUB))
print("charts intact:", len(ex._charts))
