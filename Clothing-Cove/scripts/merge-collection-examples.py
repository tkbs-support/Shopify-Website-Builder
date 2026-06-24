# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
NAVY=PatternFill("solid",fgColor="1F3864"); WHITE=Font(color="FFFFFF",bold=True)
TH=Side(style="thin",color="D9D9D9"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
rb=openpyxl.load_workbook("The Clothing Cove - Collection SEO Rebuilt (June 2026).xlsx")["Collection SEO (Rebuilt)"]
examples=[]
for r in range(5, rb.max_row+1):
    kind=rb.cell(r,1).value
    if not kind or kind=="UTILITY": continue
    examples.append({"url":rb.cell(r,2).value,"old":rb.cell(r,3).value,"new":rb.cell(r,4).value,
                     "meta":rb.cell(r,5).value,"desc":rb.cell(r,6).value})
    if len(examples)>=30: break
wb=openpyxl.load_workbook("The Clothing Cove - SEO & AEO Plan.xlsx")
idx=wb.sheetnames.index("Collection SEO"); del wb["Collection SEO"]
ws=wb.create_sheet("Collection SEO", idx); ws.sheet_view.showGridLines=False
ws.cell(1,1,"Collection SEO — Before → After (sample of 30)").font=Font(size=14,bold=True,color="1F3864")
c=ws.cell(2,1,"Representative examples of the rewrite. Full worklist of 255 navigable collections is provided as a "
              "delivery file. Each rewrite gives a keyword-rich, location-aware title + meta + on-page description.")
c.font=Font(italic=True,color="595959"); ws.merge_cells("A2:E2"); c.alignment=Alignment(wrap_text=True)
for i,h in enumerate(["Collection","Current title","Proposed SEO title","Meta description","Collection description"],start=1):
    cc=ws.cell(4,i,h); cc.font=WHITE; cc.fill=NAVY; cc.alignment=Alignment(wrap_text=True,vertical="center")
for ri,e in enumerate(examples,start=5):
    for ci,v in enumerate([e["url"],e["old"],e["new"],e["meta"],e["desc"]],start=1):
        cell=ws.cell(ri,ci,v); cell.border=BD; cell.alignment=Alignment(wrap_text=True,vertical="top")
        if ci==2: cell.font=Font(color="9C5700")
        if ci==3: cell.font=Font(bold=True,color="1F3864")
for col,w in zip("ABCDE",[30,26,40,46,60]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"
wb.save("The Clothing Cove - SEO & AEO Plan.xlsx")
# verify no brand+category word-doubling in the 30 examples
import re
dupes=[e["new"] for e in examples if re.search(r'\b(\w+)\s+\1\b', e["new"], re.I)]
print("examples merged:", len(examples), "| word-doubling found:", dupes if dupes else "NONE")
