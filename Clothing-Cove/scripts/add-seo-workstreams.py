# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
F="The Clothing Cove - SEO & AEO Plan.xlsx"
NAVY=PatternFill("solid",fgColor="1F3864"); BAND=PatternFill("solid",fgColor="FFF2CC"); GREEN=PatternFill("solid",fgColor="C6EFCE")
WHITE=Font(color="FFFFFF",bold=True); TH=Side(style="thin",color="D9D9D9"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
wb=openpyxl.load_workbook(F); fr=wb["Fix Roadmap"]
r=18
h=fr.cell(r,1,"ADDED FROM TECHNICAL REVIEW (June 2026) — run Phase 0 first"); h.font=Font(bold=True,color="1F3864"); r+=1
rows=[
 ("P0","Connect Search Console + GA4; pull 16-month traffic baseline + current ranking & competitor snapshot",
   "Measurement setup","4–8 hrs","Diagnoses the multi-year decline and makes the projected lift measurable — not promised"),
 ("P2","Indexation hygiene — noindex/prune ~200 thin, empty, event & duplicate collections; dedupe '-copy' products",
   "~200+ collections","8–12 hrs","Stops thin/duplicate pages competing with real category pages for crawl budget"),
 ("P2","Install + seed a product reviews app (none today); add FAQPage schema to /faqs",
   "Store-wide","6–10 hrs + ongoing","Unlocks star-rating + FAQ rich results — the highest-CTR results for a boutique"),
]
for pr,act,scope,eff,imp in rows:
    fr.cell(r,1,pr).font=Font(bold=True); fr.cell(r,1).fill=GREEN if pr=="P0" else BAND
    for ci,v in zip((2,3,4,5),[act,scope,eff,imp]):
        c=fr.cell(r,ci,v); c.border=BD; c.alignment=Alignment(wrap_text=True,vertical="top")
    fr.cell(r,1).border=BD
    r+=1
r+=1
note=fr.cell(r,1,"Scope note: the $14,000 Foundation (see Projected Impact) covers Phase 0 + theme/homepage + priority "
   "collection content + the core filters. P2–P4 full-catalog enrichment (titles, descriptions, tags — 1,300+ hrs) is "
   "scoped and quoted separately, per batch.")
note.font=Font(italic=True,size=9,color="595959"); note.alignment=Alignment(wrap_text=True,vertical="top")
fr.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
wb.save(F)
print("workstreams added. Fix Roadmap maxrow:", fr.max_row, "| charts:", len(wb["Executive Summary"]._charts))
