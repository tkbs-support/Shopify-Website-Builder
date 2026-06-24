# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill
F="The Clothing Cove - SEO & AEO Plan.xlsx"
RED=PatternFill("solid",fgColor="FFC7CE"); AMBER=PatternFill("solid",fgColor="FFEB9C"); GREEN=PatternFill("solid",fgColor="C6EFCE")
wb=openpyxl.load_workbook(F)

ex=wb["Executive Summary"]
# images -> live reality
ex["A13"]="Missing images (live store; ~3,800 more are unpublished)"; ex["B13"]=24; ex["E13"]="OK"; ex["E13"].fill=GREEN
# duplicate titles -> published
ex["A15"]="Duplicate titles (live store)"; ex["B15"]=36; ex["E15"]="OK"; ex["E15"].fill=GREEN
# titles/meta -> hygiene (downgrade from CRITICAL)
ex["A12"]="No meta description (page falls back to body)"
ex["E10"]="HYGIENE"; ex["E10"].fill=AMBER
ex["E12"]="HYGIENE"; ex["E12"].fill=AMBER

ps=wb["Product SEO"]
ps["A5"]="No custom SEO title (runs on product-name default)"; ps["F5"]="HYGIENE"; ps["F5"].fill=AMBER
ps["A7"]="No custom meta description (falls back to body)"; ps["F7"]="HYGIENE"; ps["F7"].fill=AMBER
ps["G7"]="Page renders the body text as the description; the work is optimization/quality, not filling a void."
ps["A8"]="Missing images (live store)"; ps["B8"]=24; ps["F8"]="OK"; ps["F8"].fill=GREEN
ps["G8"]="Live storefront is fine; ~3,800 active in-stock products are UNPUBLISHED (not on the store) — handled as the publish/photograph backlog."
ps["H8"]="Publish/photograph the backlog (see Indexation & Backlog)"
ps["H6"]="Tag from type + vendor + attributes"
ps["A13"]="Duplicate titles (live store; 13 shared)"; ps["B13"]=36; ps["F13"]="OK"; ps["F13"].fill=GREEN
ps["G13"]="Minor on the live store (909 across ALL active incl. unpublished/archived)."

aeo=wb["AEO Readiness"]
aeo["E7"]="Tag from type/vendor/attributes"
aeo["C7"]="WARNING"; aeo["C7"].fill=AMBER

fr=wb["Fix Roadmap"]
fr["B10"]="Tag in-stock products"
fr["C9"]="36 live (909 incl. unpublished)"

wb.save(F)
print("metric corrections applied. charts:", len(ex._charts))
# sanity: no remaining 3838 or 909 in these cells
for co in ["Executive Summary!B13","Executive Summary!B15","Product SEO!B8","Product SEO!B13"]:
    s,c=co.split("!"); print(" ",co,"=",wb[s][c].value)
