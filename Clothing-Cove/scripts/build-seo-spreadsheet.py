from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1B2631")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1B2631")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=11, color="1B2631")
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", bold=True, size=10)
RED = Font(name="Arial", bold=True, size=10, color="C0392B")
GREEN = Font(name="Arial", bold=True, size=10, color="27AE60")
ORANGE = Font(name="Arial", bold=True, size=10, color="E67E22")
GRAY = Font(name="Arial", size=10, color="7F8C8D")
ALT = PatternFill("solid", fgColor="F2F3F4")
WHITE = PatternFill("solid", fgColor="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D5DBDB")
RED_FILL = PatternFill("solid", fgColor="FADBD8")
GREEN_FILL = PatternFill("solid", fgColor="D5F5E3")
YELLOW_FILL = PatternFill("solid", fgColor="FEF9E7")
BORDER = Border(bottom=Side(style="thin", color="BDC3C7"))
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def data_row(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = ALT if alt else WHITE
        cell.font = BODY
        cell.border = BORDER


# ════════════════════════════════════════════════════════════
# Sheet 1: Executive Summary
# ════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Executive Summary"
ws.sheet_properties.tabColor = "1B2631"
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 50

ws.merge_cells("A1:D1")
ws["A1"] = "The Clothing Cove - SEO & AEO Audit"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Audit Date: May 2026 | Store: theclothingcove.com | 35,815 active / 11,818 in stock"
ws["A2"].font = GRAY
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 14

def section(ws, row, title):
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    ws.cell(row=row, column=1, value=title).font = SUBTITLE_FONT
    return row + 1

def metric(ws, row, label, value, grade, note="", in_stock_val=None, in_stock_pct=None):
    ws.cell(row=row, column=1, value=label).font = BODY
    c = ws.cell(row=row, column=2, value=value)
    c.font = BOLD
    if isinstance(value, (int, float)):
        c.number_format = "#,##0"
    g = ws.cell(row=row, column=3, value=grade)
    if grade == "CRITICAL":
        g.font = RED; g.fill = RED_FILL
    elif grade == "WARNING":
        g.font = ORANGE; g.fill = YELLOW_FILL
    elif grade == "OK":
        g.font = GREEN; g.fill = GREEN_FILL
    else:
        g.font = BODY
    g.alignment = CENTER
    if note:
        ws.cell(row=row, column=4, value=note).font = BODY
        ws.cell(row=row, column=4).alignment = WRAP
    if in_stock_val is not None:
        isv = ws.cell(row=row, column=5, value=in_stock_val)
        isv.font = Font(name="Arial", bold=True, size=10, color="2980B9")
        if isinstance(in_stock_val, (int, float)):
            isv.number_format = "#,##0"
        isv.alignment = CENTER
    if in_stock_pct is not None:
        isp = ws.cell(row=row, column=6, value=in_stock_pct)
        isp.font = Font(name="Arial", size=10, color="2980B9")
        isp.alignment = CENTER
    for c in range(1, 7):
        ws.cell(row=row, column=c).border = BORDER
    return row + 1

# Column headers for in-stock scope
r = 3
ws.cell(row=r, column=1, value="Metric").font = BOLD
ws.cell(row=r, column=2, value="Full Catalog").font = BOLD
ws.cell(row=r, column=3, value="Severity").font = BOLD
ws.cell(row=r, column=4, value="Notes").font = BOLD
ws.cell(row=r, column=5, value="In-Stock (Work Scope)").font = Font(name="Arial", bold=True, size=10, color="2980B9")
ws.cell(row=r, column=6, value="% of In-Stock").font = Font(name="Arial", bold=True, size=10, color="2980B9")
for c in range(1, 7):
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="AEB6BF")
    ws.cell(row=r, column=c).alignment = CENTER

r = 4
r = section(ws, r, "PRODUCT SEO (35,815 active / 11,818 in stock)")
r = metric(ws, r, "Products missing body/description", 7694, "CRITICAL", "21.5% of full catalog", 3648, "30.9%")
r = metric(ws, r, "Products with short description (<100 chars)", 2595, "WARNING", "7.2% of full catalog", 622, "5.3%")
r = metric(ws, r, "Products missing ALL images", 8095, "CRITICAL", "22.6% of full catalog", 3967, "33.6%")
r = metric(ws, r, "Images missing alt text", 3705, "WARNING", "7.0% of 53,227 images", 1314, "7.0% of 18,645")
r = metric(ws, r, "Products with zero alt text on any image", 2471, "WARNING", "6.9% of full catalog", 744, "6.3%")
r = metric(ws, r, "Products missing tags", 21643, "CRITICAL", "60.4% of full catalog", 11061, "93.6%")
r = metric(ws, r, "Estimated missing SEO title tag", 12460, "CRITICAL", "34.8% use auto-generated titles")
r = metric(ws, r, "Estimated missing SEO meta description", 358, "OK", "~1% — good coverage via SEOMetaManager")
r = metric(ws, r, "Duplicate product titles", "1,679 / 150 titles", "WARNING", "Multiple products sharing same title", "910 / 59 titles", "7.7%")
r = metric(ws, r, "Missing product type", 165, "OK", "Minor — 0.5%", 50)
r = metric(ws, r, "Missing vendor", 208, "OK", "Minor — 0.6%", 93)
r += 1

r = section(ws, r, "INVENTORY / FILTERING (in-stock focus)")
r = metric(ws, r, "Missing Color option", 18938, "", "53% of full catalog", 6354, "53.8%")
r = metric(ws, r, "Missing Size option", "", "", "", 8662, "73.3%")
r = metric(ws, r, "Default Title (no options at all)", 17253, "", "48.3% of full catalog", 5789, "49.0%")
r += 1

r = section(ws, r, "COLLECTION SEO")
r = metric(ws, r, "Total collections", 259, "", "")
r = metric(ws, r, "Collections missing description", 173, "CRITICAL", "66.8% have no SEO-friendly description")
r = metric(ws, r, "Collections with short description", 2, "OK", "")
r = metric(ws, r, "Collections with good description", 84, "WARNING", "Only 32.4% have substantive descriptions")
r += 1

r = section(ws, r, "THEME / TECHNICAL SEO")
r = metric(ws, r, "Canonical tags", "Present", "OK", "Shopify handles canonical URLs")
r = metric(ws, r, "Open Graph tags", "Missing", "WARNING", "No og: meta tags — hurts social sharing previews")
r = metric(ws, r, "Twitter Card tags", "Missing", "WARNING", "No twitter: meta tags")
r = metric(ws, r, "JSON-LD structured data (in theme)", "Missing", "CRITICAL", "No JSON-LD in theme.liquid — relying on microdata only")
r = metric(ws, r, "Microdata schema snippet", "Present", "OK", "snippets/microdata-schema.liquid exists")
r = metric(ws, r, "Viewport meta", "Present", "OK", "Mobile-friendly")
r = metric(ws, r, "Custom robots.txt", "Missing", "OK", "Using Shopify defaults — usually fine")
r = metric(ws, r, "Hreflang tags", "Missing", "OK", "Single-language store, not needed")
r = metric(ws, r, "Preconnect hints", "Missing", "WARNING", "Could improve load time for external resources")
r = metric(ws, r, "Custom homepage title", "Present", "OK", "Already customized via push-seo-fixes.py")
r = metric(ws, r, "Custom homepage meta description", "Present", "OK", "Already customized")
r += 1

r = section(ws, r, "PAGES & BLOG")
r = metric(ws, r, "Total pages", 31, "", "")
r = metric(ws, r, "Published pages", 11, "WARNING", "20 pages in draft — potential content not being used")
r = metric(ws, r, "Blog articles", 41, "WARNING", "Content exists but unclear if actively maintained")
r += 1

r = section(ws, r, "LOCAL SEO & CITATIONS")
r = metric(ws, r, "Google search visibility (branded)", "Strong", "OK", "Ranks #1 for 'The Clothing Cove Milford MI'")
r = metric(ws, r, "Google search (category queries)", "Strong", "OK", "Ranks #1 for 'dress shop Milford MI'")
r = metric(ws, r, "Brighton.com store listing", "Present", "OK", "Listed as official Brighton retailer")
r = metric(ws, r, "Yelp listing", "80 reviews", "OK", "Active with photos and positive reviews")
r = metric(ws, r, "Facebook page", "Active", "OK", "Main page + Mother of Bride specialty page")
r = metric(ws, r, "Instagram", "Active", "OK", "@theclothingcove_")
r = metric(ws, r, "Yellow Pages", "Present", "OK", "Listed correctly")
r = metric(ws, r, "LinkedIn", "Present", "OK", "Company page exists")
r = metric(ws, r, "Waze listing", "Present", "OK", "Correct address")
r = metric(ws, r, "Open Graph / social previews", "Missing", "WARNING", "Links shared on social media won't have rich previews")
r += 1

r = section(ws, r, "AEO (AI ENGINE OPTIMIZATION)")
r = metric(ws, r, "JSON-LD structured data", "Minimal", "CRITICAL", "No JSON-LD in theme — AI crawlers rely heavily on structured data")
r = metric(ws, r, "FAQ page content", "12,004 chars", "OK", "Good FAQ page exists — needs FAQ schema markup for AI")
r = metric(ws, r, "LocalBusiness schema", "Added via script", "OK", "add-local-business-schema.py was run previously")
r = metric(ws, r, "Product schema", "Microdata only", "WARNING", "Prestige theme uses microdata — JSON-LD preferred for AI")
r = metric(ws, r, "BreadcrumbList schema", "Unknown", "WARNING", "Need to verify on product/collection pages")
r = metric(ws, r, "Content depth for AI citation", "Weak", "CRITICAL", "60% of products have no tags, 21% no description — thin content for AI")


# ════════════════════════════════════════════════════════════
# Sheet 2: Product SEO Detail
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Product SEO")
ws2.sheet_properties.tabColor = "2980B9"

headers = ["Issue", "Full Catalog", "% Catalog", "In-Stock Only", "% In-Stock", "Severity", "Impact", "Fix Approach"]
widths = [34, 14, 12, 14, 12, 12, 40, 44]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    ws2.cell(row=1, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = w
header_row(ws2, 1, len(headers))

product_issues = [
    ("Missing product description", 7694, "21.5%", 3648, "30.9%", "CRITICAL", "No content for Google or AI to index.", "Bulk generate from title + vendor + type + tags"),
    ("Short description (<100 chars)", 2595, "7.2%", 622, "5.3%", "WARNING", "Thin content — ranks poorly, AI ignores.", "Expand with feature/benefit copy"),
    ("Missing ALL product images", 8095, "22.6%", 3967, "33.6%", "CRITICAL", "No visual search, Google Shopping, or social.", "Upload photos or hide/archive these products"),
    ("Images without alt text", 3705, "7.0%", 1314, "7.0%", "WARNING", "Missed image SEO. Accessibility issue.", "Script to auto-generate alt from title + color"),
    ("Products with NO alt on any image", 2471, "6.9%", 744, "6.3%", "WARNING", "Entire product invisible to image search.", "Same script — prioritize these"),
    ("Missing tags", 21643, "60.4%", 11061, "93.6%", "CRITICAL", "Cannot filter, categorize, or surface.", "Bulk tag from type + vendor + category"),
    ("Missing SEO title tag", "~12,460", "~34.8%", "", "", "CRITICAL", "Auto-generated titles, not optimized.", "Bulk generate from title + category + brand"),
    ("Missing SEO meta description", "~358", "~1.0%", "", "", "OK", "Good coverage via SEOMetaManager.", "Fix remaining ~358 manually"),
    ("Duplicate product titles", "1,679", "4.7%", "910", "7.7%", "WARNING", "Competing with own products in search.", "Append color/size/variant to differentiate"),
    ("Missing Color option", "", "", 6354, "53.8%", "CRITICAL", "Cannot filter by color.", "Add color options by product type"),
    ("Missing Size option", "", "", 8662, "73.3%", "WARNING", "Cannot filter by size.", "Add size options where relevant"),
    ("Default Title (no options)", "", "", 5789, "49.0%", "WARNING", "Single-variant — no filtering possible.", "Add options to relevant product types"),
]

for i, row_data in enumerate(product_issues):
    issue, full, pct_f, instock, pct_i, sev, impact, fix = row_data
    r = i + 2
    ws2.cell(row=r, column=1, value=issue)
    fc = ws2.cell(row=r, column=2, value=full)
    if isinstance(full, int): fc.number_format = "#,##0"
    ws2.cell(row=r, column=3, value=pct_f).alignment = CENTER
    isc = ws2.cell(row=r, column=4, value=instock)
    if isinstance(instock, int): isc.number_format = "#,##0"
    isc.font = Font(name="Arial", bold=True, size=10, color="2980B9")
    ws2.cell(row=r, column=5, value=pct_i).alignment = CENTER
    ws2.cell(row=r, column=5).font = Font(name="Arial", size=10, color="2980B9")
    s = ws2.cell(row=r, column=6, value=sev)
    s.alignment = CENTER
    if sev == "CRITICAL": s.font = RED
    elif sev == "WARNING": s.font = ORANGE
    else: s.font = GREEN
    ws2.cell(row=r, column=7, value=impact).alignment = WRAP
    ws2.cell(row=r, column=8, value=fix).alignment = WRAP
    ws2.row_dimensions[r].height = 35
    data_row(ws2, r, len(headers), alt=(i % 2 == 1))


# ════════════════════════════════════════════════════════════
# Sheet 3: Collection SEO
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Collection SEO")
ws3.sheet_properties.tabColor = "27AE60"

ws3.merge_cells("A1:D1")
ws3["A1"] = "Collection Description Coverage"
ws3["A1"].font = TITLE_FONT
ws3["A2"] = "259 total collections | 173 missing descriptions (66.8%) | 84 with descriptions (32.4%)"
ws3["A2"].font = GRAY

ws3.cell(row=4, column=1, value="Status").font = HEADER_FONT; ws3.cell(row=4, column=1).fill = HEADER_FILL
ws3.cell(row=4, column=2, value="Count").font = HEADER_FONT; ws3.cell(row=4, column=2).fill = HEADER_FILL
ws3.cell(row=4, column=3, value="% of Total").font = HEADER_FONT; ws3.cell(row=4, column=3).fill = HEADER_FILL
ws3.cell(row=4, column=4, value="SEO Impact").font = HEADER_FONT; ws3.cell(row=4, column=4).fill = HEADER_FILL
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 50

coll_data = [
    ("Missing description", 173, "66.8%", "No unique content for Google. Collection pages look thin. AI ignores these."),
    ("Short description (<100 chars)", 2, "0.8%", "Barely any content — almost as bad as missing."),
    ("Has good description", 84, "32.4%", "These collection pages are SEO-ready."),
]
for i, (status, count, pct, impact) in enumerate(coll_data):
    r = i + 5
    ws3.cell(row=r, column=1, value=status)
    ws3.cell(row=r, column=2, value=count).number_format = "#,##0"
    ws3.cell(row=r, column=3, value=pct).alignment = CENTER
    ws3.cell(row=r, column=4, value=impact).alignment = WRAP
    data_row(ws3, r, 4, alt=(i % 2 == 1))

ws3.cell(row=9, column=1, value="NOTE:").font = BOLD
ws3.cell(row=9, column=2, value="A collection-seo-plan.md and spreadsheet already exist in the repo. Check execution status.").font = BODY
ws3.merge_cells("B9:D9")


# ════════════════════════════════════════════════════════════
# Sheet 4: Theme & Technical SEO
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Technical SEO")
ws4.sheet_properties.tabColor = "8E44AD"

t_headers = ["Check", "Status", "Severity", "Details / Recommendation"]
t_widths = [30, 12, 12, 60]
for i, (h, w) in enumerate(zip(t_headers, t_widths), 1):
    ws4.cell(row=1, column=i, value=h)
    ws4.column_dimensions[get_column_letter(i)].width = w
header_row(ws4, 1, len(t_headers))

tech_checks = [
    ("Canonical tags", "Present", "OK", "Shopify auto-generates canonical URLs. No issues found."),
    ("Open Graph tags (og:)", "Missing", "WARNING", "Theme.liquid has no og: meta tags. Social shares show generic/no preview. Add og:title, og:description, og:image, og:url to theme.liquid."),
    ("Twitter Card tags", "Missing", "WARNING", "No twitter: meta tags. Twitter/X shares show plain links. Add twitter:card, twitter:title, twitter:description, twitter:image."),
    ("JSON-LD structured data", "Missing", "CRITICAL", "No application/ld+json in theme.liquid. Only microdata via snippets/microdata-schema.liquid. JSON-LD is preferred by Google and essential for AI engines. Add Product, Organization, BreadcrumbList, and FAQ schema in JSON-LD format."),
    ("Microdata schema", "Present", "OK", "snippets/microdata-schema.liquid exists. Legacy format but functional for basic product data."),
    ("Viewport meta tag", "Present", "OK", "Mobile-friendly viewport configuration in place."),
    ("Charset declaration", "Present", "OK", "UTF-8 charset properly declared."),
    ("Custom robots.txt", "Not present", "OK", "Using Shopify defaults. Appropriate for most stores."),
    ("Hreflang tags", "Not present", "OK", "Single-language English store. Not needed."),
    ("Preconnect hints", "Missing", "WARNING", "No resource hints for external domains (CDN, fonts, analytics). Adding preconnect can improve page load by 100-300ms."),
    ("Homepage title override", "Present", "OK", "Custom homepage title implemented via push-seo-fixes.py script."),
    ("Homepage meta description", "Present", "OK", "Custom meta description for homepage is in place."),
    ("Hidden H1 on homepage", "Present", "OK", "Visually hidden H1 added via push-seo-fixes.py for SEO."),
    ("Product template", "JSON-based", "OK", "templates/product.json with sections: main, recommendations, recently-viewed."),
    ("Collection template", "JSON-based", "OK", "templates/collection.json exists."),
    ("LocalBusiness schema", "Added", "OK", "Previously implemented via add-local-business-schema.py."),
    ("FAQ schema", "Added", "OK", "Previously implemented via add-faq-schema.py."),
    ("Image alt text (theme-level)", "Fixed", "OK", "Slideshow and shop-the-look alt fallbacks added via fix-alt-text.py."),
]

for i, (check, status, sev, detail) in enumerate(tech_checks):
    r = i + 2
    ws4.cell(row=r, column=1, value=check)
    s = ws4.cell(row=r, column=2, value=status)
    if "Missing" in status or "Not present" in status:
        s.font = RED
    else:
        s.font = GREEN
    s.alignment = CENTER
    sv = ws4.cell(row=r, column=3, value=sev)
    sv.alignment = CENTER
    if sev == "CRITICAL": sv.font = RED
    elif sev == "WARNING": sv.font = ORANGE
    else: sv.font = GREEN
    ws4.cell(row=r, column=4, value=detail).alignment = WRAP
    ws4.row_dimensions[r].height = 40
    data_row(ws4, r, len(t_headers), alt=(i % 2 == 1))


# ════════════════════════════════════════════════════════════
# Sheet 5: Pages & Blog
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Pages & Blog")
ws5.sheet_properties.tabColor = "E67E22"

p_headers = ["Page Title", "Handle", "Status", "Content Length", "Notes"]
p_widths = [38, 40, 12, 16, 40]
for i, (h, w) in enumerate(zip(p_headers, p_widths), 1):
    ws5.cell(row=1, column=i, value=h)
    ws5.column_dimensions[get_column_letter(i)].width = w
header_row(ws5, 1, len(p_headers))

pages = [
    ("About Us", "about-us", "Published", 2782, "Core page — good"),
    ("Contact Us/Hours", "contact-us", "Published", 819, "Core page — NAP info"),
    ("FAQs", "faqs", "Published", 12004, "Strong content — has FAQ schema"),
    ("Event Calendar", "event-calendar", "Published", 1952, "Active events page"),
    ("Join Our Team", "cove-joinourteam", "Published", 258, "Thin content — expand"),
    ("Our Favorite Brands", "our-favorite-brands", "Published", 35, "VERY thin — only 35 chars. Needs full brand content."),
    ("Gift Card Balance", "https-the-clothing-cove-store...", "Published", 647, "Bad handle URL — investigate"),
    ("Milford Garden Walk", "milford-garden-walk", "Published", 721, "Community event page"),
    ("Privacy/Security Policy", "privacy-security-policy", "Published", 1864, "Required legal page"),
    ("Return And Exchange Policy", "return-and-exchange-policy", "Published", 3724, "Good policy content"),
    ("Shipping Information", "shipping-information", "Published", 9566, "Detailed shipping page"),
    ("Brighton Event", "brighton-event", "Draft", 842, "Unpublished event"),
    ("Brighton Spring Trunk Show", "brighton-spring-trunk-show", "Draft", 874, "Unpublished event"),
    ("Christmas Open House", "christmas-open-house", "Draft", 2564, "Seasonal — publish when relevant"),
    ("Cyber Monday", "cyber-monday", "Draft", 2882, "Seasonal — publish when relevant"),
    ("GIVE BACK FRIDAY", "give-back-friday", "Draft", 27249, "Massive content — worth publishing?"),
    ("Community Sharing Fashion Show", "community-sharing-spring-fashion-show", "Draft", 1780, "Event content"),
    ("Pink Friday", "pink-friday", "Draft", 3402, "Event content"),
    ("Shop Small Saturday", "shop-small-saturday", "Draft", 5240, "Event content"),
    ("The Clothing Cove Rewards", "the-clothing-cove-rewards", "Draft", 0, "EMPTY — no content at all"),
]

for i, (title, handle, status, length, notes) in enumerate(pages):
    r = i + 2
    ws5.cell(row=r, column=1, value=title)
    ws5.cell(row=r, column=2, value=handle)
    s = ws5.cell(row=r, column=3, value=status)
    s.alignment = CENTER
    s.font = GREEN if status == "Published" else ORANGE
    ws5.cell(row=r, column=4, value=length).number_format = "#,##0"
    ws5.cell(row=r, column=4).alignment = CENTER
    ws5.cell(row=r, column=5, value=notes).alignment = WRAP
    data_row(ws5, r, len(p_headers), alt=(i % 2 == 1))

blog_row = len(pages) + 4
ws5.cell(row=blog_row, column=1, value="BLOG").font = SUBTITLE_FONT
ws5.cell(row=blog_row, column=1).fill = SECTION_FILL
for c in range(2, 6):
    ws5.cell(row=blog_row, column=c).fill = SECTION_FILL
ws5.cell(row=blog_row + 1, column=1, value="Blog name:").font = BOLD
ws5.cell(row=blog_row + 1, column=2, value="Blog").font = BODY
ws5.cell(row=blog_row + 2, column=1, value="Total articles:").font = BOLD
ws5.cell(row=blog_row + 2, column=2, value=41).font = BODY
ws5.cell(row=blog_row + 3, column=1, value="Status:").font = BOLD
ws5.cell(row=blog_row + 3, column=2, value="Content exists but needs audit for freshness, keyword targeting, and internal linking").font = BODY
ws5.merge_cells(start_row=blog_row + 3, start_column=2, end_row=blog_row + 3, end_column=5)


# ════════════════════════════════════════════════════════════
# Sheet 6: Local SEO & Citations
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Local SEO")
ws6.sheet_properties.tabColor = "16A085"

l_headers = ["Platform", "Status", "URL / Details", "Action Needed"]
l_widths = [22, 12, 50, 40]
for i, (h, w) in enumerate(zip(l_headers, l_widths), 1):
    ws6.cell(row=1, column=i, value=h)
    ws6.column_dimensions[get_column_letter(i)].width = w
header_row(ws6, 1, len(l_headers))

citations = [
    ("Google (branded search)", "Strong", "Ranks #1 for 'The Clothing Cove Milford MI'", "Maintain"),
    ("Google (category search)", "Strong", "Ranks #1 for 'dress shop Milford MI'", "Maintain"),
    ("Brighton.com listing", "Present", "brighton.com/pages/store-milford-mi-cl9010000", "Verify hours are current"),
    ("Yelp", "Active", "80 reviews, 51 photos — strong presence", "Continue encouraging reviews"),
    ("Facebook (main)", "Active", "facebook.com/TheClothingCove/", "Verify NAP matches website"),
    ("Facebook (MOB specialty)", "Active", "facebook.com/ClothingCoveSpecialOccasion/", "Cross-link to main site"),
    ("Instagram", "Active", "@theclothingcove_", "Add website link in bio, ensure consistent NAP"),
    ("Yellow Pages", "Present", "yellowpages.com/milford-mi/mip/clothing-cove-2179174", "Verify accuracy"),
    ("LinkedIn", "Present", "linkedin.com/company/the-clothing-cove", "Update if needed"),
    ("Waze", "Present", "Correct address: 414 N Main St", "Maintain"),
    ("Nextdoor", "Present", "nextdoor.com/pages/the-clothing-cove/", "Maintain"),
    ("MeetMeInMilford.com", "Listed", "meetmeinmilford.com directory", "Ensure listing is complete"),
    ("BBB (Better Business Bureau)", "Not found", "Not appearing in BBB search", "Consider creating a BBB profile"),
    ("Google Business Profile", "Not verified", "Could not verify — check in Shopify admin", "Verify GBP is claimed and optimized"),
]

for i, (platform, status, detail, action) in enumerate(citations):
    r = i + 2
    ws6.cell(row=r, column=1, value=platform).font = BOLD
    s = ws6.cell(row=r, column=2, value=status)
    s.alignment = CENTER
    if status in ("Strong", "Active", "Present", "Listed", "Added"):
        s.font = GREEN
    elif "Not" in status:
        s.font = RED
    else:
        s.font = ORANGE
    ws6.cell(row=r, column=3, value=detail).alignment = WRAP
    ws6.cell(row=r, column=4, value=action).alignment = WRAP
    data_row(ws6, r, len(l_headers), alt=(i % 2 == 1))

nap_row = len(citations) + 4
ws6.cell(row=nap_row, column=1, value="NAP CONSISTENCY CHECK").font = SUBTITLE_FONT
ws6.cell(row=nap_row, column=1).fill = SECTION_FILL
for c in range(2, 5):
    ws6.cell(row=nap_row, column=c).fill = SECTION_FILL
nap_data = [
    ("Business Name", "The Clothing Cove"),
    ("Address", "414 N Main St, Milford, MI 48381"),
    ("Phone", "(248) 685-2500"),
    ("Hours (website)", "Mon-Fri 10-6, Sat 9:30-6, Sun Closed"),
    ("Hours (Brighton.com)", "Mon-Wed 10-6, Thu-Fri 10-8, Sat 9:30-6"),
    ("DISCREPANCY", "Brighton.com shows Thu-Fri until 8pm, website shows 6pm — VERIFY"),
]
for i, (label, val) in enumerate(nap_data):
    r = nap_row + 1 + i
    ws6.cell(row=r, column=1, value=label).font = BOLD if "DISCREPANCY" not in label else RED
    ws6.cell(row=r, column=2, value=val).font = BODY if "DISCREPANCY" not in label else RED
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)


# ════════════════════════════════════════════════════════════
# Sheet 7: AEO Gaps
# ════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("AEO Gaps")
ws7.sheet_properties.tabColor = "C0392B"

a_headers = ["Gap", "Current State", "Severity", "Impact on AI Visibility", "Recommended Fix"]
a_widths = [28, 24, 12, 44, 44]
for i, (h, w) in enumerate(zip(a_headers, a_widths), 1):
    ws7.cell(row=1, column=i, value=h)
    ws7.column_dimensions[get_column_letter(i)].width = w
header_row(ws7, 1, len(a_headers))

aeo_gaps = [
    ("No JSON-LD schema", "Microdata only", "CRITICAL", "AI engines (ChatGPT, Perplexity, Google AI Overviews) strongly prefer JSON-LD. Microdata is parsed but less reliably.", "Add Product, Organization, BreadcrumbList JSON-LD to theme."),
    ("Thin product content", "7,694 products with no description", "CRITICAL", "AI cannot cite or recommend products it can't understand. These products are invisible to AI search.", "Generate descriptions from product attributes — even 2-3 sentences helps."),
    ("No FAQ schema on key pages", "FAQ schema exists on FAQ page", "WARNING", "FAQ schema on product/collection pages would surface answers in AI snippets and People Also Ask.", "Add FAQ schema to top collection pages and product pages with common questions."),
    ("Missing product tags", "60.4% have no tags", "CRITICAL", "Tags feed into content taxonomy that AI uses to understand product relationships and categories.", "Bulk-tag products with category, occasion, material, style attributes."),
    ("No blog content strategy", "41 articles, unclear freshness", "WARNING", "Fresh, topical blog content gets cited by AI when answering shopping queries ('best dresses for X').", "Create content around key shopping queries: 'mother of bride dress guide', 'what to wear to X'."),
    ("Missing Open Graph tags", "Not present", "WARNING", "When AI or users share links, there's no structured preview. Reduces click-through from AI citations.", "Add og:title, og:description, og:image to theme.liquid."),
    ("Product schema incomplete", "Basic microdata", "WARNING", "Missing Review, AggregateRating, offers.availability in structured data reduces rich result eligibility.", "Enhance product schema with review data, availability, shipping details."),
    ("No entity disambiguation", "Store name is generic", "WARNING", "'The Clothing Cove' is a common name. Without strong entity markup, AI may confuse with other stores.", "Strengthen Organization schema with sameAs links, founding date, area served."),
    ("Collection pages thin", "173 of 259 have no description", "CRITICAL", "Collection pages are category-level content — exactly what AI uses to understand what the store sells.", "Write 150-300 word descriptions for top 50 collections covering what's in each."),
    ("No 'About' entity content", "About page exists (2,782 chars)", "OK", "About page provides entity context. Currently decent.", "Enhance with founding year, owner name, community involvement for richer entity graph."),
]

for i, (gap, current, sev, impact, fix) in enumerate(aeo_gaps):
    r = i + 2
    ws7.cell(row=r, column=1, value=gap)
    ws7.cell(row=r, column=2, value=current)
    s = ws7.cell(row=r, column=3, value=sev)
    s.alignment = CENTER
    if sev == "CRITICAL": s.font = RED
    elif sev == "WARNING": s.font = ORANGE
    else: s.font = GREEN
    ws7.cell(row=r, column=4, value=impact).alignment = WRAP
    ws7.cell(row=r, column=5, value=fix).alignment = WRAP
    ws7.row_dimensions[r].height = 50
    data_row(ws7, r, len(a_headers), alt=(i % 2 == 1))


# ════════════════════════════════════════════════════════════
# Sheet 8: Fix Roadmap
# ════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("SEO Fix Roadmap")
ws8.sheet_properties.tabColor = "F39C12"

rm_headers = ["Priority", "Task", "Effort", "Impact", "Products/Pages Affected"]
rm_widths = [10, 60, 14, 12, 22]
for i, (h, w) in enumerate(zip(rm_headers, rm_widths), 1):
    ws8.cell(row=1, column=i, value=h)
    ws8.column_dimensions[get_column_letter(i)].width = w
header_row(ws8, 1, len(rm_headers))

roadmap = [
    ("P1", "Add Open Graph + Twitter Card meta tags to theme.liquid", "1-2 hours", "HIGH", "All pages"),
    ("P1", "Add JSON-LD Product schema to product template (replace/supplement microdata)", "3-4 hours", "HIGH", "All products"),
    ("P1", "Add JSON-LD BreadcrumbList to theme.liquid", "1 hour", "MEDIUM", "All pages"),
    ("P1", "Add preconnect hints for external resources", "30 min", "LOW", "All pages"),
    ("P2", "Write collection descriptions for top 50 collections (by traffic)", "8-12 hours", "HIGH", "50 collections"),
    ("P2", "Bulk-generate SEO title tags for in-stock products missing them", "3-4 hours", "HIGH", "In-stock scope"),
    ("P2", "Bulk-generate product descriptions for 3,648 in-stock products with none", "6-12 hours", "HIGH", "3,648 in-stock"),
    ("P2", "Fix duplicate product titles (append variant info)", "1-2 hours", "MEDIUM", "910 in-stock"),
    ("P3", "Bulk-tag 11,061 in-stock untagged products (category + attributes)", "10-16 hours", "HIGH", "11,061 in-stock"),
    ("P3", "Generate/fix image alt text for 1,314 in-stock images without alt", "1-2 hours", "MEDIUM", "744 in-stock products"),
    ("P3", "Write remaining 123 collection descriptions", "15-20 hours", "MEDIUM", "123 collections"),
    ("P3", "Audit and refresh 41 blog articles for freshness + internal links", "8-12 hours", "MEDIUM", "41 articles"),
    ("P4", "Publish useful draft pages (events, Give Back Friday content)", "2-3 hours", "LOW", "~10 pages"),
    ("P4", "Fix 'Our Favorite Brands' page (35 chars — needs real content)", "1-2 hours", "MEDIUM", "1 page"),
    ("P4", "Verify NAP consistency across all citations (hours discrepancy found)", "1-2 hours", "MEDIUM", "All citations"),
    ("P4", "Enhance Organization schema with sameAs, founding date, area served", "1 hour", "MEDIUM", "Theme"),
    ("P4", "Add FAQ schema to top collection pages", "2-3 hours", "MEDIUM", "10-20 collections"),
    ("P4", "Create BBB profile", "30 min", "LOW", "1 listing"),
]

for i, (pri, task, effort, impact, affected) in enumerate(roadmap):
    r = i + 2
    p = ws8.cell(row=r, column=1, value=pri)
    p.alignment = CENTER
    if pri == "P1": p.font = RED
    elif pri == "P2": p.font = ORANGE
    elif pri == "P3": p.font = Font(name="Arial", bold=True, size=10, color="2980B9")
    else: p.font = GRAY
    ws8.cell(row=r, column=2, value=task).alignment = WRAP
    ws8.cell(row=r, column=3, value=effort).alignment = CENTER
    imp = ws8.cell(row=r, column=4, value=impact)
    imp.alignment = CENTER
    if impact == "HIGH": imp.font = RED
    elif impact == "MEDIUM": imp.font = ORANGE
    else: imp.font = GRAY
    ws8.cell(row=r, column=5, value=affected).alignment = CENTER
    ws8.row_dimensions[r].height = 35
    data_row(ws8, r, len(rm_headers), alt=(i % 2 == 1))

summary_row = len(roadmap) + 4
ws8.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
ws8.cell(row=summary_row, column=1, value="EFFORT ESTIMATE").font = SUBTITLE_FONT
ws8.cell(row=summary_row, column=1).fill = SECTION_FILL
estimates = [
    "WORK SCOPE: In-stock products only (11,818). Full catalog numbers shown for reference.",
    "",
    "P1 (Theme fixes — no product changes): ~6-8 hours",
    "P2 (Bulk product SEO — titles, descriptions, dupes for in-stock): ~12-22 hours",
    "P3 (Deep content — tags, alt text, collections, blog): ~34-50 hours",
    "P4 (Polish — pages, citations, schema enhancements): ~8-12 hours",
    "TOTAL ESTIMATED: ~60-92 hours across all priorities (in-stock scope)",
]
for i, est in enumerate(estimates):
    r = summary_row + 1 + i
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws8.cell(row=r, column=1, value=f"  {est}").font = BOLD if "TOTAL" in est else BODY


out = "A:\\TKBS Marketing - Git\\Shopify-Website-Builder\\The Clothing Cove - SEO & AEO Audit.xlsx"
wb.save(out)
print(f"Saved to: {out}")
