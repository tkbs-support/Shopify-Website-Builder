# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
F="The Clothing Cove - SEO & AEO Plan.xlsx"
B="https://theclothingcove.com/products/"
NAVY=PatternFill("solid",fgColor="1F3864"); SUB=PatternFill("solid",fgColor="2E5496")
BAND=PatternFill("solid",fgColor="FFF2CC"); WHITE=Font(color="FFFFFF",bold=True)
LINK=Font(color="0563C1",underline="single"); TH=Side(style="thin",color="D9D9D9"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
title_=[("100% Cotton Pot Holder & Oven Mitt Set 10095","100-cttn-pot-hldr-oven-mitt-1"),
 ("100% Cotton Printed Tea Towel 10161","100-cotton-printed-tea-towel"),
 ("100% Cotton Solid Tea Towel 10198","100-cttn-solid-tea-towel"),
 ("100% Cotton Striped Apron 10134","100-cotton-striped-apron"),
 ("100% Cotton Tea Towel 10083","100-cotton-tea-towel"),
 ("100% Cotton Throw With Fringes 10110","100-cttn-throw-w-fringes"),
 ("101819 Dream Big Bib Set","101819-dream-big-bib-set"),
 ("101956 Galaxy Milestone Blocks","galaxy-milestone-blocks"),
 ("104374 Stay A While Mug","104374-stay-a-while-mug"),
 ("100% Cotton Solid/Print Tea Towel 10041","100-cttn-solid-print-tea-twl")]
desc_=[("Braza 3500 Strapless Angel","braza-strapless-angel"),
 ("Braza 4310 Cool One Pair Pack","braza-4310-braza-cool"),
 ("Brighton G20030 Tango Square Clock","brighton-g20030-tango-square-clock"),
 ("Sterling CZ Dancing Stone Pendant Necklace","01908-sterling-cz-dancing-stone-butterfly-pendant-necklace-copy"),
 ("1.9 oz Square Wrap Soap","1-9-oz-wrap-soap"),
 ("Woven Infinity Scarf","1004290093-woven-infinity-scarf-give-away"),
 ("107266 Red & Black Buffalo Check Stocking","107266-red-black-buffalo-check-stocking"),
 ("10765 Sterling Silver CZ Dancing Stone Heart Necklace","10765-sterling-silver-cz-dancing-stone-heart-necklace"),
 ("10779 Sterling Silver Dancing Stone Butterfly Necklace","10779-sterling-silver-dancing-stone-butterfly-necklace"),
 ('11" Faux Diamond Icicle Ornament',"11-faux-diamond-icicle-ornament-mtx76355")]
both_=[("111050 Blue Plaid Stocking","111050-blue-plaid-stocking"),
 ("111310 Small Black & White Buffalo Check Stocking","111310-sm-b-w-buffalo-check-stocking"),
 ("111768 Happy Plaid Kitchen Towel","111768-happy-plaid-kitchen-towel"),
 ("112176 Squeeze The Day Inset Box Sign","112716-squeeze-the-day-inset-box-sign"),
 ("12123 Sterling Silver CZ Dancing Stone Heart Necklace","12123-sterling-silver-cz-dancing-stone-heart-necklace"),
 ("14260 Tall Wooden Tophat Snowman With Lights","14260-tall-wooden-tophat-snowman-with-lights"),
 ("15477 Stone Smart Watch Case","15477-stone-smart-watch-case"),
 ("15692 Belt Bag","15692-belt-bag"),
 ("15767 3D Belt Bag","15767-3d-belt-bag"),
 ("14322 Traditional White Large Bell","14323-traditional-red-bell-copy")]
wb=openpyxl.load_workbook(F)
if "SEO Examples" in wb.sheetnames: del wb["SEO Examples"]
idx=wb.sheetnames.index("Product SEO")+1
ws=wb.create_sheet("SEO Examples", idx); ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=4; ws.column_dimensions['B'].width=58; ws.column_dimensions['C'].width=58
ws.cell(1,1,"SEO Examples — live products with missing titles/descriptions").font=Font(size=14,bold=True,color="1F3864")
note=ws.cell(2,1,"All in stock and published today. HOW TO READ THIS: Shopify PRE-FILLS the ‘Page title’ box with the "
  "product name when no custom SEO title has been written — so the box looks populated and shows a character count even "
  "though nothing was written. The tell that it’s the un-written default: the title is the raw product name, SKU and all "
  "(e.g. ‘…10095’) and often starts with ‘100%’ etc. A real SEO title reads like a shopper’s search and never includes an "
  "internal part number. These products have no custom-written title; they run on that default. (Ground truth is the "
  "stored SEO field via Shopify’s API / an SEO crawler — what this audit measured.)")
note.alignment=Alignment(wrap_text=True,vertical="top"); ws.merge_cells("A2:C6")
ws.row_dimensions[2].height=90
r=8
def block(r,heading,rows,issue):
    ws.cell(r,1,heading).font=Font(size=12,bold=True,color="1F3864"); r+=1
    for i,h in enumerate(["#","Product (click to open)","What Google shows now / the gap"],start=1):
        c=ws.cell(r,i,h); c.font=WHITE; c.fill=SUB
    r+=1
    for n,(name,handle) in enumerate(rows,start=1):
        ws.cell(r,1,n).border=BD
        c=ws.cell(r,2,name); c.hyperlink=B+handle; c.font=LINK; c.border=BD; c.alignment=Alignment(wrap_text=True)
        ic=ws.cell(r,3,issue); ic.border=BD; ic.alignment=Alignment(wrap_text=True)
        r+=1
    return r+1
r=block(r,"Missing SEO title (these have a description)",title_,"No title set → Google shows the raw product name as the headline")
r=block(r,"Missing meta description (first 3 have a title)",desc_,"No description → Google writes its own snippet, or shows nothing")
r=block(r,"Missing BOTH — worst case",both_,"No title and no description → Google shows a bare part-number name")
ws.freeze_panes="A8"
wb.save(F)
print("added SEO Examples tab. sheets:", wb.sheetnames)
print("Exec Summary charts intact:", len(wb["Executive Summary"]._charts))
