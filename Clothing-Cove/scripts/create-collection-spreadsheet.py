from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Priority Collections"

header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
header_fill = PatternFill("solid", fgColor="1B2838")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin", color="D9E0E6"),
    right=Side(style="thin", color="D9E0E6"),
    top=Side(style="thin", color="D9E0E6"),
    bottom=Side(style="thin", color="D9E0E6"),
)
tier1_fill = PatternFill("solid", fgColor="E8F5E9")
tier2_fill = PatternFill("solid", fgColor="FFF8E1")
tier3_fill = PatternFill("solid", fgColor="E3F2FD")
no_fill = PatternFill("solid", fgColor="FFCDD2")
yes_fill = PatternFill("solid", fgColor="C8E6C9")
partial_fill = PatternFill("solid", fgColor="FFF9C4")

headers = ["Tier", "Collection URL", "Current Title", "Has Description", "In Navigation", "Proposed SEO Title", "Status"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 35
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 62
ws.column_dimensions["G"].width = 14

data = [
    ("Tier 1", "/clothing", "Clothing", "No", "Yes", "Women's Clothing | The Clothing Cove — Milford, MI"),
    ("Tier 1", "/dresses", "Dresses", "No", "Yes", "Designer Dresses — Casual to Evening | The Clothing Cove"),
    ("Tier 1", "/brighton", "Brighton", "No", "Yes", "Brighton Jewelry & Accessories | The Clothing Cove"),
    ("Tier 1", "/accessories", "Accessories", "Yes", "Yes", "Women's Accessories — Jewelry, Handbags & More | The Clothing Cove"),
    ("Tier 1", "/new-arrivals", "New Products", "Yes", "Homepage only", "New Arrivals — Latest Women's Fashion | The Clothing Cove"),
    ("Tier 1", "/sale", "Sale", "Yes", "Yes", "Sale — Women's Clothing & Accessories | The Clothing Cove"),
    ("Tier 1", "/shop-with-purpose", "Shop With Purpose", "Yes", "Yes", "Shop With Purpose — Ethically Made Fashion | The Clothing Cove"),
    ("Tier 2", "/clothing-tops", "Clothing \\ Tops", "No", "Yes", "Women's Tops — Casual, Dressy & Layering | The Clothing Cove"),
    ("Tier 2", "/clothing-bottoms", "Clothing \\ Bottoms", "No", "Yes", "Women's Pants, Skirts & Bottoms | The Clothing Cove"),
    ("Tier 2", "/clothing-denim", "Clothing \\ Denim", "No", "Yes", "Women's Denim — Jeans, Jackets & Jeggings | The Clothing Cove"),
    ("Tier 2", "/clothing-jumpsuits", "Clothing \\ Jumpsuits", "No", "Yes", "Women's Jumpsuits | The Clothing Cove"),
    ("Tier 2", "/dresses-evening", "Dresses \\ Evening", "No", "Yes", "Evening & Formal Dresses | The Clothing Cove"),
    ("Tier 2", "/dresses-casual", "Dresses \\ Casual", "No", "Yes", "Casual Dresses | The Clothing Cove"),
    ("Tier 2", "/dresses-mother-of-bride", "Dresses \\ Mother Of Bride", "No", "Yes", "Mother of the Bride Dresses | The Clothing Cove"),
    ("Tier 2", "/dresses-pant-sets", "Dresses \\ Pant Sets", "No", "Yes", "Dress Pant Sets | The Clothing Cove"),
    ("Tier 2", "/dresses-plus", "Dresses \\ Plus", "No", "Yes", "Plus Size Dresses | The Clothing Cove"),
    ("Tier 2", "/dresses-petite", "Dresses \\ Petite", "No", "Yes", "Petite Dresses | The Clothing Cove"),
    ("Tier 2", "/brighton-jewelry", "Brighton \\ Jewelry", "No", "Yes", "Brighton Jewelry — Necklaces, Bracelets & Rings | The Clothing Cove"),
    ("Tier 2", "/brighton-handbags", "Brighton \\ Handbags", "No", "Yes", "Brighton Handbags & Wallets | The Clothing Cove"),
    ("Tier 2", "/brighton-collections", "Brighton \\ Collections", "No", "Yes", "Brighton Collections | The Clothing Cove"),
    ("Tier 2", "/accessories-jewelry", "Accessories \\ Jewelry", "No", "Yes", "Women's Jewelry — Fashion & Designer Pieces | The Clothing Cove"),
    ("Tier 2", "/accessories-handbags", "Accessories \\ Handbags", "No", "Yes", "Women's Handbags & Purses | The Clothing Cove"),
    ("Tier 2", "/accessories-footwear", "Accessories \\ Footwear", "No", "Yes", "Women's Shoes & Footwear | The Clothing Cove"),
    ("Tier 3", "/brands-sympli", "Brands \\ Sympli", "No", "Yes", "Sympli Clothing | The Clothing Cove — Milford, MI"),
    ("Tier 3", "/brands-brighton-jewelry-and-accessories", "Brands \\ Brighton Jewelry and Accessories", "No", "Yes", "Brighton Jewelry & Accessories | The Clothing Cove"),
    ("Tier 3", "/brands-joseph-ribkoff", "Brands \\ Joseph Ribkoff", "No", "Yes", "Joseph Ribkoff Dresses & Fashion | The Clothing Cove"),
    ("Tier 3", "/brands-frank-lyman", "Brands \\ Frank Lyman", "No", "Yes", "Frank Lyman Dresses & Separates | The Clothing Cove"),
    ("Tier 3", "/brands-renuar", "Brands \\ Renuar", "Yes", "Yes", "Renuar Perfect Fit Pants & Clothing | The Clothing Cove"),
    ("Tier 3", "/brands-liverpool-jeans", "Brands \\ Liverpool Jeans", "Yes", "Yes", "Liverpool Jeans & Denim | The Clothing Cove"),
    ("Tier 3", "/brands-rachel-marie-designs", "Brands \\ Rachel Marie Designs", "Yes", "Yes", "Rachel Marie Designs Jewelry | The Clothing Cove — Michigan Made"),
    ("Tier 3", "/brands-charlie-b", "Brands \\ Charlie B", "No", "Yes", "Charlie B Clothing | The Clothing Cove"),
    ("Tier 3", "/brands-tribal", "Brands \\ Tribal", "No", "Yes", "Tribal Sportswear | The Clothing Cove"),
    ("Tier 3", "/brands-democracy", "Brands \\ Democracy", "No", "Yes", "Democracy Clothing — Ab-Solution Jeans & More | The Clothing Cove"),
    ("Tier 3", "/brands-liv-by-habitat", "Brands \\ Liv by Habitat", "No", "Yes", "Liv by Habitat — Sustainable Women's Fashion | The Clothing Cove"),
    ("Tier 3", "/brands-damee", "Brands \\ Damee", "No", "Yes", "Damee Jackets & Wearable Art | The Clothing Cove"),
]

for row_idx, (tier, url, title, has_desc, in_nav, proposed) in enumerate(data, 2):
    fill = {"Tier 1": tier1_fill, "Tier 2": tier2_fill, "Tier 3": tier3_fill}[tier]
    for col, val in enumerate([tier, url, title, has_desc, in_nav, proposed, "Pending"], 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.fill = fill
        if col == 4:
            cell.fill = no_fill if val == "No" else yes_fill
        if col == 5 and val == "Homepage only":
            cell.fill = partial_fill

ws.auto_filter.ref = f"A1:G{len(data)+1}"
ws.freeze_panes = "A2"

# --- Summary sheet ---
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 30

summary = [
    ("Collection SEO Plan — The Clothing Cove", "", True, 14),
    ("", "", False, 11),
    ("Total Collections", 541, False, 11),
    ("In Navigation", 253, False, 11),
    ("Not in Navigation (sales/events/orphaned)", 288, False, 11),
    ("Collections with Description", 118, False, 11),
    ("Collections without Description", 423, False, 11),
    ("Empty Rate", "78%", False, 11),
    ("", "", False, 11),
    ("Priority Collections to Optimize", 35, True, 11),
    ("Tier 1 — Top-Level Categories", 7, False, 11),
    ("Tier 2 — Key Subcategories", 16, False, 11),
    ("Tier 3 — Brand Pages", 12, False, 11),
    ("", "", False, 11),
    ("What We Fix Per Collection:", "", True, 11),
    ("1. SEO Title", "Keyword-rich, 50-65 chars", False, 11),
    ("2. Meta Description", "140-155 chars with location & CTA", False, 11),
    ("3. Collection Description", "2-3 sentences with product attributes", False, 11),
    ("", "", False, 11),
    ("IMPORTANT:", "", True, 12),
    ("Collection changes are store-wide", "NOT theme-specific", False, 11),
    ("Changes go live immediately", "Owner approval required", False, 11),
]

for row_idx, (label, val, bold, size) in enumerate(summary, 1):
    c1 = ws2.cell(row=row_idx, column=1, value=label)
    c2 = ws2.cell(row=row_idx, column=2, value=val)
    c1.font = Font(name="Arial", size=size, bold=bold)
    c2.font = Font(name="Arial", size=11)
    if row_idx == 20:
        c1.font = Font(name="Arial", size=12, bold=True, color="B42318")

out = r"a:\TKBS Marketing - Git\Shopify-Website-Builder\The Clothing Cove - Collection SEO Plan.xlsx"
wb.save(out)
print(f"Saved: {out}")
