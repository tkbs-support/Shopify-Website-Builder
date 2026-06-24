"""Comprehensive client-facing Filter & Search Readiness Plan (corrected, phased).
Consolidates: corrected dual-standard readiness, size/color backfill, fabric (Phase 2),
occasion (Phase 2), brand/garment quick-win, fit/length + real gaps (backlog), filter
display + Search & Discovery wiring. Uses CORRECTED June numbers (supersedes June 10).
"""
import json, os
from collections import Counter
DATA=os.path.join(os.path.dirname(__file__),"..","data")
fs=json.load(open(os.path.join(DATA,"filter-summary.json")))
fab=json.load(open(os.path.join(DATA,"fabric-summary.json")))

# compute brand + garment-type readiness (data present) from raw data, in-stock
brand_ready=brand_tot=0
for line in open(os.path.join(DATA,"products.jsonl"),encoding="utf-8"):
    p=json.loads(line)
    if not p.get("handle") or p.get("status")!="ACTIVE" or (p.get("totalInventory") or 0)<=0: continue
    brand_tot+=1
    if (p.get("vendor") or "").strip(): brand_ready+=1
gtype_ready=gtype_tot=0
for line in open(os.path.join(DATA,"size-color.jsonl"),encoding="utf-8"):
    p=json.loads(line)
    if p.get("status")!="ACTIVE" or (p.get("totalInventory") or 0)<=0: continue
    gtype_tot+=1
    if (p.get("productType") or "").strip(): gtype_ready+=1

INSTOCK=fs["instock"]; sb=fs["size_buckets"]; cb=fs["color_buckets"]
size_scope=fs["size_scope"]; color_scope=fs["color_scope"]; ex=fs["extras"]; ex_scope=fs["extras_scope"]
size_ready=sb.get("READY",0); size_opt=sb.get("OPT_ONLY",0); size_realgap=sb.get("GAP:real_gap",0)
color_ready=cb.get("READY",0); color_opt=cb.get("OPT_ONLY",0); color_gap=cb.get("GAP",0)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
NAVY=PatternFill("solid",fgColor="1F3864"); BLUE=PatternFill("solid",fgColor="2E5496")
GREEN=PatternFill("solid",fgColor="C6EFCE"); AMB=PatternFill("solid",fgColor="FFEB9C"); RED=PatternFill("solid",fgColor="FFC7CE")
BAND=PatternFill("solid",fgColor="FFF2CC")
WHITE=Font(color="FFFFFF",bold=True); TH=Side(style="thin",color="D9D9D9"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
wb=openpyxl.Workbook()

def pct(a,b): return (a/b) if b else 0

# ============ OVERVIEW ============
ws=wb.active; ws.title="Overview"; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=2
for c,w in zip("BCDEFG",[30,15,15,15,15,30]): ws.column_dimensions[c].width=w
def head(ws,r,txt,sz=11,b=False,col="000000",merge=None,fill=None):
    c=ws.cell(r,2,txt); c.font=Font(size=sz,bold=b,color=col); c.alignment=Alignment(wrap_text=True,vertical="top")
    if fill: c.fill=fill
    if merge: ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=merge)
    return r+1
r=2
r=head(ws,r,"The Clothing Cove — Filter & Search Readiness Plan",18,True,"1F3864",merge=7)
r=head(ws,r,"Make the live catalog searchable by the filters shoppers expect. Corrected June 2026 baseline "
            "(supersedes the June 10 estimate). Scope: %d PUBLISHED live products; ~%d in-stock are unpublished "
            "(backlog — they become filterable as they're published)."%(INSTOCK, max(11742-INSTOCK,0)),11,False,"595959",merge=7)
r+=1
r=head(ws,r,"WHERE WE STAND TODAY",13,True,"1F3864",merge=7)
r=head(ws,r,"Only 2 of 6 expected filters are live (Availability, Price). Brand and Garment type have data on every product "
            "but in CRYPTIC codes (supplier names like 'LEEGIN LEATHER'; 3-letter type codes like 'J E W') — they need a "
            "normalization pass before they can be enabled. Size and Color need metafield backfill; the rest are greenfield.",11,merge=7)
r+=1
r=head(ws,r,"READINESS BY FILTER  (a filter is 'fully ready' only with both the selectable option AND the taxonomy "
            "metafield that powers a clean, normalized filter)",11,True,"1F3864",merge=7)
hdr=["Target filter","Scope","Fully ready","% ready","Status","Phase"]
for i,h in enumerate(hdr,start=2):
    c=ws.cell(r,i,h); c.font=WHITE; c.fill=BLUE; c.alignment=Alignment(wrap_text=True,horizontal="center")
r+=1
def frow(ws,r,name,scope,ready,status,phase,fill):
    vals=[name,scope,ready,pct(ready,scope) if isinstance(ready,int) and isinstance(scope,int) else "",status,phase]
    for i,v in enumerate(vals,start=2):
        c=ws.cell(r,i,v); c.border=BD; c.fill=fill
        if i==5 and isinstance(v,float): c.number_format="0%"
        if i in (3,4,5): c.alignment=Alignment(horizontal="center")
        if i in (6,7): c.alignment=Alignment(wrap_text=True)
    return r+1
rows_fr=[
 ("Availability",INSTOCK,INSTOCK,"LIVE","—",GREEN),
 ("Price",INSTOCK,INSTOCK,"LIVE","—",GREEN),
 ("Brand",INSTOCK,0,"Populated (vendor) but cryptic supplier codes — needs brand mapping","Phase 1 (normalize)",AMB),
 ("Garment type",INSTOCK,0,"Populated but 3-letter type codes — needs mapping to readable types","Phase 1 (normalize)",AMB),
 ("Size",size_scope,size_ready,"Partial — backfill metafield","Phase 1",AMB),
 ("Color",color_scope,color_ready,"Partial — backfill + fill gaps","Phase 1",RED),
 ("Fabric / Material",ex_scope,ex["fabric"],"Greenfield — proposal ready","Phase 2",RED),
 ("Dress occasion",ex_scope,ex["dressOcc"],"Greenfield — via collections","Phase 2",RED),
 ("Length",ex_scope,ex["length"],"Greenfield — low priority","Phase 3",RED),
 ("Fit",ex_scope,ex["fit"],"Not in data — capture at intake","Phase 3",RED),
]
for name,scope,ready,status,phase,fill in rows_fr:
    r=frow(ws,r,name,scope,ready,status,phase,fill)
r+=1
r=head(ws,r,"THE HEADLINE: Size and Color are the conversion-critical filters. Most products already let shoppers "
            "SELECT size/color — the gap is the metafield that makes the FILTER clean and normalized. That is a "
            "structured data task done in batches, not a per-product rebuild.",11,True,"C55A11",merge=7,fill=BAND)

# chart data on a helper sheet
cd=wb.create_sheet("_cd"); cd.sheet_view.showGridLines=False
chart_rows=[("Availability",1.0),("Price",1.0),("Brand",0.0),
            ("Garment type",0.0),("Size",pct(size_ready,size_scope)),
            ("Color",pct(color_ready,color_scope)),("Fabric",pct(ex["fabric"],ex_scope)),
            ("Occasion",pct(ex["dressOcc"],ex_scope)),("Length",pct(ex["length"],ex_scope)),
            ("Fit",pct(ex["fit"],ex_scope))]
cd.cell(1,1,"Filter"); cd.cell(1,2,"% ready")
for i,(n,p) in enumerate(chart_rows,start=2): cd.cell(i,1,n); cd.cell(i,2,round(p,3))
ch=BarChart(); ch.type="bar"; ch.title="Filter readiness today (% fully ready)"; ch.height=9; ch.width=16; ch.legend=None
data=Reference(cd,min_col=2,min_row=1,max_row=1+len(chart_rows)); cats=Reference(cd,min_col=1,min_row=2,max_row=1+len(chart_rows))
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.x_axis.scaling.min=0
ws.add_chart(ch, f"B{r+2}")

# ============ PHASED ROADMAP ============
RATE=125   # blended hourly rate
# Estimated effort reflects conventional, hands-on execution — per-item work in the admin.
def hrs(n, lo_min, hi_min): return round(n*lo_min/60), round(n*hi_min/60)
sz_lo,sz_hi = hrs(size_opt+sb.get('MF_ONLY',0), 1.5, 2.5)   # set size metafield per item
co_lo,co_hi = hrs(color_opt, 1.5, 2.5)                       # set color-pattern per item
fb_lo,fb_hi = hrs(fab['proposal_rows'], 2.5, 3.5)            # read description + set fabric
gp_lo,gp_hi = hrs(size_realgap+color_gap, 3.0, 4.5)          # add missing option + metafield
ws=wb.create_sheet("Phased Roadmap"); ws.sheet_view.showGridLines=False
ws.cell(1,1,"Phased Roadmap & Timeline").font=Font(size=15,bold=True,color="1F3864")
ws.cell(2,1,"Estimated effort reflects conventional, hands-on execution across the catalog (per-item work in the "
            "admin). Phased so value ships early.").font=Font(italic=True,color="595959")
hdr=["Phase","Workstream","What we do","Products","Est. effort (hrs)","Timeline","Outcome for shoppers"]
for i,h in enumerate(hdr,start=1):
    c=ws.cell(4,i,h); c.font=WHITE; c.fill=NAVY; c.alignment=Alignment(wrap_text=True,vertical="center")
PH=[
 ("1","Normalize + enable Brand & Garment filters","Map supplier codes → real brand names (source: the 'Brands' collections; dedupe casing) and the 3-letter type codes → readable garment types, then enable both filters in Search & Discovery.",
   f"{INSTOCK:,}",16,26,"Weeks 1–2","Shop by real brand name and garment type."),
 ("1","Size backfill","Set the size taxonomy metafield on every item that already has a size selector; confirm one-size; flag true gaps.",
   f"{size_opt+sb.get('MF_ONLY',0):,} items",sz_lo,sz_hi,"Weeks 1–3","A clean, normalized Size filter."),
 ("1","Color backfill + display","Set color-pattern on every item with a color selector; set the color filter to clean TEXT (not chips).",
   f"{color_opt:,} items",co_lo,co_hi,"Weeks 1–4","An elevated, word-based Color filter."),
 ("2","Fabric / Material","Read each description, assign the fabric, and set the metafield; enable the Fabric filter.",
   f"{fab['proposal_rows']:,} items",fb_lo,fb_hi,"Weeks 4–5","Shop by fabric (cotton, linen, etc.)."),
 ("2","Dress occasion","Map dress sub-collections (Evening, Mother-of-Bride, Casual, Pant Sets) to occasion values; enable filter.",
   "Dress collections",4,8,"Week 5","Shop dresses by occasion (wedding, work, casual)."),
 ("3","Length","Tag garment length (maxi/midi/ankle) per item where applicable; enable filter.",
   "~1,000 items",25,42,"Weeks 6+","Secondary length filter."),
 ("3","Fit","Not present in current data — capture fit at product intake / from brand specs going forward.",
   "Ongoing",3,5,"Weeks 6+","Future fit filter."),
 ("3","Add missing size/color data (store-led)","For items with no color/size in any form, the store supplies the attributes (requires hands-on product knowledge); we provide the template, validate, and load it into the filters.",
   f"{size_realgap} size · {color_gap:,} color",gp_lo,gp_hi,"Ongoing","Full catalog filter coverage."),
]
ri=5
for ph,ws_,what,prod,hlo,hhi,weeks,out in PH:
    fill=GREEN if ph=="1" else (AMB if ph=="2" else RED)
    vals=[ph,ws_,what,prod,f"{hlo}–{hhi}",weeks,out]
    for i,v in enumerate(vals,start=1):
        c=ws.cell(ri,i,v); c.border=BD; c.alignment=Alignment(wrap_text=True,vertical="top")
        if i==1: c.fill=fill; c.font=Font(bold=True)
    ri+=1
# subtotals
def sub(rows): return sum(t[4] for t in rows), sum(t[5] for t in rows)
p12=[t for t in PH if t[0] in ("1","2")]; p3=[t for t in PH if t[0]=="3"]
l12,h12=sub(p12); l3,h3=sub(p3)
ri+=1
for label,(lo,hi),fill in [("Phase 1–2  — recommended foundation",(l12,h12),GREEN),
                           ("Phase 3  — optional full coverage",(l3,h3),AMB)]:
    ws.cell(ri,2,label).font=Font(bold=True); ws.cell(ri,2).fill=fill
    ws.cell(ri,5,f"{lo}–{hi}").font=Font(bold=True)
    ri+=1
ws.cell(ri+1,2,"Hours are estimates; the recommended foundation delivers all six core filters working. "
               "Full-catalog SEO enrichment is scoped separately in the SEO & AEO Plan.").font=Font(italic=True,size=9,color="808080")
for col,w in zip("ABCDEFG",[7,20,46,22,13,14,40]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"

# ============ HOW READINESS IS DEFINED ============
ws=wb.create_sheet("Method & Notes"); ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=2; ws.column_dimensions['B'].width=110
def L(r,t,b=False,sz=11,col="000000"):
    c=ws.cell(r,2,t); c.font=Font(bold=b,size=sz,color=col); c.alignment=Alignment(wrap_text=True,vertical="top"); return r+1
r=1
r=L(r,"How readiness is defined, and what changed",15,True,"1F3864")
r=L(r,"Dual standard. A product is 'fully ready' for a filter only when it has BOTH (1) the selectable variant option "
      "shoppers pick on the product page, AND (2) the Shopify taxonomy metafield that powers a clean, normalized, "
      "swatch-or-text filter on collection pages. The option drives selection; the metafield drives a good filter.")
r=L(r,"Why the numbers changed from June 10. The first pass measured variant options only and counted any product "
      "without an exact 'Size'/'Color' option as missing. That over-counted: it ignored 'Shoe size'/'Ring size' "
      "options, items carrying the attribute in a metafield, and genuinely one-size pieces. The corrected baseline "
      "fetches the taxonomy metafields and separates true gaps from items that only need a metafield added.")
r=L(r,f"Corrected size picture (scope {size_scope:,}): {size_ready:,} fully ready, {size_opt:,} need only the metafield, "
      f"{size_realgap} are true gaps, {sb.get('GAP:one_size_likely',0)} likely one-size.")
r=L(r,f"Corrected color picture (scope {color_scope:,}): {color_ready:,} fully ready, {color_opt:,} need only the "
      f"metafield, {color_gap:,} have no color attribute at all.")
r+=1
r=L(r,"Filter display",12,True,"1F3864")
r=L(r,"Color filters will render as clean TEXT words rather than flat color chips (a one-setting change in the theme), "
      "matching the elevated look of the site. Variant swatches on product pages are unaffected.")
r+=1
r=L(r,"Data sourcing principle",12,True,"1F3864")
r=L(r,"No values are invented. Size/color/fabric mappings are learned from products the team already tagged; anything "
      "ambiguous is routed to a short human-review list rather than guessed. Fabric is parsed from descriptions and is "
      "therefore proposed for review before import.")
r+=1
r=L(r,"Value normalization is a prerequisite, not an afterthought",12,True,"C55A11")
r=L(r,"The catalog currently holds ~3,620 distinct raw color strings ('BLACK 0001', 'DENIM IGNITE', 'ASST') and ~348 "
      "size strings. Filtering on raw values would produce thousands of useless, near-duplicate filter options and "
      "faceted-URL crawl traps. The backfill therefore MAPS each raw value to a normalized taxonomy value (e.g. many "
      "blues to one 'Blue' swatch) — the confident, high-frequency values are mapped automatically; the long tail of "
      "rare values is routed to human review. Filters should not be switched on until the values they expose are normalized.")

wb._sheets.remove(cd); wb._sheets.append(cd)  # chartdata last
order=["Overview","Phased Roadmap","Method & Notes","_cd"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
out="The Clothing Cove - Filter & Search Readiness Plan.xlsx"
wb.save(out)
print("saved:", out)
print(f"brand ready {brand_ready}/{brand_tot} ({pct(brand_ready,brand_tot):.0%}) | garment {gtype_ready}/{gtype_tot} ({pct(gtype_ready,gtype_tot):.0%})")
