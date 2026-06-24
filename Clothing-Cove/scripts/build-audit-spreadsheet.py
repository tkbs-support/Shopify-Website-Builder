from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import json
import os

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="34495E")
SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2C3E50")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=11, color="2C3E50")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
ACCENT_FONT = Font(name="Arial", bold=True, size=10, color="E74C3C")
ROW_ALT = PatternFill("solid", fgColor="ECF0F1")
ROW_WHITE = PatternFill("solid", fgColor="FFFFFF")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF3CD")
GREEN_FILL = PatternFill("solid", fgColor="D4EDDA")
RED_FILL = PatternFill("solid", fgColor="F8D7DA")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="BDC3C7")
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def style_data_row(ws, row, cols, alt=False):
    fill = ROW_ALT if alt else ROW_WHITE
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.border = THIN_BORDER


# ── Sheet 1: Summary ──
ws = wb.active
ws.title = "Summary"
ws.sheet_properties.tabColor = "2C3E50"

ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 14

ws.merge_cells("A1:C1")
ws["A1"] = "The Clothing Cove — Inventory Recovery Plan"
ws["A1"].font = Font(name="Arial", bold=True, size=16, color="2C3E50")
ws["A1"].alignment = LEFT

ws["A2"] = "Store: the-clothing-cove-store.myshopify.com"
ws["A2"].font = Font(name="Arial", size=10, color="7F8C8D")
ws["A3"] = "Audit Date: May 2026 | Work scope: In-stock products only (11,818)"
ws["A3"].font = Font(name="Arial", size=10, color="7F8C8D")

row = 5
sections = [
    ("CATALOG OVERVIEW", [
        ("Total active products", 35721, None),
        ("In stock", 11747, "=B7/B6"),
        ("Out of stock", 23974, "=B8/B6"),
        ("Default Title (no options)", 17245, "=B9/B6"),
    ]),
    ("COLOR OPTION STATUS", [
        ("Products WITH Color option", 16783, "=B12/B6"),
        ("Products WITHOUT Color option", 18938, "=B13/B6"),
        ("In-stock WITH color", 5420, "=B14/B7"),
        ("In-stock WITHOUT color", 6327, "=B15/B7"),
    ]),
    ("COLOR OPTION NAMING", [
        ('Option name "Color"', 12747, "=B18/B12"),
        ('Option name "COLOR"', 4019, "=B19/B12"),
        ('Option name "color"', 17, "=B20/B12"),
    ]),
    ("METADATA STATUS", [
        ("Products with shopify.color-pattern metafield", 1374, "3.8%"),
        ("Unique color metaobject references", 405, None),
        ("Unique color option values", 6052, None),
        ("Products with ANY shopify.* metafield", 1514, "4.2%"),
    ]),
    ("IN-STOCK WORK SCOPE (11,818 products)", [
        ("In-stock missing Color option", 6354, "53.8%"),
        ("In-stock missing Size option", 8662, "73.3%"),
        ("In-stock Default Title (no options)", 5789, "49.0%"),
        ("In-stock missing body/description", 3648, "30.9%"),
        ("In-stock missing ALL images", 3967, "33.6%"),
        ("In-stock missing tags", 11061, "93.6%"),
    ]),
]

for section_title, items in sections:
    ws.cell(row=row, column=1, value=section_title).font = SUBTITLE_FONT
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D5DBDB")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="D5DBDB")
    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor="D5DBDB")
    row += 1
    for label, value, pct in items:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = LEFT
        c = ws.cell(row=row, column=2, value=value)
        c.font = BOLD_FONT
        c.alignment = RIGHT
        if isinstance(value, (int, float)):
            c.number_format = "#,##0"
        if pct:
            p = ws.cell(row=row, column=3, value=pct)
            p.font = BODY_FONT
            p.alignment = RIGHT
            if isinstance(pct, str) and pct.startswith("="):
                p.number_format = "0.0%"
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3).border = THIN_BORDER
        row += 1
    row += 1

# ── Sheet 2: By Product Type ──
ws2 = wb.create_sheet("By Product Type")
ws2.sheet_properties.tabColor = "2980B9"

headers = ["Product Type", "Total Products", "In Stock", "Has Color Option", "In Stock + Color", "Color Coverage %", "In-Stock Color Coverage %"]
widths = [16, 16, 12, 18, 18, 18, 22]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    ws2.cell(row=1, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws2, 1, len(headers))

product_type_data = [
    ("J E W", 12798, 3991, 3551, 1056),
    ("H O A", 3880, 1599, 649, 298),
    ("B A G", 2247, 628, 1479, 465),
    ("T O C", 2019, 488, 1798, 448),
    ("A C C", 1911, 739, 827, 311),
    ("J A K", 1270, 251, 1079, 232),
    ("S O M", 905, 408, 893, 403),
    ("C H A", 891, 382, 127, 61),
    ("B A B", 778, 273, 208, 54),
    ("C H H", 707, 205, 99, 40),
    ("D R C", 616, 163, 533, 152),
    ("S O X", 507, 169, 200, 62),
    ("P A M", 485, 175, 474, 174),
    ("K N W", 468, 115, 404, 104),
    ("S H O", 456, 219, 415, 185),
    ("S O C", 456, 208, 446, 202),
    ("A C S", 448, 148, 292, 121),
    ("H A T", 407, 80, 329, 55),
    ("W A L", 393, 92, 221, 36),
    ("T O B", 311, 122, 298, 122),
    ("D R E", 265, 55, 213, 53),
    ("T O P", 263, 50, 233, 50),
    ("P A J", 251, 82, 216, 72),
    ("R D R", 229, 121, 87, 56),
    ("A C H", 219, 132, 62, 36),
    ("S L P", 211, 73, 76, 23),
    ("S U N", 205, 111, 93, 55),
    ("(empty)", 165, 49, 81, 29),
    ("T A N", 151, 51, 145, 51),
    ("T O L", 145, 28, 103, 27),
    ("K N C", 140, 43, 130, 42),
    ("W A T", 119, 43, 17, 5),
    ("M E N", 114, 34, 56, 20),
    ("JEW", 102, 8, 47, 4),
    ("S K S", 93, 38, 84, 33),
    ("P A C", 81, 25, 76, 22),
    ("T O V", 75, 12, 64, 11),
    ("V E S", 62, 19, 61, 18),
    ("P A P", 57, 35, 56, 35),
    ("L E G", 52, 24, 48, 21),
    ("P R M", 51, 24, 0, 0),
    ("J E A", 49, 25, 37, 19),
    ("F O D", 46, 27, 3, 2),
    ("P E T", 43, 16, 7, 7),
    ("S O S", 42, 12, 42, 12),
    ("S H M", 42, 11, 39, 10),
    ("A C B", 41, 20, 20, 8),
    ("J U M", 39, 10, 35, 10),
    ("A C P", 32, 2, 19, 2),
    ("S O W", 28, 9, 28, 9),
    ("K N T", 28, 6, 23, 6),
    ("P O N", 27, 3, 25, 3),
    ("T H T", 27, 14, 27, 14),
    ("S O P", 26, 10, 25, 10),
    ("B R A", 23, 16, 23, 16),
    ("C H C", 23, 4, 7, 2),
    ("T H K", 21, 2, 20, 2),
    ("L U G", 16, 1, 1, 1),
    ("T H J", 16, 4, 13, 3),
    ("H O S", 13, 8, 13, 8),
    ("T H A", 13, 4, 12, 4),
    ("O U J", 12, 0, 9, 0),
    ("S O J", 11, 4, 10, 4),
    ("L I P", 10, 2, 9, 2),
    ("S O B", 9, 3, 9, 3),
    ("S P C", 9, 1, 6, 1),
    ("F R G", 8, 3, 4, 3),
    ("Y O S", 8, 4, 8, 4),
    ("P A W", 8, 3, 8, 3),
    ("T H D", 8, 0, 0, 0),
    ("O U C", 7, 1, 6, 1),
    ("T H B", 5, 3, 4, 2),
    ("L I S", 4, 1, 4, 1),
    ("P A S", 3, 2, 3, 2),
    ("T H V", 3, 0, 2, 0),
    ("T H C", 3, 0, 0, 0),
    ("P S T", 3, 0, 3, 0),
    ("J EW", 3, 0, 3, 0),
    ("A T V", 2, 0, 2, 0),
    ("K N V", 2, 1, 2, 1),
    ("Gift Card", 1, 1, 0, 0),
    ("S H W", 1, 0, 1, 0),
    ("WAT", 1, 1, 0, 0),
    ("THK", 1, 1, 1, 1),
    ("J E W\\", 1, 0, 0, 0),
]

for i, (pt, total, instock, has_color, instock_color) in enumerate(product_type_data):
    r = i + 2
    ws2.cell(row=r, column=1, value=pt).alignment = LEFT
    ws2.cell(row=r, column=2, value=total).number_format = "#,##0"
    ws2.cell(row=r, column=3, value=instock).number_format = "#,##0"
    ws2.cell(row=r, column=4, value=has_color).number_format = "#,##0"
    ws2.cell(row=r, column=5, value=instock_color).number_format = "#,##0"
    ws2.cell(row=r, column=6, value=f"=IF(B{r}=0,\"-\",D{r}/B{r})").number_format = "0.0%"
    ws2.cell(row=r, column=7, value=f"=IF(C{r}=0,\"-\",E{r}/C{r})").number_format = "0.0%"
    style_data_row(ws2, r, len(headers), alt=(i % 2 == 1))

ws2.auto_filter.ref = f"A1:G{len(product_type_data)+1}"

# ── Sheet 3: Top Color Values ──
ws3 = wb.create_sheet("Color Values")
ws3.sheet_properties.tabColor = "27AE60"

color_headers = ["Rank", "Color Value", "Product Count", "% of Color Products"]
color_widths = [8, 24, 16, 20]
for i, (h, w) in enumerate(zip(color_headers, color_widths), 1):
    ws3.cell(row=1, column=i, value=h)
    ws3.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws3, 1, len(color_headers))

color_data = [
    ("BLACK", 3289), ("SILVER", 1112), ("WHITE", 1107), ("NAVY", 1067),
    ("BLUE", 810), ("PINK", 782), ("GREY", 768), ("GOLD", 767),
    ("RED", 704), ("CLEAR", 629), ("BEIGE", 427), ("GREEN", 389),
    ("BROWN", 382), ("IVORY", 340), ("PURPLE", 313), ("MULTI", 296),
    ("DENIM", 232), ("TEAL", 220), ("OLIVE", 216), ("TAUPE", 203),
    ("TURQUOISE", 184), ("YELLOW", 177), ("TAN", 170), ("AQUA", 161),
    ("CHARCOAL", 160), ("ORANGE", 148), ("MINT", 145), ("CHAMPAGNE", 142),
    ("BURGUNDY", 142), ("CREAM", 140), ("CORAL", 131), ("PEWTER", 128),
    ("AURORA BOREALIS", 126), ("SILVER NIGHT", 121), ("FUCHSIA", 114),
    ("LAVENDER", 114), ("WINE", 112), ("BLUSH", 109), ("KHAKI", 107),
    ("CAMEL", 105), ("ROYAL", 102), ("BLACK & WHITE", 98),
    ("BLACK/WHITE", 98), ("LIGHT BLUE", 90), ("ROSE", 89),
    ("ROYAL BLUE", 87), ("OFF WHITE", 85), ("MIDNIGHT", 84),
    ("ROSE GOLD", 79), ("INDIGO", 75),
]

for i, (color, count) in enumerate(color_data):
    r = i + 2
    ws3.cell(row=r, column=1, value=i + 1).alignment = CENTER
    ws3.cell(row=r, column=2, value=color).alignment = LEFT
    ws3.cell(row=r, column=3, value=count).number_format = "#,##0"
    ws3.cell(row=r, column=4, value=f"=C{r}/16783").number_format = "0.0%"
    style_data_row(ws3, r, len(color_headers), alt=(i % 2 == 1))

note_row = len(color_data) + 3
ws3.cell(row=note_row, column=1, value="Note:").font = BOLD_FONT
ws3.cell(row=note_row, column=2, value="6,052 total unique color values. Top 50 shown above. Full list in color-audit-results.json").font = BODY_FONT

ws3.auto_filter.ref = f"A1:D{len(color_data)+1}"

# ── Sheet 4: Metafield Status ──
ws4 = wb.create_sheet("Metafield Status")
ws4.sheet_properties.tabColor = "8E44AD"

mf_headers = ["Metafield Key", "Products (Full Catalog)", "Confirmed Count", "Coverage %"]
mf_widths = [30, 28, 20, 14]
for i, (h, w) in enumerate(zip(mf_headers, mf_widths), 1):
    ws4.cell(row=1, column=i, value=h)
    ws4.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws4, 1, len(mf_headers))

mf_data = [
    ("shopify.color-pattern", 1374, 1374, "3.8%"),
    ("shopify.size", 934, 934, "2.6%"),
    ("shopify.age-group", 407, 407, "1.1%"),
    ("shopify.fabric", 379, 379, "1.1%"),
    ("shopify.target-gender", 377, 377, "1.1%"),
    ("shopify.neckline", 221, 221, "0.6%"),
    ("shopify.sleeve-length-type", 207, 207, "0.6%"),
    ("shopify.top-length-type", 145, 145, "0.4%"),
]

for i, (key, found, est, pct) in enumerate(mf_data):
    r = i + 2
    ws4.cell(row=r, column=1, value=key).alignment = LEFT
    ws4.cell(row=r, column=2, value=found).number_format = "#,##0"
    ws4.cell(row=r, column=2).alignment = CENTER
    ws4.cell(row=r, column=3, value=est).number_format = "#,##0"
    ws4.cell(row=r, column=3).alignment = CENTER
    ws4.cell(row=r, column=4, value=pct).alignment = CENTER
    style_data_row(ws4, r, len(mf_headers), alt=(i % 2 == 1))

note_start = len(mf_data) + 3
notes = [
    "KEY FINDINGS:",
    "- 1,374 products have shopify.color-pattern — referencing 405 unique color metaobjects",
    "- Heavily concentrated in newest products (recent tagging initiative)",
    "- API token lacks read_metaobjects scope — cannot read metaobject contents",
    "- 1,514 total products (4.2%) have ANY Shopify standard metafield",
    "- The other 15,409 products with color use variant OPTIONS only (not metafields)",
    "",
    "OTHER METAFIELDS (present on most products but NOT color-related):",
    "- global.description_tag: ~98% of products (SEO meta descriptions)",
    "- mc-facebook.google_product_category: ~76% (Google Shopping categories)",
    "- global.title_tag: ~74% (SEO title overrides)",
    "- SEOMetaManager.config: ~69% (SEO app configuration)",
    "- seo.hidden: ~46% (SEO visibility flags)",
]
for i, note in enumerate(notes):
    cell = ws4.cell(row=note_start + i, column=1, value=note)
    cell.font = BOLD_FONT if note.endswith(":") else BODY_FONT
    ws4.merge_cells(start_row=note_start + i, start_column=1, end_row=note_start + i, end_column=4)

# ── Sheet 5: Fix Roadmap ──
ws5 = wb.create_sheet("Fix Roadmap")
ws5.sheet_properties.tabColor = "E67E22"

road_headers = ["Phase", "Description", "Effort", "Risk", "Products Modified"]
road_widths = [10, 60, 14, 28, 20]
for i, (h, w) in enumerate(zip(road_headers, road_widths), 1):
    ws5.cell(row=1, column=i, value=h)
    ws5.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws5, 1, len(road_headers))

roadmap = [
    ("Phase 1", "Activate Color Swatches — populate theme color_swatch_config with hex codes for all existing color option values. Makes filters show colored squares instead of text.", "2-3 hours", "None — theme display setting only, easily reversible", "0"),
    ("Phase 2", "Standardize Option Names — rename \"COLOR\" (4,019) and \"color\" (17) to \"Color\" so Search & Discovery treats them as one unified filter.", "1-2 hours", "Low — cosmetic option label change only, does not affect inventory or variants", "4,036"),
    ("Phase 3", "Improve Swatch Appearance — evaluate default Prestige theme swatch CSS and make adjustments for shape, size, spacing, mobile tap targets.", "1-2 hours", "None — cosmetic CSS changes tested on dev theme first", "0"),
    ("Phase 4", "Add Color Data to Products — add Color options to in-stock products that currently lack them. Priority: clothing types (T O C, J A K, D R C, etc.)", "1-4 weeks", "Medium — modifies product data. Use dry-run testing. Can be done incrementally by product type.", "TBD (up to 6,327 in-stock)"),
    ("Phase 5", "Search & Discovery Optimization — verify/configure the Color filter in S&D app. Ensure proper sort order and display.", "30 min", "None", "0"),
    ("Phase 6", "Publish to Live — apply swatch config from dev theme (145504469107) to live theme (138024386675). Verify on production.", "30 min", "Low — all changes tested on dev theme first", "0"),
]

for i, (phase, desc, effort, risk, modified) in enumerate(roadmap):
    r = i + 2
    ws5.cell(row=r, column=1, value=phase).font = BOLD_FONT
    ws5.cell(row=r, column=1).alignment = CENTER
    ws5.cell(row=r, column=2, value=desc).alignment = WRAP
    ws5.cell(row=r, column=3, value=effort).alignment = CENTER
    ws5.cell(row=r, column=4, value=risk).alignment = WRAP
    ws5.cell(row=r, column=5, value=modified).alignment = CENTER
    ws5.row_dimensions[r].height = 45
    style_data_row(ws5, r, len(road_headers), alt=(i % 2 == 1))

rec_row = len(roadmap) + 4
ws5.merge_cells(start_row=rec_row, start_column=1, end_row=rec_row, end_column=5)
ws5.cell(row=rec_row, column=1, value="RECOMMENDATION").font = SUBTITLE_FONT
ws5.cell(row=rec_row, column=1).fill = PatternFill("solid", fgColor="D5DBDB")

recs = [
    "Start with Phases 1-3 (quick wins, ~1 day total, no product data risk) — immediately makes existing color and size data functional and visual.",
    "Evaluate Phase 4 after seeing Phases 1-3 live — the built-in Prestige styling may be sufficient.",
    "Scope Phase 5 based on priorities — focus on clothing/fashion product types first (T O C, J A K, D R C), exclude Home (H O A) and Food (F O D).",
    "Phase 5 can be done incrementally, one product type at a time, with review between batches.",
    "1,374 products already have shopify.color-pattern metafields. Adding read_metaobjects API scope would let us audit what color data those contain.",
]
for i, rec in enumerate(recs):
    r = rec_row + 1 + i
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws5.cell(row=r, column=1, value=f"  {i+1}. {rec}").font = BODY_FONT


# ── Sheet 6: Size Analysis ──
ws6 = wb.create_sheet("Size Analysis")
ws6.sheet_properties.tabColor = "16A085"

ws6.merge_cells("A1:D1")
ws6["A1"] = "Size Option Analysis"
ws6["A1"].font = TITLE_FONT

ws6["A3"] = "Size Option Overview"
ws6["A3"].font = SUBTITLE_FONT
ws6["A3"].fill = PatternFill("solid", fgColor="D5DBDB")
ws6["B3"].fill = PatternFill("solid", fgColor="D5DBDB")
ws6["C3"].fill = PatternFill("solid", fgColor="D5DBDB")

size_overview = [
    ("Products with Size option", 9902, None),
    ("In-stock with Size", 3129, None),
    ('Option name "Size"', 7097, "71.7%"),
    ('Option name "SIZE"', 2805, "28.3%"),
    ("Unique size values", 479, None),
]
for i, (label, val, pct) in enumerate(size_overview):
    r = 4 + i
    ws6.cell(row=r, column=1, value=label).font = BODY_FONT
    c = ws6.cell(row=r, column=2, value=val)
    c.font = BOLD_FONT
    c.number_format = "#,##0"
    if pct:
        ws6.cell(row=r, column=3, value=pct).font = BODY_FONT
    ws6.cell(row=r, column=1).border = THIN_BORDER
    ws6.cell(row=r, column=2).border = THIN_BORDER
    ws6.cell(row=r, column=3).border = THIN_BORDER
ws6.column_dimensions["A"].width = 32
ws6.column_dimensions["B"].width = 14
ws6.column_dimensions["C"].width = 12
ws6.column_dimensions["D"].width = 14

size_val_start = 11
ws6.cell(row=size_val_start, column=1, value="Size Value").font = HEADER_FONT
ws6.cell(row=size_val_start, column=1).fill = HEADER_FILL
ws6.cell(row=size_val_start, column=2, value="Products").font = HEADER_FONT
ws6.cell(row=size_val_start, column=2).fill = HEADER_FILL
ws6.cell(row=size_val_start, column=3, value="Category").font = HEADER_FONT
ws6.cell(row=size_val_start, column=3).fill = HEADER_FILL

size_values = [
    ("L", 5124, "Letter"), ("M", 5123, "Letter"), ("S", 4955, "Letter"),
    ("XL", 4272, "Letter"), ("10", 1537, "Numeric"), ("8", 1345, "Numeric"),
    ("12", 1281, "Numeric"), ("14", 1084, "Numeric"), ("6", 1059, "Numeric"),
    ("16", 770, "Numeric"), ("XS", 701, "Letter"), ("XXL", 635, "Letter"),
    ("4", 524, "Numeric"), ("S/M", 437, "Combo"), ("L/XL", 373, "Combo"),
    ("7", 332, "Numeric"), ("9", 328, "Numeric"), ("2", 324, "Numeric"),
    ("18", 276, "Numeric"), ("ONE SIZE", 268, "Universal"),
    ("38", 202, "European"), ("PL", 202, "Petite"), ("1X", 202, "Plus"),
    ("PS", 197, "Petite"), ("PM", 197, "Petite"),
]

for i, (val, count, cat) in enumerate(size_values):
    r = size_val_start + 1 + i
    ws6.cell(row=r, column=1, value=val).alignment = LEFT
    ws6.cell(row=r, column=2, value=count).number_format = "#,##0"
    ws6.cell(row=r, column=3, value=cat).alignment = CENTER
    style_data_row(ws6, r, 3, alt=(i % 2 == 1))

cat_note = size_val_start + len(size_values) + 2
ws6.cell(row=cat_note, column=1, value="SIZE VALUE CATEGORIES (479 unique values)").font = SUBTITLE_FONT
cats = [
    ("Letter sizes (S, M, L, XL, XXL, XS)", "Most common — standard clothing"),
    ("Numeric sizes (2, 4, 6, 8, 10, 12, 14, 16, 18)", "Dresses, pants, formal wear"),
    ("European sizes (36, 37, 38, 39, 40, 41, 42)", "Shoes and some imported clothing"),
    ("Petite (PS, PM, PL)", "Petite line sizing"),
    ("Plus (1X, 2X, 3X)", "Extended sizes"),
    ("Combo (S/M, L/XL)", "Two-size ranges"),
    ("Universal (ONE SIZE, O/S, OSFM)", "One-size-fits-most items"),
    ("Shoe sizes (6, 6.5, 7, 7.5, 8, ...)", "Overlaps with numeric — context-dependent"),
]
for i, (cat, desc) in enumerate(cats):
    r = cat_note + 1 + i
    ws6.cell(row=r, column=1, value=cat).font = BOLD_FONT
    ws6.cell(row=r, column=2, value=desc).font = BODY_FONT
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

# Size by product type
pt_note = cat_note + len(cats) + 2
ws6.cell(row=pt_note, column=1, value="Size by Product Type").font = SUBTITLE_FONT
ws6.cell(row=pt_note, column=1).fill = PatternFill("solid", fgColor="D5DBDB")
ws6.cell(row=pt_note, column=2).fill = PatternFill("solid", fgColor="D5DBDB")
ws6.cell(row=pt_note, column=3).fill = PatternFill("solid", fgColor="D5DBDB")

size_pt = [
    ("T O C (Tops/Clothing)", 1941), ("J A K (Jackets)", 1221),
    ("S O M", 895), ("D R C (Dresses)", 600),
    ("P A M", 449), ("S O C", 447), ("S H O (Shoes)", 431),
    ("K N W", 401), ("T O B", 296), ("D R E", 223),
]
for i, (pt, count) in enumerate(size_pt):
    r = pt_note + 1 + i
    ws6.cell(row=r, column=1, value=pt).font = BODY_FONT
    ws6.cell(row=r, column=2, value=count).number_format = "#,##0"
    ws6.cell(row=r, column=2).font = BOLD_FONT
    ws6.cell(row=r, column=1).border = THIN_BORDER
    ws6.cell(row=r, column=2).border = THIN_BORDER


# ── Sheet 7: Data Quality Issues ──
ws7 = wb.create_sheet("Data Quality")
ws7.sheet_properties.tabColor = "C0392B"

ws7.merge_cells("A1:E1")
ws7["A1"] = "Data Quality Issues & Cleanup Tasks"
ws7["A1"].font = TITLE_FONT

dq_headers = ["Issue", "Option", "Current Value", "Should Be", "Products Affected"]
dq_widths = [30, 12, 22, 16, 18]
for i, (h, w) in enumerate(zip(dq_headers, dq_widths), 1):
    ws7.cell(row=3, column=i, value=h)
    ws7.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws7, 3, len(dq_headers))

dq_issues = [
    ("Inconsistent casing", "Color", '"COLOR"', '"Color"', "4,019"),
    ("Inconsistent casing", "Color", '"color"', '"Color"', "17"),
    ("Inconsistent casing", "Size", '"SIZE"', '"Size"', "2,805"),
    ("Trailing space in name", "Color", '"COLOR " (trailing space)', '"Color"', "11"),
    ("Trailing space in name", "Color", '"Color " (trailing space)', '"Color"', "2"),
    ("Trailing space in name", "Color", '"color " (trailing space)', '"Color"', "1"),
    ("Trailing space in name", "Size", '"SIZE " (trailing space)', '"Size"', "3"),
    ("Trailing space in name", "Style", '"STYLE " (trailing space)', '"Style"', "5"),
    ("Trailing space in name", "Style", '"style " (trailing space)', '"Style"', "1"),
    ("Typo / misspelling", "Color", '"OCLOR"', '"Color"', "1"),
    ("Typo / misspelling", "Color", '"Coilor"', '"Color"', "1"),
    ("Typo / misspelling", "Color", '"COOR"', '"Color"', "1"),
    ("Typo / misspelling", "Size", '"S1ze" (one instead of i)', '"Size"', "1"),
    ("Typo / misspelling", "Size", '"Siz"', '"Size"', "1"),
    ("Should merge into Color", "Colors", '"COLORS"', '"Color"', "7"),
    ("Should merge into Color", "Color & Size", '"COLOR & SIZE"', 'Split to Color + Size', "1"),
    ("Should merge into Color", "Style/Color", '"Style/Color"', 'Split to Style + Color', "1"),
    ("Not a real option", "N/A", '"Download Our App Now!"', "Remove", "1"),
    ("Not a real option", "N/A", '"polyester & spandex"', "Remove or move to metafield", "1"),
    ("Not a real option", "N/A", '"poly spandex"', "Remove or move to metafield", "1"),
    ("Not a real option", "N/A", '"Multi"', "Review — may be Color value", "1"),
    ("Not a real option", "N/A", '"BLACK" (as option name)', "Should be Color", "2"),
    ("Not a real option", "N/A", '"MATERIAL"', "Move to metafield", "1"),
    ("Ambiguous option", "N/A", '"MEN\'S ACCESSORY"', "Review", "1"),
    ("Ambiguous option", "N/A", '"Bracelet design"', "Review", "1"),
    ("Ambiguous option", "N/A", '"LENGTH"', "Review", "1"),
]

for i, (issue, opt, current, should_be, affected) in enumerate(dq_issues):
    r = i + 4
    ws7.cell(row=r, column=1, value=issue).alignment = LEFT
    ws7.cell(row=r, column=2, value=opt).alignment = CENTER
    ws7.cell(row=r, column=3, value=current).alignment = LEFT
    ws7.cell(row=r, column=4, value=should_be).alignment = LEFT
    ws7.cell(row=r, column=5, value=affected).alignment = CENTER
    style_data_row(ws7, r, len(dq_headers), alt=(i % 2 == 1))
    if "Typo" in issue or "Not a real" in issue:
        ws7.cell(row=r, column=1).font = Font(name="Arial", size=10, color="C0392B")

dq_summary_row = len(dq_issues) + 6
ws7.cell(row=dq_summary_row, column=1, value="SUMMARY").font = SUBTITLE_FONT
ws7.cell(row=dq_summary_row, column=1).fill = PatternFill("solid", fgColor="D5DBDB")
for c in range(2, 6):
    ws7.cell(row=dq_summary_row, column=c).fill = PatternFill("solid", fgColor="D5DBDB")

dq_summaries = [
    "Total products with naming inconsistencies (casing): ~6,841 (COLOR: 4,019 + SIZE: 2,805 + color: 17)",
    "Total products with trailing spaces: ~22",
    "Total products with typos/misspellings: ~5",
    "Total products with invalid/junk options: ~8",
    'All naming fixes are low-risk — they only change the option LABEL, not values, inventory, or variants.',
    "Recommended: fix all naming issues in one batch before turning on filters, so Search & Discovery sees clean data.",
]
for i, s in enumerate(dq_summaries):
    r = dq_summary_row + 1 + i
    ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws7.cell(row=r, column=1, value=f"  {s}").font = BODY_FONT


# ── Sheet 8: All Option Types ──
ws8 = wb.create_sheet("All Option Types")
ws8.sheet_properties.tabColor = "7F8C8D"

ws8.merge_cells("A1:F1")
ws8["A1"] = "Complete Inventory of Product Option Types"
ws8["A1"].font = TITLE_FONT
ws8["A2"] = f"35,749 active products scanned | 17,253 Default Title (no options) | 18,496 with options"
ws8["A2"].font = Font(name="Arial", size=10, color="7F8C8D")

ao_headers = ["Option Name", "Products", "In Stock", "Unique Values", "Naming Variants", "Notes"]
ao_widths = [18, 12, 12, 14, 36, 30]
for i, (h, w) in enumerate(zip(ao_headers, ao_widths), 1):
    ws8.cell(row=4, column=i, value=h)
    ws8.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws8, 4, len(ao_headers))

all_options = [
    ("color", 16798, 5431, 6052, 'Color:12762, COLOR:4019, color:17', "PRIMARY — needs swatch config + naming fix"),
    ("size", 9902, 3129, 479, 'Size:7097, SIZE:2805', "PRIMARY — needs naming fix"),
    ("style", 214, 105, 646, 'STYLE:116, Style:98', "Minor — mostly H O A and J E W"),
    ("scent", 27, 24, 51, 'SCENT:24, Scent:3', "Niche — home goods and candles"),
    ("shoe size", 24, 23, 17, 'Shoe size:24', "Clean — shoes only"),
    ('color (trailing space)', 14, 2, 16, 'COLOR :11, Color :2, color :1', "BUG — trailing space, merge into Color"),
    ("colors", 7, 5, 18, 'COLORS:7', "BUG — merge into Color"),
    ('style (trailing space)', 6, 5, 19, 'STYLE :5, style :1', "BUG — trailing space, merge into Style"),
    ("print", 6, 3, 20, 'PRINT:4, Print:2', "Niche — baby/kids patterns"),
    ("strength", 5, 3, 14, 'STRENGTH:5', "Reading glasses magnification"),
    ("ring size", 4, 3, 4, 'Ring size:4', "Niche — jewelry"),
    ('size (trailing space)', 3, 1, 8, 'SIZE :3', "BUG — trailing space, merge into Size"),
    ("fragrance", 2, 1, 6, 'FRAGRANCE:2', "Niche"),
    ("black", 2, 0, 14, 'BLACK:2', "BUG — color value used as option name"),
    ("initial", 1, 1, 21, 'INITIAL:1', "Niche — monogram jewelry"),
    ("oclor", 1, 0, 1, 'OCLOR:1', "TYPO — should be Color"),
    ("coilor", 1, 1, 1, 'Coilor:1', "TYPO — should be Color"),
    ("coor", 1, 1, 4, 'COOR:1', "TYPO — should be Color"),
    ("s1ze", 1, 0, 3, 'S1ze:1', "TYPO — should be Size"),
    ("siz", 1, 0, 2, 'Siz:1', "TYPO — should be Size"),
    ("material", 1, 0, 1, 'MATERIAL:1', "Should be a metafield, not variant option"),
    ("flavor", 1, 1, 4, 'Flavor:1', "Niche — food"),
    ("lip balm tint", 1, 1, 5, 'LIP BALM TINT:1', "Niche"),
    ("design", 1, 1, 7, 'DESIGN:1', "Niche"),
    ("download our app now!", 1, 0, 1, 'Download Our App Now!:1', "JUNK — not a real option"),
]

for i, (name, prods, stock, vals, casings, notes) in enumerate(all_options):
    r = i + 5
    ws8.cell(row=r, column=1, value=name).alignment = LEFT
    ws8.cell(row=r, column=2, value=prods).number_format = "#,##0"
    ws8.cell(row=r, column=3, value=stock).number_format = "#,##0"
    ws8.cell(row=r, column=4, value=vals).number_format = "#,##0"
    ws8.cell(row=r, column=5, value=casings).alignment = LEFT
    ws8.cell(row=r, column=5).font = Font(name="Arial", size=9)
    ws8.cell(row=r, column=6, value=notes).alignment = WRAP
    style_data_row(ws8, r, len(ao_headers), alt=(i % 2 == 1))
    if "BUG" in notes or "TYPO" in notes or "JUNK" in notes:
        ws8.cell(row=r, column=6).font = Font(name="Arial", size=10, color="C0392B")
    elif "PRIMARY" in notes:
        ws8.cell(row=r, column=6).font = Font(name="Arial", bold=True, size=10, color="27AE60")


# ── Update Summary sheet with V2 data ──
# Add size stats to summary
row = 27
ws.cell(row=row, column=1, value="SIZE OPTION STATUS").font = SUBTITLE_FONT
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D5DBDB")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="D5DBDB")
ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor="D5DBDB")
row += 1
size_stats = [
    ("Products WITH Size option", 9902, None),
    ("In-stock WITH size", 3129, None),
    ('Option name "Size"', 7097, "71.7%"),
    ('Option name "SIZE"', 2805, "28.3%"),
    ("Unique size values", 479, None),
]
for label, value, pct in size_stats:
    ws.cell(row=row, column=1, value=label).font = BODY_FONT
    c = ws.cell(row=row, column=2, value=value)
    c.font = BOLD_FONT
    c.number_format = "#,##0"
    if pct:
        ws.cell(row=row, column=3, value=pct).font = BODY_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3).border = THIN_BORDER
    row += 1

row += 1
ws.cell(row=row, column=1, value="DATA QUALITY").font = SUBTITLE_FONT
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D5DBDB")
ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="D5DBDB")
ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor="D5DBDB")
row += 1
dq_stats = [
    ("Products with naming inconsistencies", "~6,841", None),
    ("Products with trailing spaces in option name", "~22", None),
    ("Products with typos in option name", "~5", None),
    ("Products with junk/invalid options", "~8", None),
    ("Default Title products (no options)", 17253, None),
]
for label, value, _ in dq_stats:
    ws.cell(row=row, column=1, value=label).font = BODY_FONT
    c = ws.cell(row=row, column=2, value=value)
    c.font = BOLD_FONT
    if isinstance(value, int):
        c.number_format = "#,##0"
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1


# ── Update Roadmap to V2 (comprehensive) ──
# Unmerge any merged cells in the roadmap area first
for merge in list(ws5.merged_cells.ranges):
    ws5.unmerge_cells(str(merge))
for r in range(2, 20):
    for c in range(1, 6):
        ws5.cell(row=r, column=c).value = None

roadmap_v2 = [
    ("Phase 1", "Activate Color Swatches — populate theme color_swatch_config with hex codes for all existing color values. Makes filters show colored squares instead of text.", "2-3 hours", "None — theme display setting only, easily reversible", "0"),
    ("Phase 2", "Data Quality Cleanup — fix all option naming issues: rename COLOR→Color (4,019), SIZE→Size (2,805), fix trailing spaces (22), typos (5), merge COLORS→Color (7). One batch.", "2-3 hours", "Low — only changes option labels. Does not affect inventory, variants, or pricing.", "~6,876"),
    ("Phase 3", "Search & Discovery Configuration — configure Color and Size filters in S&D app. Verify unified filter groups after naming cleanup.", "30 min", "None", "0"),
    ("Phase 4", "Improve Swatch Appearance — evaluate default Prestige theme swatch CSS. Adjust shape, size, spacing if needed.", "1-2 hours", "None — cosmetic CSS changes tested on dev theme first", "0"),
    ("Phase 5", "Add Missing Product Data — add Color/Size options to in-stock products that currently lack them. Priority by product type.", "2-6 weeks", "Medium — modifies product data. Incremental by product type, dry-run first.", "TBD"),
    ("Phase 6", "Metafield Expansion — continue the shopify.color-pattern tagging initiative (1,374 started). Extend to size, fabric, gender.", "Ongoing", "Low — additive metadata, doesn't change existing data.", "TBD"),
    ("Phase 7", "Publish to Live — apply all theme changes from dev (145504469107) to live (138024386675).", "30 min", "Low — all changes tested on dev theme first", "0"),
]

for i, (phase, desc, effort, risk, modified) in enumerate(roadmap_v2):
    r = i + 2
    ws5.cell(row=r, column=1, value=phase).font = BOLD_FONT
    ws5.cell(row=r, column=1).alignment = CENTER
    ws5.cell(row=r, column=2, value=desc).alignment = WRAP
    ws5.cell(row=r, column=3, value=effort).alignment = CENTER
    ws5.cell(row=r, column=4, value=risk).alignment = WRAP
    ws5.cell(row=r, column=5, value=modified).alignment = CENTER
    ws5.row_dimensions[r].height = 50
    style_data_row(ws5, r, len(road_headers), alt=(i % 2 == 1))


output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "The Clothing Cove - Inventory Recovery Plan V2.xlsx")
wb.save(output_path)
print(f"Saved V2 to: {output_path}")
