import openpyxl
F="The Clothing Cove - SEO & AEO Plan.xlsx"
wb=openpyxl.load_workbook(F)
edits={
 "Product SEO": {"H5":"Write from title, brand & category","H7":"Write benefit-led snippets",
                 "H9":"Write from title, vendor & type, then review","H11":"Alt text = title + color",
                 "H12":"Same approach — do these first"},
 "AEO Readiness": {"E6":"Write 2–3 sentences each"},
 "Fix Roadmap": {
   "B7":"Write & optimize SEO titles for in-stock products","D7":"~280–350 hrs",
   "B8":"Write meta descriptions for in-stock products","D8":"~270–340 hrs",
   "D10":"~270–350 hrs","E10":"Filtering + AI taxonomy + smart-collection sorting",
   "B11":"Write product descriptions (missing + thin)","D11":"~450–600 hrs",
   "B12":"Write alt-text for in-stock images","D12":"~30–45 hrs","D13":"~50–75 hrs"},
}
for sheet,cells in edits.items():
    ws=wb[sheet]
    for coord,val in cells.items(): ws[coord]=val
wb.save(F)
# verify nothing automation-y remains in effort/method areas
import re
pat=re.compile(r'script|automat|bulk[- ]?generat', re.I)
left=[f"{ws.title}!{c.coordinate}: {c.value[:60]}" for ws in wb.worksheets for row in ws.iter_rows() for c in row
      if isinstance(c.value,str) and pat.search(c.value)]
print("remaining automation refs:", left if left else "none")
