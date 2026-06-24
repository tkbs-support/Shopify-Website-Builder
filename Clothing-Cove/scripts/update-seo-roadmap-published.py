# -*- coding: utf-8 -*-
import openpyxl, re
F="The Clothing Cove - SEO & AEO Plan.xlsx"
reps=[  # order matters (specific first)
 ("3,522 missing + 828 thin","143 missing + 541 thin"),
 ("1,277 images (738 products with zero alt)","842 images (487 products with zero alt)"),
 ("in-stock untagged (94%)","live untagged (96%)"),
 ("8,459","4,579"),("6,564","2,797"),("10,966","7,365"),("3,522","143"),
 ("1,277 images","842 images"),("738 products","487 products"),
 ("for in-stock products","for live products"),("Tag in-stock products","Tag live products"),
 ("in-stock untagged","live untagged"),("in-stock products","live products"),
 ("in-stock images","live images"),
]
wb=openpyxl.load_workbook(F); n=0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str):
                v=c.value
                for a,b in reps: v=v.replace(a,b)
                if v!=c.value: c.value=v; n+=1
wb.save(F)
print(f"replaced in {n} cells")
# re-scan for any remaining in-stock product counts
leftover=re.compile(r'8,459|6,564|10,966|3,522|1,277|\b909\b|in-stock')
hits=[f"{ws.title}!{c.coordinate}: {c.value[:46]}" for ws in wb.worksheets for r in ws.iter_rows() for c in r if isinstance(c.value,str) and leftover.search(c.value)]
print("leftover in-stock refs:", hits if hits else "NONE")
