"""Build the two client deliverables from data/analysis.json + analysis2.json.

- The Clothing Cove - Inventory Recovery Plan (June 2026).xlsx
- The Clothing Cove - SEO & AEO Plan (June 2026).xlsx

All numbers come from the fresh June 2026 audit data. In-stock scope leads.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), "..")
A = json.load(open(os.path.join(ROOT, "data", "analysis.json"), encoding="utf-8"))
A2 = json.load(open(os.path.join(ROOT, "data", "analysis2.json"), encoding="utf-8"))

AUDIT_DATE = "June 10, 2026"
NAVY = "1F3864"
TEAL = "2E9BA6"
RED = "C00000"
ORANGE = "ED7D31"
GREEN = "548235"
GRAY = "808080"
LIGHT = "F2F2F2"
SECTION_FILL = "D9E2F3"

F = lambda **kw: Font(name="Arial", **kw)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEV_FILL = {"CRITICAL": PatternFill("solid", start_color="F8CBAD"),
            "WARNING": PatternFill("solid", start_color="FFE699"),
            "OK": PatternFill("solid", start_color="C6EFCE")}
SEV_FONT = {"CRITICAL": F(bold=True, color="9C0006"),
            "WARNING": F(color="9C6500"),
            "OK": F(color="276221")}

T = A["totals"]
SA = A["seo_active"]
SS = A["seo_stock"]
IN_STOCK = T["in_stock"]
ACTIVE = T["active"]


def set_font_all(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.font is None or c.font.name != "Arial":
                c.font = F(size=c.font.size if c.font and c.font.size else 11,
                           bold=c.font.bold if c.font else False,
                           color=c.font.color if c.font else None)


def title_block(ws, title, subtitle, span=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = F(size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = F(size=10, italic=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="2F5496")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def kpi_block(ws, row, col, label, value, sub, color, num_fmt=None):
    """3-row KPI card across 2 columns."""
    for r in range(row, row + 3):
        for cc in range(col, col + 2):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", start_color=color)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    c = ws.cell(row=row, column=col, value=label)
    c.font = F(size=9, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    c = ws.cell(row=row + 1, column=col, value=value)
    c.font = F(size=20, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    if num_fmt:
        c.number_format = num_fmt
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
    c = ws.cell(row=row + 2, column=col, value=sub)
    c.font = F(size=8, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)


def section_header(ws, row, text, span=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F(size=11, bold=True, color=NAVY)
    c.fill = PatternFill("solid", start_color=SECTION_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def table_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def data_row(ws, row, values, start_col=1, fmts=None, fills=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = F(size=10)
        c.border = BORDER
        if fmts and fmts[i]:
            c.number_format = fmts[i]
        if fills and fills[i]:
            c.fill = fills[i]


def sev_cell(ws, row, col, sev):
    c = ws.cell(row=row, column=col, value=sev)
    c.fill = SEV_FILL.get(sev, PatternFill())
    c.font = SEV_FONT.get(sev, F(size=10))
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER


def color_points(series, colors):
    for i, col in enumerate(colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = col
        series.dPt.append(pt)


def clean_labels(show_val=False, show_percent=False):
    d = DataLabelList()
    d.showVal = show_val
    d.showPercent = show_percent
    d.showSerName = False
    d.showCatName = False
    d.showLegendKey = False
    d.showBubbleSize = False
    return d


def bar_chart(ws, anchor, title, data_ref, cat_ref, color=TEAL, horizontal=True,
              width=13, height=7, show_vals=True):
    ch = BarChart()
    ch.type = "bar" if horizontal else "col"
    ch.title = title
    ch.add_data(data_ref, titles_from_data=False)
    ch.set_categories(cat_ref)
    ch.legend = None
    ch.width = width
    ch.height = height
    ch.gapWidth = 60
    s = ch.series[0]
    s.graphicalProperties.solidFill = color
    if show_vals:
        ch.dataLabels = clean_labels(show_val=True)
    ws.add_chart(ch, anchor)
    return ch


# =====================================================================
# WORKBOOK 1 — FILTER & SEARCH READINESS PLAN
# =====================================================================
wb = Workbook()

# --- derived filter-readiness figures ---
COLOR_PROB = A["color_problems"]
SIZE_PROB = A["size_problems"]
COLOR_ALL = A["missing_color_alltypes"]
SIZE_ALL = A["missing_size_alltypes"]
stock_color_products = IN_STOCK - SS["missing_color"]
stock_size_products = IN_STOCK - SS["missing_size"]
swatch_pct = SS["has_color_pattern_mf"] / IN_STOCK
brand = A["brand"]
material = A["material"]
# readiness among the product types that SHOULD have each option
_csz = [t for t in A["by_type"] if t["expects_size"]]
_ccl = [t for t in A["by_type"] if t["expects_color"]]
size_ready = (sum(t["stock_size"] for t in _csz) / max(sum(t["in_stock"] for t in _csz), 1))
color_ready = (sum(t["stock_color"] for t in _ccl) / max(sum(t["in_stock"] for t in _ccl), 1))
# aggregate raw counts (measured) so the Summary readiness %s are formulas over them
color_have = sum(t["stock_color"] for t in _ccl)
color_scope = sum(t["in_stock"] for t in _ccl)
size_have = sum(t["stock_size"] for t in _csz)
size_scope = sum(t["in_stock"] for t in _csz)

# ---------- Summary ----------
ws = wb.active
ws.title = "Summary"
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKLM", [32, 12, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11]):
    ws.column_dimensions[col].width = w

title_block(ws, "The Clothing Cove — Filter & Search Readiness Plan",
            f"Audit date: {AUDIT_DATE}  |  theclothingcove.com  |  Goal: make the catalog searchable by color, size, type, brand & material  |  "
            "Scope: in-stock products", span=13)

# KPI row
kpi_block(ws, 4, 1, "FILTERS LIVE TODAY", "2 of 6", "Availability and Price only", NAVY)
kpi_block(ws, 4, 4, "APPAREL MISSING SIZE", SIZE_PROB, "in-stock apparel & footwear without a Size option", RED, "#,##0")
kpi_block(ws, 4, 7, "PRODUCTS MISSING COLOR", COLOR_PROB, "in-stock products that carry colors but have no Color option", RED, "#,##0")
kpi_block(ws, 4, 10, "COLOR SWATCH-READY", swatch_pct, f"{SS['has_color_pattern_mf']:,} products mapped to the palette metafield", ORANGE, "0.0%")
ws.row_dimensions[4].height = 26
ws.row_dimensions[5].height = 34

r = 8
section_header(ws, r, "OBJECTIVE", span=5)
r += 1
goal = ws.cell(row=r, column=1, value="Make the in-stock catalog searchable by the filters shoppers expect — color, size, garment type, brand and material. "
              "Each filter requires its product metadata to exist, be standardized, and be enabled in the storefront. "
              "The plan below brings each filter to that state, in priority order.")
goal.font = F(size=9, italic=True, color=GRAY)
goal.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=5)
r += 3

# Filter readiness snapshot (chart source)
fr_start = r
section_header(ws, r, "FILTER READINESS AT A GLANCE  (see ‘Filter Readiness Map’ for the full plan)", span=5)
r += 1
table_header(ws, r, ["Target filter", "Products ready", "In scope", "Data readiness"])
fr_hdr = r
# (label, ready_count, scope_count) — counts are measured; readiness is a formula =ready/scope
fr_rows = [
    ("Availability", IN_STOCK, IN_STOCK),
    ("Price", IN_STOCK, IN_STOCK),
    ("Brand", IN_STOCK - brand["missing"], IN_STOCK),
    ("Garment type", IN_STOCK - SS["missing_type"], IN_STOCK),
    ("Color (option)", color_have, color_scope),
    ("Size (option)", size_have, size_scope),
    ("Color swatches", SS["has_color_pattern_mf"], IN_STOCK),
    ("Material / fabric", material["structured_option"], IN_STOCK),
]
for i, (label, ready, scope) in enumerate(fr_rows):
    rr = r + 1 + i
    data_row(ws, rr, [label, ready, scope, f"=IF(C{rr}=0,0,B{rr}/C{rr})"],
             fmts=[None, "#,##0", "#,##0", "0.0%"])
last_fr = r + len(fr_rows)
ws.conditional_formatting.add(f"D{fr_hdr + 1}:D{last_fr}",
                              DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                                          color=TEAL, showValue=True))
bar_chart(ws, "F" + str(fr_start), "Filter data readiness (in-stock)",
          Reference(ws, min_col=4, min_row=fr_hdr + 1, max_row=last_fr),
          Reference(ws, min_col=1, min_row=fr_hdr + 1, max_row=last_fr),
          color=TEAL, width=15, height=8.5, show_vals=False)
ch = ws._charts[-1]
color_points(ch.series[0], [GREEN, GREEN, TEAL, TEAL, ORANGE, ORANGE, RED, RED])
r = max(r + len(fr_rows) + 2, fr_start + 17)

# Attribute gaps by priority (data only)
ta_start = r
section_header(ws, r, "ATTRIBUTE GAPS BY PRIORITY  (in-stock products to update)", span=5)
r += 1
ta_hdr = r
table_header(ws, r, ["Attribute gap (priority order)", "", "Products"])
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
# counts are COUNTIF formulas over the Problem Worklist priority column
ta_rows = [
    ("1 — Apparel & footwear missing a Size option", 1),
    ("2 — Apparel missing a Color option", 2),
    ("3 — Accessories missing a Color option", 3),
    ("4 — Jewelry missing a Color option", 4),
]
for i, (label, prio) in enumerate(ta_rows):
    rr = r + 1 + i
    data_row(ws, rr, [label, "", f"=COUNTIF('Problem Worklist'!$A:$A,{prio})"], fmts=[None, None, "#,##0"])
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
rr = r + len(ta_rows) + 1
data_row(ws, rr, ["Total products to update", "", f"=SUM(C{ta_hdr + 1}:C{ta_hdr + len(ta_rows)})"], fmts=[None, None, "#,##0"])
ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
ws.cell(row=rr, column=1).font = F(size=10, bold=True)
ws.cell(row=rr, column=3).font = F(size=10, bold=True)
r = max(rr + 2, ta_start + 8)

# Catalog overview (doughnut)
co_start = r
section_header(ws, r, "CATALOG OVERVIEW", span=5)
r += 1
table_header(ws, r, ["Status", "Products", "% of active"])
co_hdr = r
active_row = co_hdr + 3
co_rows = [("In stock", T["in_stock"]), ("Out of stock (still published)", T["out_of_stock"])]
for i, (label, n) in enumerate(co_rows):
    rr = r + 1 + i
    data_row(ws, rr, [label, n, f"=B{rr}/$B${active_row}"], fmts=[None, "#,##0", "0.0%"])
rr = active_row
data_row(ws, rr, ["Active products (total)", ACTIVE, f"=B{rr}/$B${rr}"], fmts=[None, "#,##0", "0.0%"])
# scope cells other sheets reference for their %s (these hold raw measured totals)
INSTOCK_REF = f"'Summary'!$B${co_hdr + 1}"
ACTIVE_REF = f"'Summary'!$B${active_row}"
ws.cell(row=rr, column=1).font = F(size=10, bold=True)
data_row(ws, rr + 1, ["Draft products (not published)", T["draft"], ""], fmts=[None, "#,##0", None])
note = ws.cell(row=rr + 2, column=1, value="Two-thirds of the published catalog is out of stock — those pages dilute filtering, search, and SEO.")
note.font = F(size=9, italic=True, color=GRAY)
dch = DoughnutChart()
dch.title = "Active catalog: in stock vs out of stock"
dch.add_data(Reference(ws, min_col=2, min_row=co_hdr + 1, max_row=co_hdr + 2), titles_from_data=False)
dch.set_categories(Reference(ws, min_col=1, min_row=co_hdr + 1, max_row=co_hdr + 2))
dch.height = 7
dch.width = 11
dch.dataLabels = clean_labels(show_percent=True)
color_points(dch.series[0], [GREEN, "D9D9D9"])
ws.add_chart(dch, "F" + str(co_start))

# taller rows so KPI sub-text and the wrapped objective note don't clip/overlap
ws.row_dimensions[6].height = 30      # KPI card sub-text
ws.row_dimensions[9].height = 24      # objective paragraph
ws.row_dimensions[10].height = 24

# ---------- Filter Readiness Map (the centerpiece) ----------
ws = wb.create_sheet("Filter Readiness Map")
ws.sheet_properties.tabColor = TEAL
ws.sheet_view.showGridLines = False
title_block(ws, "Filter Readiness Map — the metadata behind every filter",
            "For each target filter: where the data lives, whether it’s standardized, what field Shopify can filter on, and the work to turn it on. "
            "Shopify Search & Discovery builds facets from product options, metafields, or tags only.", span=8)
hdr = 4
table_header(ws, hdr, ["Target filter", "Metadata source", "Data present?", "Shopper-ready?",
                       "Filterable via", "On site?", "Work to enable", "Effort"])
fmap = [
    ("Availability", "Shopify native", "Yes", "Yes", "Native facet", "LIVE",
     "Already on — keep", "—"),
    ("Price", "Shopify native", "Yes", "Yes", "Native facet", "LIVE",
     "Already on — keep", "—"),
    ("Brand", f"Vendor field ({brand['distinct']} brands)", "Yes — 0 missing", "No — supplier codes",
     "Vendor facet", "OFF",
     "Normalize vendor names (e.g. LEEGIN LEATHER → Brighton), then enable the Brand filter", "Low–Med"),
    ("Garment type", "productType codes", "Yes", "No — cryptic 3-letter codes",
     "Product-type facet", "OFF",
     "Map product-type codes to shopper-facing labels, then enable the Product-type filter", "Low"),
    ("Color (text)", "Color option", f"{stock_color_products:,} of {IN_STOCK:,}", "No — 6 spellings",
     "Option facet", "OFF",
     "Unify Color/COLOR/colors spellings; add a Color option where missing; enable Color filter", "Med"),
    ("Color (swatches)", "shopify.color-pattern metafield", f"{SS['has_color_pattern_mf']:,} ({swatch_pct:.0%})", "No — 10% mapped",
     "Metafield facet", "OFF",
     f"Map ~{A2['color_values_stock_unique']:,} raw values → palette metafield; turn on swatches", "Med–High"),
    ("Size", "Size option", "Apparel mostly yes", "No — Size/SIZE split",
     "Option facet", "OFF",
     f"Unify naming; add Size to {SIZE_PROB} apparel/footwear gaps; enable Size filter", "Low–Med"),
    ("Material / fabric", "Description text only", f"~{material['in_desc']:,} mention fabric; 0 structured", "No",
     "Metafield (to build)", "OFF",
     "Extract fabric words from descriptions → a Material metafield/tag, then enable", "High"),
    ("Occasion / style / fit", "None yet", "No", "No",
     "Tags/metafields (to build)", "OFF",
     "Net-new attribute tagging — longest-horizon; what the biggest sites offer", "High (future)"),
]
sev_for = {"LIVE": "OK", "OFF": "CRITICAL"}
for i, (filt, src, present, ready, via, on, work, effort) in enumerate(fmap):
    rr = hdr + 1 + i
    data_row(ws, rr, [filt, src, present, ready, via, None, work, effort])
    ws.cell(row=rr, column=1).font = F(size=10, bold=True, color=NAVY)
    sev_cell(ws, rr, 6, sev_for[on])
    ws.cell(row=rr, column=6).value = on
    for cc in (2, 3, 4, 7):
        ws.cell(row=rr, column=cc).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 34
for col, w in zip("ABCDEFGH", [18, 26, 18, 20, 18, 9, 50, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------- Filter Readiness by Garment Type ----------
ws = wb.create_sheet("By Garment Type")
ws.sheet_view.showGridLines = False
title_block(ws, "Filter Readiness by Garment Type (in-stock)",
            "Decoded product types. Color % = Has Color ÷ In Stock; Color gap = In Stock − Has Color where the option applies "
            "(blank where it does not, e.g. Size for jewelry). Every % and gap is a live Excel formula over the count columns.", span=13)
hdr = 4
table_header(ws, hdr, ["Garment Type", "Code", "In Stock", "Has Color", "Color %", "Has Size", "Size %",
                       "Color filter?", "Size filter?", "Color gap", "Size gap", "Active (all)", "Note"])
rows = [t for t in sorted(A["by_type"], key=lambda x: -x["in_stock"]) if t["in_stock"] > 0]
for i, t in enumerate(rows):
    rr = hdr + 1 + i
    note = "High-volume Color gap" if t["color_problems"] >= 200 else ("Size gaps to fill" if t["size_problems"] >= 15 else "")
    data_row(ws, rr, [t.get("label", t["type"]), t["type"], t["in_stock"], t["stock_color"],
                      f"=IF(C{rr}=0,0,D{rr}/C{rr})", t["stock_size"], f"=IF(C{rr}=0,0,F{rr}/C{rr})",
                      "Yes" if t["expects_color"] else "—", "Yes" if t["expects_size"] else "—",
                      f'=IF(H{rr}="Yes",C{rr}-D{rr},"")', f'=IF(I{rr}="Yes",C{rr}-F{rr},"")',
                      t["total"], note],
             fmts=[None, None, "#,##0", "#,##0", "0.0%", "#,##0", "0.0%", None, None, "#,##0", "#,##0", "#,##0", None])
    for cc in (8, 9):
        ws.cell(row=rr, column=cc).alignment = Alignment(horizontal="center")
last = hdr + len(rows)
for col in ("E", "G"):
    ws.conditional_formatting.add(f"{col}{hdr + 1}:{col}{last}",
                                  DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                                              color=TEAL, showValue=True))
for col, w in zip("ABCDEFGHIJKLM", [24, 8, 9, 10, 9, 10, 9, 11, 11, 10, 10, 11, 22]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "C5"

# ---------- Problem Worklist (actionable SKU list) ----------
ws = wb.create_sheet("Problem Worklist")
ws.sheet_properties.tabColor = RED
ws.sheet_view.showGridLines = False
wl = A["problem_worklist"]
n_p1 = sum(1 for w in wl if w["priority"] == 1)
n_p2 = sum(1 for w in wl if w["priority"] == 2)
title_block(ws, "Product Worklist — in-stock products needing a Color or Size option",
            f"{len(wl):,} products. Priority 1 = apparel & footwear missing Size ({n_p1}); 2 = apparel missing Color ({n_p2}); "
            f"3 = accessories missing Color; 4 = jewelry missing Color. Use the column filters to work one batch at a time.", span=7)
hdr = 4
table_header(ws, hdr, ["Priority", "Product", "Garment Type", "Brand", "Missing Color", "Missing Size", "Fix (Shopify admin)"])
PR_FILL = {1: PatternFill("solid", start_color="F8CBAD"), 2: PatternFill("solid", start_color="FFD9A0"),
           3: PatternFill("solid", start_color="FFE699"), 4: PatternFill("solid", start_color="FFF2CC")}
for i, w in enumerate(wl):
    rr = hdr + 1 + i
    data_row(ws, rr, [w["priority"], w["title"], w["type"], w["vendor"],
                      w["missing_color"], w["missing_size"], ""],
             fmts=[None, None, None, None, None, None, None])
    ws.cell(row=rr, column=1).fill = PR_FILL[w["priority"]]
    ws.cell(row=rr, column=1).alignment = Alignment(horizontal="center")
    lc = ws.cell(row=rr, column=7, value="Edit in Shopify ↗")
    lc.hyperlink = w["admin_url"]
    lc.font = F(size=10, color="0563C1", underline="single")
    lc.border = BORDER
for col, w in zip("ABCDEFG", [8, 46, 20, 26, 13, 12, 20]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{hdr}:G{hdr + len(wl)}"

# ---------- Color Values ----------
ws = wb.create_sheet("Color Values")
ws.sheet_view.showGridLines = False
title_block(ws, "Color Values — In-Stock Products",
            f"{A2['color_values_stock_unique']:,} unique color values across {stock_color_products:,} in-stock products with a Color option "
            f"({A['color_values_unique']:,} unique across full active catalog). Value sprawl is why a clean palette metafield is needed.", span=5)
scope_lbl = ws.cell(row=3, column=1, value="In-stock products with a Color option (scope for %):")
scope_lbl.font = F(size=9, italic=True, color=GRAY)
ws.merge_cells("A3:B3")
ws.cell(row=3, column=3, value=stock_color_products).font = F(size=9, bold=True)
ws.cell(row=3, column=3).number_format = "#,##0"
hdr = 4
table_header(ws, hdr, ["Rank", "Color Value", "In-Stock Products", "% of in-stock color products"])
for i, (val, n) in enumerate(A2["color_values_stock"][:60]):
    rr = hdr + 1 + i
    data_row(ws, rr, [i + 1, val.title(), n, f"=C{rr}/$C$3"],
             fmts=[None, None, "#,##0", "0.0%"])
for col, w in zip("ABCD", [7, 26, 16, 24]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------- Size Values ----------
ws = wb.create_sheet("Size Values")
ws.sheet_view.showGridLines = False
title_block(ws, "Size Values — In-Stock Products",
            f"{A2['size_values_stock_unique']:,} unique size values across {stock_size_products:,} in-stock products with a Size option", span=5)
hdr = 4
table_header(ws, hdr, ["Rank", "Size Value", "In-Stock Products", "Category"])
def size_cat(v):
    s = v.strip().upper()
    if s.replace("X", "") in ("S", "M", "L", "XS", "XL", "1X", "2X", "3X") or s in ("XXL", "XXS", "S/M", "M/L", "L/XL", "ONE SIZE", "OS", "O/S"):
        return "Letter"
    try:
        float(s)
        return "Numeric"
    except ValueError:
        return "Other"
for i, (val, n) in enumerate(A2["size_values_stock"][:60]):
    rr = hdr + 1 + i
    data_row(ws, rr, [i + 1, val, n, size_cat(val)], fmts=[None, None, "#,##0", None])
for col, w in zip("ABCD", [7, 26, 16, 14]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------- Option Types ----------
ws = wb.create_sheet("Option Types")
ws.sheet_view.showGridLines = False
title_block(ws, "Complete Inventory of Product Option Types",
            f"{ACTIVE:,} active products scanned  |  {SA['default_title']:,} have no options at all (“Default Title”). "
            "‘Spellings found’ shows every naming variant that fragments a filter (e.g. Size vs SIZE).", span=7)
hdr = 4
table_header(ws, hdr, ["Option (canonical)", "Active Products", "In Stock", "Unique Values", "Spellings found", "Assessment"])
notes = {"color": "PRIMARY filter — fix naming, then map values",
         "size": "PRIMARY filter — fix naming",
         "style": "Minor — high value-sprawl",
         "scent": "Niche (candles, home)",
         "shoe size": "Clean — footwear only",
         "colors": "BUG — merge into Color",
         "print": "Niche", "ring size": "Niche — jewelry", "strength": "Reading glasses"}
for i, o in enumerate(A["option_types"]):
    rr = hdr + 1 + i
    spell = ", ".join(f"{k}:{v}" for k, v in o["raw_names"].items())
    data_row(ws, rr, [o["canonical"].title(), o["products"], o["in_stock"], o["unique_values"],
                      spell, notes.get(o["canonical"], "Review — non-standard option")],
             fmts=[None, "#,##0", "#,##0", "#,##0", None, None])
for col, w in zip("ABCDEF", [18, 14, 10, 13, 46, 40]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------- Data Quality ----------
ws = wb.create_sheet("Data Quality")
ws.sheet_view.showGridLines = False
fixable = [d for d in A["data_quality"] if not d.get("nonstandard")]
review = sorted([d for d in A["data_quality"] if d.get("nonstandard")], key=lambda d: -d["products"])
total_fixable = sum(d["products"] for d in fixable)
title_block(ws, "Data Quality — Option Naming Cleanup List",
            f"{total_fixable:,} products carry a misspelled or inconsistent option name. "
            "Fixes change only the option LABEL — no variants, prices, or inventory are touched.", span=5)
hdr = 4
table_header(ws, hdr, ["Issue", "Current Spelling", "Should Be", "Products Affected"])
rr = hdr
for d in fixable:
    rr += 1
    data_row(ws, rr, [d["issue"], repr(d["raw"])[1:-1] if d["raw"].strip() != d["raw"] else d["raw"],
                      d["canonical"], d["products"]], fmts=[None, None, None, "#,##0"])
rr += 1
data_row(ws, rr, ["TOTAL (rename in one scripted batch)", "", "", f"=SUM(D{hdr + 1}:D{rr - 1})"], fmts=[None, None, None, "#,##0"])
ws.cell(row=rr, column=1).font = F(size=10, bold=True)
ws.cell(row=rr, column=4).font = F(size=10, bold=True)
rr += 2
section_header(ws, rr, "NON-STANDARD OPTION NAMES — typos & values used as options", span=4)
rr += 1
table_header(ws, rr, ["Verdict", "Option Name", "", "Products"])
for d in review:
    rr += 1
    data_row(ws, rr, [d["issue"], repr(d["raw"])[1:-1] if d["raw"].strip() != d["raw"] else d["raw"],
                      "", d["products"]], fmts=[None, None, None, "#,##0"])
    if d["issue"].startswith(("Typo", "Invalid")):
        ws.cell(row=rr, column=1).font = F(size=10, bold=True, color="9C0006")
for col, w in zip("ABCD", [28, 34, 4, 16]):
    ws.column_dimensions[col].width = w

# ---------- Color Swatch Metadata ----------
ws = wb.create_sheet("Color Swatch Metadata")
ws.sheet_view.showGridLines = False
title_block(ws, "Color Swatch Metadata — the swatch filter’s foundation",
            "Shopify’s shopify.color-pattern metafield maps messy color names to a standard palette. Swatches can’t render without it.", span=4)
hdr = 4
table_header(ws, hdr, ["Metric", "Value", "% of scope"])
mf_rows = [
    ("In-stock products with color-pattern metafield", SS["has_color_pattern_mf"], INSTOCK_REF),
    ("Active products with color-pattern metafield", SA["has_color_pattern_mf"], ACTIVE_REF),
    ("Unique palette colors referenced (in-stock)", A2["metaobject_refs_unique_stock"], None),
    ("Unique palette colors referenced (active)", A2["metaobject_refs_unique_active"], None),
    ("Unique raw color values to map (in-stock)", A2["color_values_stock_unique"], None),
]
for i, (label, v, scope_ref) in enumerate(mf_rows):
    rr = hdr + 1 + i
    pct = f"=B{rr}/{scope_ref}" if scope_ref else ""
    data_row(ws, rr, [label, v, pct], fmts=[None, "#,##0", "0.0%"])
rr = hdr + len(mf_rows) + 2
note = ws.cell(row=rr, column=1, value=f"Read: only {SS['has_color_pattern_mf'] / IN_STOCK:.0%} of in-stock products have a mapped palette color. "
                                       f"The remaining ~{A2['color_values_stock_unique']:,} raw values need mapping to ~{A2['metaobject_refs_unique_stock']} palette colors for clean swatch filtering.")
note.font = F(size=9, italic=True, color=GRAY)
note.alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=rr, start_column=1, end_row=rr + 1, end_column=3)
for col, w in zip("ABC", [48, 12, 12]):
    ws.column_dimensions[col].width = w

# ---------- Material & Brand Readiness ----------
ws = wb.create_sheet("Material & Brand")
ws.sheet_view.showGridLines = False
title_block(ws, "Material & Brand — facets that need metadata built",
            "Two high-value filters whose data isn’t yet in a filterable field.", span=5)
section_header(ws, 4, "MATERIAL / FABRIC — data is trapped in product descriptions", span=5)
table_header(ws, 5, ["Metric", "Value", "% of in-stock"])
mat_rows = [
    ("In-stock products mentioning a fabric in the description", material["in_desc"]),
    ("In-stock products with a structured Material option/field", material["structured_option"]),
]
for i, (label, v) in enumerate(mat_rows):
    rr = 6 + i
    data_row(ws, rr, [label, v, f"=B{rr}/{INSTOCK_REF}"], fmts=[None, "#,##0", "0.0%"])
data_row(ws, 8, ["Most-mentioned fabrics:", ", ".join(f"{f} ({c})" for f, c in material["top_fabrics"][:8]), ""], fmts=[None, None, None])
ws.cell(row=8, column=2).alignment = Alignment(wrap_text=True)
mnote = ws.cell(row=10, column=1, value="Fabric is described in prose but not stored as a filterable attribute. A script can extract fabric keywords from "
              "descriptions into a Material metafield/tag — turning ~1 in 5 products into material-filterable with no manual tagging.")
mnote.font = F(size=9, italic=True, color=GRAY)
mnote.alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=10, start_column=1, end_row=11, end_column=5)

section_header(ws, 13, "BRAND — data is complete but values are supplier codes", span=5)
table_header(ws, 14, ["Metric", "Value", ""])
data_row(ws, 15, ["Distinct vendor values (in-stock)", brand["distinct"], ""], fmts=[None, "#,##0", None])
data_row(ws, 16, ["In-stock products missing a vendor", brand["missing"], ""], fmts=[None, "#,##0", None])
data_row(ws, 18, ["Top vendor values (note: supplier names, not shopper brands):", "", ""])
for i, (v, c) in enumerate(brand["top"][:8]):
    data_row(ws, 19 + i, [f"   {v}", c, ""], fmts=[None, "#,##0", None])
bnote = ws.cell(row=28, column=1, value="The Brand filter is data-complete (0 missing) but the values are supplier/distributor codes (e.g. ‘LEEGIN LEATHER’ = Brighton). "
              "Normalize vendor names to shopper-facing brands, then the Brand filter can be switched on immediately.")
bnote.font = F(size=9, italic=True, color=GRAY)
bnote.alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=28, start_column=1, end_row=29, end_column=5)
for col, w in zip("ABCDE", [52, 16, 10, 10, 10]):
    ws.column_dimensions[col].width = w

# ---------- Fix Roadmap ----------
ws = wb.create_sheet("Fix Roadmap")
ws.sheet_properties.tabColor = GREEN
ws.sheet_view.showGridLines = False
title_block(ws, "Filter Readiness — Fix Roadmap",
            "Sequenced to switch on the most filters per hour of work. Phases 1–3 turn on filters using data that already exists.", span=5)
hdr = 4
table_header(ws, hdr, ["Phase", "Action", "Unlocks filter", "Effort", "Risk"])
roadmap = [
    ("Phase 1", "Option naming cleanup — unify Color/COLOR/colors and Size/SIZE; fix the typos and invalid option names (see Data Quality).",
     "Color, Size (prereq)", f"{total_fixable:,} labels / 2–3 hrs", "Low — labels only; no variants/prices touched"),
    ("Phase 2", "Enable the filters whose data is already complete — Brand (after vendor-name normalization) and Garment Type (after code→label mapping).",
     "Brand, Garment type", "Brand + type / 1–2 days", "Low — mapping + Search & Discovery config"),
    ("Phase 3", "Turn on Color & Size filters in Search & Discovery + activate color swatches on the theme (currently no swatch config).",
     "Color, Size", "Theme + app / 2–4 hrs", "None — display settings, instantly reversible"),
    ("Phase 4", "Canonical color mapping — map raw color values to Shopify’s palette (color-pattern metafield) so swatches render.",
     "Color swatches", f"~{A2['color_values_stock_unique']:,} values / 1–2 wks", "Low — additive metadata"),
    ("Phase 5", "Fill remaining option gaps — add Size to apparel & footwear and Color where the product type uses it (see Product Worklist).",
     "Color, Size coverage", f"{SIZE_PROB} Size + {COLOR_PROB:,} Color / incremental", "Medium — product edits; dry-run per type"),
    ("Phase 6", "Build the Material filter — extract fabric from descriptions into a Material metafield/tag, then enable.",
     "Material / fabric", f"~{material['in_desc']:,} products / 1–2 wks", "Low — additive, script-extracted"),
    ("Phase 7", "Out-of-stock cleanup — archive stale zero-inventory listings (mostly 2018–2022) so filters/search reflect real stock.",
     "All (signal quality)", f"{T['out_of_stock']:,} OOS / 1–2 days", "Medium — reversible (archive, don’t delete)"),
    ("Phase 8", "Publish & verify — confirm every enabled filter on key collections, re-crawl, spot-check top 20 collections.",
     "Verification", "Site-wide / 2–3 hrs", "Low"),
]
for i, row in enumerate(roadmap):
    rr = hdr + 1 + i
    data_row(ws, rr, list(row))
    ws.cell(row=rr, column=1).font = F(size=10, bold=True, color=NAVY)
    for cc in range(1, 6):
        ws.cell(row=rr, column=cc).alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[rr].height = 44
for col, w in zip("ABCDE", [10, 62, 22, 26, 40]):
    ws.column_dimensions[col].width = w

out1 = os.path.join(ROOT, "The Clothing Cove - Filter & Search Readiness Plan (June 2026).xlsx")
wb.save(out1)
print("saved", out1)

# =====================================================================
# WORKBOOK 2 — SEO & AEO PLAN
# =====================================================================
wb = Workbook()
nav = A["navigable"]
cols_by_handle = {c["handle"]: c for c in A["collections_list"]}
C = A["collections"]
pages = A["pages"]
pages_pub = sum(1 for p in pages if p["published"])
blog = A["blog"]

# join navigable collections
nav_rows = []
for h, info in nav.items():
    c = cols_by_handle.get(h)
    title = c["title"] if c else h.replace("-", " ").title()
    desc_len = c["desc_len"] if c else 0
    seo_t = bool(c and c["seo_title"])
    needs = desc_len < 100 or not seo_t
    # proposed title
    base = title.split("\\")[-1].strip() if "\\" in title else title
    if info["category"] == "Brand":
        prop = f"{base} | The Clothing Cove – Milford, MI"
    elif info["category"] == "Top-Level":
        prop = f"{base} for Women | The Clothing Cove – Milford, MI"
    else:
        prop = f"Women's {base} | The Clothing Cove"
    if len(prop) > 60:
        prop = f"{base} | The Clothing Cove"
    nav_rows.append({"category": info["category"], "url": f"/collections/{h}", "title": title,
                     "products": c["products"] if c else 0, "desc_len": desc_len,
                     "has_desc": "Yes" if desc_len >= 100 else ("Short" if desc_len > 0 else "No"),
                     "has_seo_title": "Yes" if seo_t else "No",
                     "proposed": prop if needs else "", "needs": "Yes" if needs else "No"})
cat_order = {"Top-Level": 0, "Subcategory": 1, "Brand": 2}
nav_rows.sort(key=lambda x: (cat_order[x["category"]], x["url"]))
nav_needs = sum(1 for x in nav_rows if x["needs"] == "Yes")

ws = wb.active
ws.title = "Executive Summary"
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKLM", [42, 12, 12, 12, 12, 11, 11, 11, 11, 11, 11, 11, 11]):
    ws.column_dimensions[col].width = w
title_block(ws, "The Clothing Cove — SEO & AEO Plan",
            f"Audit date: {AUDIT_DATE}  |  theclothingcove.com  |  Scope: in-stock products ({IN_STOCK:,})  |  "
            "Source: Shopify Admin API + live site crawl", span=13)

kpi_block(ws, 4, 1, "IN-STOCK MISSING SEO TITLE", SS["missing_seo_title"] / IN_STOCK, f"{SS['missing_seo_title']:,} products", RED, "0.0%")
kpi_block(ws, 4, 4, "IN-STOCK MISSING TAGS", SS["missing_tags"] / IN_STOCK, f"{SS['missing_tags']:,} products", RED, "0.0%")
kpi_block(ws, 4, 7, "COLLECTIONS WITHOUT DESCRIPTION", C["missing_desc"] / C["total"], f"{C['missing_desc']} of {C['total']} collections", ORANGE, "0.0%")
kpi_block(ws, 4, 10, "LAST BLOG POST", "Apr 2023", "no fresh content for AI/Google to cite", GRAY)
ws.row_dimensions[4].height = 26
ws.row_dimensions[5].height = 30

r = 8
section_header(ws, r, "PRODUCT SEO GAPS — IN-STOCK PRODUCTS (the work scope)", span=5)
r += 1
table_header(ws, r, ["Gap", "In-Stock", "% of in-stock", "Full catalog (ref)", "Severity"])
ps_hdr = r
ps_rows = [
    ("Missing SEO title tag", SS["missing_seo_title"], SA["missing_seo_title"], "CRITICAL"),
    ("Missing tags", SS["missing_tags"], SA["missing_tags"], "CRITICAL"),
    ("Missing meta description", SS["missing_seo_desc"], SA["missing_seo_desc"], "CRITICAL"),
    ("Missing ALL product images", SS["missing_images"], SA["missing_images"], "CRITICAL"),
    ("Missing description (body)", SS["missing_body"], SA["missing_body"], "CRITICAL"),
    ("Duplicate titles", SS["dup_title_products"], SA["dup_title_products"], "WARNING"),
    ("Short description (<100 chars)", SS["short_body"], SA["short_body"], "WARNING"),
    ("No alt text on any image", SS["products_zero_alt"], SA["products_zero_alt"], "WARNING"),
]
for i, (label, n_s, n_a, sev) in enumerate(ps_rows):
    rr = r + 1 + i
    data_row(ws, rr, [label, n_s, f"=B{rr}/{IN_STOCK}", n_a, None],
             fmts=[None, "#,##0", "0.0%", "#,##0", None])
    sev_cell(ws, rr, 5, sev)
bar_chart(ws, "G" + str(ps_hdr - 1), "In-stock products affected by each SEO gap",
          Reference(ws, min_col=2, min_row=ps_hdr + 1, max_row=ps_hdr + len(ps_rows)),
          Reference(ws, min_col=1, min_row=ps_hdr + 1, max_row=ps_hdr + len(ps_rows)),
          color=RED, width=16, height=9)
r = max(r + len(ps_rows) + 2, ps_hdr + 18)

# Collections doughnut
cs_start = r
section_header(ws, r, "COLLECTION PAGES — category-level content Google & AI rely on", span=5)
r += 1
table_header(ws, r, ["Description status", "Collections", "% of all"])
cs_hdr = r
cs_rows = [("No description", C["missing_desc"]), ("Short (<100 chars)", C["short_desc"]), ("Good (100+ chars)", C["good_desc"])]
for i, (label, n) in enumerate(cs_rows):
    rr = r + 1 + i
    data_row(ws, rr, [label, n, f"=B{rr}/{C['total']}"], fmts=[None, "#,##0", "0.0%"])
rr = r + 4
extra = [("Total collections", C["total"]), ("In site navigation (shopper-visible)", len(nav_rows)),
         ("Navigable collections needing SEO work", nav_needs),
         ("Missing custom SEO title", C["missing_seo_title"]), ("Empty (0 products)", C["empty"])]
for label, n in extra:
    data_row(ws, rr, [label, n, ""], fmts=[None, "#,##0", None])
    rr += 1
dch = DoughnutChart()
dch.title = "Collection description coverage (544 collections)"
dch.add_data(Reference(ws, min_col=2, min_row=cs_hdr + 1, max_row=cs_hdr + 3), titles_from_data=False)
dch.set_categories(Reference(ws, min_col=1, min_row=cs_hdr + 1, max_row=cs_hdr + 3))
dch.height = 7.5
dch.width = 12
dch.dataLabels = clean_labels(show_percent=True)
color_points(dch.series[0], [RED, ORANGE, GREEN])
ws.add_chart(dch, "G" + str(cs_start))
r = max(rr + 1, cs_start + 17)

# Technical snapshot
section_header(ws, r, "TECHNICAL & AEO SNAPSHOT (live site, June 10, 2026)", span=5)
r += 1
table_header(ws, r, ["Check", "Status", "", "", "Severity"])
tech_rows = [
    ("Homepage title tag", "“The Clothing Cove” — no keywords or location", "CRITICAL"),
    ("May 2026 theme SEO work", "Sits on an UNPUBLISHED theme — not live", "CRITICAL"),
    ("Collection meta descriptions", "Missing on live collection pages (e.g. /dresses)", "CRITICAL"),
    ("Product structured data (JSON-LD)", "Present via app — Product, Offer, Brand", "OK"),
    ("Store schema on homepage", "ClothingStore with NAP + hours present", "OK"),
    ("Open Graph / Twitter cards", "Present via app on live pages", "OK"),
    ("Collection filters", "Only Availability + Price — no Color/Size/Brand", "CRITICAL"),
    ("Blog freshness", "Last post April 2023", "WARNING"),
]
for i, (label, status, sev) in enumerate(tech_rows):
    rr = r + 1 + i
    data_row(ws, rr, [label, status, "", "", None])
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
    sev_cell(ws, rr, 5, sev)
r += len(tech_rows) + 2

# Blog by year chart
bl_start = r
section_header(ws, r, "BLOG ACTIVITY BY YEAR — fresh content is what AI assistants cite", span=5)
r += 1
table_header(ws, r, ["Year", "Articles published"])
bl_hdr = r
by = blog["by_year"]
yrs = [str(y) for y in range(2020, 2027)]
for i, y in enumerate(yrs):
    rr = r + 1 + i
    data_row(ws, rr, [y, by.get(y, 0)], fmts=[None, "#,##0"])
bar_chart(ws, "G" + str(bl_start), "Blog posts per year",
          Reference(ws, min_col=2, min_row=bl_hdr + 1, max_row=bl_hdr + len(yrs)),
          Reference(ws, min_col=1, min_row=bl_hdr + 1, max_row=bl_hdr + len(yrs)),
          color=GRAY, horizontal=False, width=12, height=6.5)

# ---------- Product SEO ----------
ws = wb.create_sheet("Product SEO")
ws.sheet_view.showGridLines = False
title_block(ws, "Product SEO — Detail",
            f"In-stock scope: {IN_STOCK:,} products. Full-catalog figures shown for reference only.", span=8)
hdr = 4
table_header(ws, hdr, ["Issue", "In-Stock", "% In-Stock", "Full Catalog", "% Catalog", "Severity", "Impact", "Fix Approach"])
detail = [
    ("Missing SEO title tag", SS["missing_seo_title"], SA["missing_seo_title"], "CRITICAL",
     "Google shows auto-generated titles; no keyword targeting.", "Bulk-generate from title + brand + category"),
    ("Missing tags", SS["missing_tags"], SA["missing_tags"], "CRITICAL",
     "No filtering, categorization, or AI taxonomy signals.", "Bulk-tag from type + vendor + attributes"),
    ("Missing meta description", SS["missing_seo_desc"], SA["missing_seo_desc"], "CRITICAL",
     "Google writes its own snippet; lower click-through.", "Bulk-generate benefit-led snippets"),
    ("Missing ALL images", SS["missing_images"], SA["missing_images"], "CRITICAL",
     "Invisible to image search, Shopping, and social.", "Photograph or hide until photographed"),
    ("Missing description (body)", SS["missing_body"], SA["missing_body"], "CRITICAL",
     "Nothing for Google or AI to index — product is invisible.", "Generate from title/vendor/type, then review"),
    ("Short description (<100 chars)", SS["short_body"], SA["short_body"], "WARNING",
     "Thin content ranks poorly; AI skips it.", "Expand with fabric/fit/occasion copy"),
    ("Images without alt text", SS["images_no_alt"], SA["images_no_alt"], "WARNING",
     f"Of {SS['images_total']:,} in-stock images. Missed image SEO + accessibility.", "Script alt = title + color"),
    ("No alt text on ANY image", SS["products_zero_alt"], SA["products_zero_alt"], "WARNING",
     "Whole product invisible to image search.", "Same script — run these first"),
    (f"Duplicate titles ({SS['dup_title_groups']} titles shared)", SS["dup_title_products"], SA["dup_title_products"], "WARNING",
     "Products compete with each other in search.", "Differentiate with color/style descriptors"),
    ("Missing product type", SS["missing_type"], SA["missing_type"], "OK",
     "Minor gap.", "Fill from vendor catalog"),
]
for i, (label, n_s, n_a, sev, impact, fix) in enumerate(detail):
    rr = hdr + 1 + i
    data_row(ws, rr, [label, n_s, f"=B{rr}/{IN_STOCK}", n_a, f"=D{rr}/{ACTIVE}", None, impact, fix],
             fmts=[None, "#,##0", "0.0%", "#,##0", "0.0%", None, None, None])
    sev_cell(ws, rr, 6, sev)
    ws.cell(row=rr, column=7).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=rr, column=8).alignment = Alignment(wrap_text=True, vertical="top")
for col, w in zip("ABCDEFGH", [34, 10, 11, 12, 10, 12, 44, 38]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------- Collection SEO ----------
ws = wb.create_sheet("Collection SEO")
ws.sheet_view.showGridLines = False
title_block(ws, "Collection SEO — Navigable Collections Worklist",
            f"{len(nav_rows)} collections reachable from site navigation  |  {nav_needs} need SEO work  |  "
            f"store has {C['total']} collections total ({C['empty']} empty)", span=9)
hdr = 4
table_header(ws, hdr, ["Category", "URL", "Title", "Products", "Description?", "Desc Length",
                       "Custom SEO Title?", "Proposed SEO Title", "Needs Work"])
for i, x in enumerate(nav_rows):
    rr = hdr + 1 + i
    data_row(ws, rr, [x["category"], x["url"], x["title"], x["products"], x["has_desc"],
                      x["desc_len"], x["has_seo_title"], x["proposed"], x["needs"]],
             fmts=[None, None, None, "#,##0", None, "#,##0", None, None, None])
    if x["needs"] == "Yes":
        ws.cell(row=rr, column=9).fill = SEV_FILL["CRITICAL"]
    else:
        ws.cell(row=rr, column=9).fill = SEV_FILL["OK"]
for col, w in zip("ABCDEFGHI", [12, 38, 30, 10, 12, 11, 14, 52, 11]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{hdr}:I{hdr + len(nav_rows)}"

# ---------- Technical SEO ----------
ws = wb.create_sheet("Technical SEO")
ws.sheet_view.showGridLines = False
title_block(ws, "Technical SEO — Live Site & Theme",
            f"Live theme: {A['theme']['live_theme']}  |  {A['theme']['theme_count']} themes on store  |  checked {AUDIT_DATE}", span=4)
hdr = 4
table_header(ws, hdr, ["Check", "Status", "Severity", "Details / Recommendation"])
th = A["theme"]
lh = A["live_home"]
tech = [
    ("Theme publishing", "MISMATCH", "CRITICAL",
     "The May 2026 SEO theme work (custom titles, JSON-LD, OG tags, hidden H1) lives on unpublished theme “Prestige Updates TKBS - 05-15-2026”. "
     "The live theme has none of it. Port or publish before any new theme work."),
    ("Homepage title tag", "Generic", "CRITICAL",
     f"Live title is “{lh['title']}” — no keywords, no location. Should target e.g. dresses / women’s boutique / Milford MI."),
    ("Homepage meta description", "Present", "OK", lh["meta_description"][:160] + "..."),
    ("Homepage H1", "Empty", "WARNING", "One H1 exists on the homepage but contains no text — no headline keyword signal."),
    ("Collection meta descriptions", "Missing", "CRITICAL",
     "Live /collections/dresses has NO meta description (collection has no description text to draw from)."),
    ("Product JSON-LD", "Present (app)", "OK", "Product, Offer, Brand, BreadcrumbList rendered on live product pages via app."),
    ("Homepage JSON-LD", "Present (app)", "OK", f"Types found: {', '.join(t for t in lh['json_ld_types'] if t in ('ClothingStore','Organization','WebSite','BreadcrumbList','SearchAction'))}."),
    ("Collection JSON-LD", "Present (app)", "OK", "CollectionPage + ItemList on live collection pages."),
    ("Open Graph tags", "Present (app)", "OK", f"{lh['og_tags']} og: tags on homepage; present on product/collection pages too."),
    ("Twitter cards", "Present (app)", "OK", f"{lh['twitter_tags']} twitter: tags on homepage."),
    ("Theme-level og/JSON-LD backstop", "Missing", "WARNING",
     "theme.liquid itself has no og:, twitter:, or JSON-LD — all depends on the app. If the app is removed, all structured data disappears."),
    ("Canonical tags", "Present", "OK", "Shopify canonical URL in theme."),
    ("Viewport / mobile", "Present", "OK", "Mobile-friendly viewport set."),
    ("Preconnect hints", "Missing", "WARNING", "No preconnect/dns-prefetch for CDN or font hosts — easy page-speed win."),
    ("Microdata snippet", "Present", "OK", "snippets/microdata-schema.liquid exists in live theme (legacy but harmless)."),
    ("Color swatch config", "Empty", "WARNING", "Live theme color_swatch_config is empty — swatches cannot render."),
    ("robots.txt / sitemap", "Standard", "OK", "Shopify defaults; sitemap.xml present and serving."),
    ("Collection filters", "Availability + Price only", "CRITICAL",
     "No Color, Size, or Brand facets on live collection pages — see Filter & Search Readiness Plan."),
    ("Theme housekeeping", f"{th['theme_count']} themes", "WARNING",
     "11 themes stored (2022–2026 backups). Risk of editing/publishing the wrong one; archive old copies."),
]
for i, (label, status, sev, det) in enumerate(tech):
    rr = hdr + 1 + i
    data_row(ws, rr, [label, status, None, det])
    sev_cell(ws, rr, 3, sev)
    ws.cell(row=rr, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 30
for col, w in zip("ABCD", [28, 22, 12, 95]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------- Pages & Blog ----------
ws = wb.create_sheet("Pages & Blog")
ws.sheet_view.showGridLines = False
title_block(ws, "Pages & Blog",
            f"{len(pages)} pages ({pages_pub} published, {len(pages) - pages_pub} draft)  |  "
            f"{blog['articles']} blog articles, last published April 2023", span=5)
hdr = 4
table_header(ws, hdr, ["Page Title", "Handle", "Status", "Content Length", "Note"])
def page_note(p):
    if p["content_len"] == 0:
        return "EMPTY — no content"
    if p["content_len"] < 300:
        return "Very thin — expand or unpublish"
    if not p["published"]:
        return "Draft — publish if still relevant"
    return ""
for i, p in enumerate(sorted(pages, key=lambda x: (not x["published"], -x["content_len"]))):
    rr = hdr + 1 + i
    data_row(ws, rr, [p["title"], p["handle"][:60], "Published" if p["published"] else "Draft",
                      p["content_len"], page_note(p)], fmts=[None, None, None, "#,##0", None])
    if not p["published"]:
        ws.cell(row=rr, column=3).font = F(size=10, color=GRAY)
rr = hdr + len(pages) + 2
section_header(ws, rr, "BLOG", span=5)
rr += 1
blog_rows = [("Total articles", blog["articles"]), ("Published", blog["published"]),
             ("Last published", "April 5, 2023"),
             ("Posts since 2024", sum(v for k, v in blog["by_year"].items() if k >= "2024"))]
for label, v in blog_rows:
    data_row(ws, rr, [label, v, "", "", ""])
    rr += 1
for col, w in zip("ABCDE", [34, 40, 12, 14, 36]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ---------- Local SEO & Entity ----------
ws = wb.create_sheet("Local SEO & Entity")
ws.sheet_view.showGridLines = False
title_block(ws, "Local SEO & Entity Signals",
            "What the site currently tells Google/AI about the business (read from live structured data).", span=4)
hdr = 4
table_header(ws, hdr, ["Signal", "Value on site", "Status", "Action"])
local = [
    ("Business name", "The Clothing Cove", "OK", "Keep consistent everywhere"),
    ("Address (schema)", "414 North Main, Milford, Michigan 48381", "OK", "Verify matches Google Business Profile exactly"),
    ("Phone (schema)", "1-248-685-2500", "OK", "Verify matches GBP & directories"),
    ("Hours (schema)", "Mon–Wed 10–6, Thu–Fri 10–8, Sat 9:30–6, Sun closed", "VERIFY", "Confirm these are the real current hours — schema says Thu/Fri to 8pm"),
    ("Schema type", "ClothingStore (homepage)", "OK", "Good local entity type"),
    ("Social profiles (sameAs)", "Facebook, Twitter/X, Instagram, Pinterest, YouTube", "OK", "Verify links live and active"),
    ("Organization schema", "Missing phone + address", "WARNING", "Second Organization block has no NAP — consolidate with ClothingStore data"),
    ("Google Business Profile", "Not verifiable from site", "VERIFY", "Confirm claimed, categories set, products/photos current"),
    ("Local landing content", "No Milford/Michigan keywords in homepage title", "WARNING", "Add location to homepage title + key page copy"),
]
for i, (label, val, status, action) in enumerate(local):
    rr = hdr + 1 + i
    data_row(ws, rr, [label, val, None, action])
    sev_cell(ws, rr, 3, {"OK": "OK", "WARNING": "WARNING", "VERIFY": "WARNING"}[status])
    if status == "VERIFY":
        ws.cell(row=rr, column=3).value = "VERIFY"
    ws.cell(row=rr, column=4).alignment = Alignment(wrap_text=True, vertical="top")
for col, w in zip("ABCD", [26, 50, 12, 56]):
    ws.column_dimensions[col].width = w

# ---------- AEO Readiness ----------
ws = wb.create_sheet("AEO Readiness")
ws.sheet_view.showGridLines = False
title_block(ws, "AEO Readiness — Will AI Assistants Recommend This Store?",
            "AI engines (ChatGPT, Perplexity, Google AI Overviews) need machine-readable content. Status of each ingredient:", span=5)
hdr = 4
table_header(ws, hdr, ["Ingredient", "Current State", "Severity", "Why It Matters", "Fix"])
aeo = [
    ("Product structured data", "JSON-LD present via app", "OK",
     "AI can read price, availability, brand.", "Maintain; add review/rating data when available"),
    ("Product descriptions", f"{SS['missing_body']:,} in-stock products have none", "CRITICAL",
     "AI cannot recommend a product it can't understand.", "Bulk-generate, even 2–3 sentences each"),
    ("Product tags / taxonomy", f"{SS['missing_tags']:,} in-stock untagged ({SS['missing_tags'] / IN_STOCK:.0%})", "CRITICAL",
     "Tags feed category understanding (occasion, style, material).", "Bulk-tag from type/vendor/attributes"),
    ("Collection page content", f"{C['missing_desc']} of {C['total']} have no description", "CRITICAL",
     "Category pages are exactly what AI reads to learn what the store sells.", f"Write descriptions for the {nav_needs} navigable collections first"),
    ("Fresh, citable content", "Last blog post April 2023", "WARNING",
     "AI cites recent content for “best X for Y” queries.", "Revive blog around shopping questions (mother-of-the-bride, Brighton, occasion dressing)"),
    ("FAQ content", "Strong FAQ page exists (published)", "OK",
     "Q&A format is the easiest content for AI to quote.", "Extend FAQ blocks to top collections"),
    ("Entity clarity", "Generic homepage title; split Organization/ClothingStore schema", "WARNING",
     "“The Clothing Cove” must be unambiguous to AI.", "Unify schema, keyword homepage title, consistent NAP"),
    ("Hours/location accuracy", "Schema present; hours need verification", "WARNING",
     "AI answers “is X open now” from this data.", "Verify hours; keep schema synced with reality"),
]
for i, (label, state, sev, why, fix) in enumerate(aeo):
    rr = hdr + 1 + i
    data_row(ws, rr, [label, state, None, why, fix])
    sev_cell(ws, rr, 3, sev)
    for cc in (2, 4, 5):
        ws.cell(row=rr, column=cc).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 30
for col, w in zip("ABCDE", [26, 38, 12, 48, 48]):
    ws.column_dimensions[col].width = w

# ---------- Fix Roadmap ----------
ws = wb.create_sheet("Fix Roadmap")
ws.sheet_properties.tabColor = GREEN
ws.sheet_view.showGridLines = False
title_block(ws, "SEO & AEO — Fix Roadmap",
            "Priority order: publish what exists, then bulk wins, then content depth. In-stock scope.", span=5)
hdr = 4
table_header(ws, hdr, ["Priority", "Action", "Scope", "Effort", "Expected Impact"])
seo_road = [
    ("P1", "Publish / port the existing theme SEO work to the LIVE theme (homepage title + meta, hidden H1, JSON-LD backstop, OG backstop, preconnect). It is currently sitting unpublished.",
     "Theme only — 0 products", "4–8 hours", "Immediate: keyword-bearing homepage title, schema independence from app"),
    ("P1", "Fix homepage title + empty H1 on live theme", "2 elements", "Included above", "Core ranking signal for “dress shop Milford”-type queries"),
    ("P2", "Bulk-generate SEO titles for in-stock products", f"{SS['missing_seo_title']:,} products", "8–12 hours scripted + spot review", "Every product page gets a keyworded title"),
    ("P2", "Bulk-generate meta descriptions for in-stock products", f"{SS['missing_seo_desc']:,} products", "6–10 hours scripted", "Higher click-through from search results"),
    ("P2", "De-duplicate product titles", f"{SS['dup_title_products']:,} products / {SS['dup_title_groups']} shared titles", "3–5 hours", "Stops self-competition in search"),
    ("P3", "Bulk-tag in-stock products", f"{SS['missing_tags']:,} products", "1–2 weeks scripted + review", "Filtering + AI taxonomy + collection automation"),
    ("P3", "Generate product descriptions (missing + thin)", f"{SS['missing_body']:,} missing + {SS['short_body']:,} thin", "2–3 weeks scripted + review", "Products become indexable and AI-recommendable"),
    ("P3", "Alt-text generation for in-stock images", f"{SS['images_no_alt']:,} images ({SS['products_zero_alt']:,} products with zero alt)", "4–6 hours scripted", "Image search visibility + accessibility"),
    ("P3", "Write descriptions + SEO titles for navigable collections (proposals included on Collection SEO sheet)", f"{nav_needs} collections", "15–25 hours", "Category pages start ranking and feeding AI answers"),
    ("P4", "Publish or prune draft pages; expand thin pages", f"{len(pages) - pages_pub} drafts, several thin pages", "4–8 hours", "Recovers existing content investment"),
    ("P4", "Revive blog: 1–2 posts/month around shopping queries", "Ongoing", "2–4 hours/post", "Fresh citable content for AI + long-tail search"),
    ("P4", "Verify hours, GBP, and unify Organization/ClothingStore schema", "Entity layer", "2–3 hours", "Accurate AI answers about the store"),
]
for i, row in enumerate(seo_road):
    rr = hdr + 1 + i
    data_row(ws, rr, list(row))
    ws.cell(row=rr, column=1).font = F(size=10, bold=True, color=NAVY)
    for cc in range(1, 6):
        ws.cell(row=rr, column=cc).alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[rr].height = 40
for col, w in zip("ABCDE", [9, 70, 28, 26, 48]):
    ws.column_dimensions[col].width = w

out2 = os.path.join(ROOT, "The Clothing Cove - SEO & AEO Plan (June 2026).xlsx")
wb.save(out2)
print("saved", out2)
