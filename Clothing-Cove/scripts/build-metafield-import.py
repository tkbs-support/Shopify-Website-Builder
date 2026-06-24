"""Build a Matrixify metafield import that populates shopify.size and shopify.color-pattern
for option-only products, using a mapping LEARNED from products staff already mapped
(never fabricated). Ambiguous/unseen raw values go to a Needs-Review tab instead of guessing.
"""
import json, os, sys, urllib.request, time
from collections import Counter, defaultdict
sys.path.insert(0, os.path.dirname(__file__))
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
STORE=os.getenv("SHOPIFY_STORE_DOMAIN"); TOKEN=os.getenv("SHOPIFY_ACCESS_TOKEN")
GQL=f"https://{STORE}/admin/api/2024-10/graphql.json"
DATA=os.path.join(os.path.dirname(__file__),"..","data")

def gql(q,v=None):
    p=json.dumps({"query":q,"variables":v or {}}).encode()
    r=urllib.request.Request(GQL,data=p,method="POST",headers={"X-Shopify-Access-Token":TOKEN,"Content-Type":"application/json"})
    for a in range(6):
        try:
            with urllib.request.urlopen(r,timeout=90) as resp:
                out=json.loads(resp.read())
                if out.get("errors"): raise RuntimeError(json.dumps(out["errors"])[:400])
                return out["data"]
        except Exception:
            if a<5: time.sleep(5); continue
            raise

def norm(s): return " ".join((s or "").strip().upper().split())
def mfvals(mf):
    if not mf: return []
    v=mf.get("value") if isinstance(mf,dict) else mf
    try: return json.loads(v)
    except Exception: return [v] if v else []

rows=[json.loads(l) for l in open(os.path.join(DATA,"size-color.jsonl"),encoding="utf-8")]
instock=[p for p in rows if p.get("status")=="ACTIVE" and (p.get("totalInventory") or 0)>0]

SIZE_OPT={"size","shoe size","ring size"}
def opt_values(p, kind):
    out=[]
    for o in p.get("options") or []:
        nm=(o["name"] or "").strip().lower()
        if (kind=="color" and nm in ("color","colors")) or (kind=="size" and nm in SIZE_OPT):
            out+= [v.strip() for v in o["values"]]
    return out
def size_gids(p):
    g=[];
    for k in ("sizeStd","shoeSize","ringSize","accSize"): g+=mfvals(p.get(k))
    return g

# ---- learn raw value -> metaobject GID from UNAMBIGUOUS products (1 value <-> 1 GID) ----
def learn(kind):
    votes=defaultdict(Counter)
    for p in instock:
        vals=opt_values(p,kind)
        gids=mfvals(p.get("colorPat")) if kind=="color" else size_gids(p)
        uv=set(norm(v) for v in vals if v)
        ug=list(dict.fromkeys(gids))
        if len(uv)==1 and len(ug)==1:
            votes[next(iter(uv))][ug[0]]+=1
    learned={}; ambiguous={}
    for raw,c in votes.items():
        top,cnt=c.most_common(1)[0]; total=sum(c.values())
        if cnt>=2 and cnt/total>=0.70: learned[raw]=(top,cnt,total)
        else: ambiguous[raw]=(c.most_common(3),total)
    return learned, ambiguous

color_map, color_amb = learn("color")
size_map,  size_amb  = learn("size")
print(f"learned color values: {len(color_map)} confident, {len(color_amb)} ambiguous")
print(f"learned size  values: {len(size_map)} confident, {len(size_amb)} ambiguous")

# Token lacks read_metaobjects scope, so we can't fetch taxonomy display names.
# The import needs only GIDs (which we have). For human-readable labels we derive each
# GID's meaning from the raw values staff mapped to it — fully data-driven, no guessing.
gid_to_raws=defaultdict(Counter)
for raw,(g,cnt,tot) in list(color_map.items())+list(size_map.items()):
    gid_to_raws[g][raw]+=1
def label(g):
    rs=[r for r,_ in gid_to_raws.get(g,Counter()).most_common(4)]
    return " / ".join(rs) if rs else g.split("/")[-1]
names={g: label(g) for g in gid_to_raws}

# ---- apply to OPTION-ONLY products (have option, lack metafield) ----
def build_import(kind, learned):
    imp=[]; review_vals=Counter()
    for p in instock:
        vals=opt_values(p,kind)
        has_opt=bool(vals)
        has_mf=bool(mfvals(p.get("colorPat"))) if kind=="color" else bool(size_gids(p))
        if not has_opt or has_mf: continue          # only option-only products
        mapped=[]; unmapped=[]
        for v in vals:
            nv=norm(v)
            if nv in learned: mapped.append(learned[nv][0])
            elif v.strip(): unmapped.append(v.strip())
        for u in unmapped: review_vals[norm(u)]+=1
        if mapped:
            gids=list(dict.fromkeys(mapped))
            imp.append({"id":p["id"],"handle":p["handle"],"title":p["title"],
                        "raw":", ".join(dict.fromkeys(vals)),
                        "gids":gids,"names":", ".join(names.get(g,g) for g in gids),
                        "unmapped":", ".join(dict.fromkeys(unmapped))})
    return imp, review_vals

color_imp, color_rev = build_import("color", color_map)
size_imp,  size_rev  = build_import("size", size_map)
print(f"COLOR import-ready products: {len(color_imp)} | review values: {len(color_rev)}")
print(f"SIZE  import-ready products: {len(size_imp)} | review values: {len(size_rev)}")

# ---- write Matrixify workbook ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
H=PatternFill("solid",fgColor="1F3864"); SUB=PatternFill("solid",fgColor="2E5496")
WHITE=Font(color="FFFFFF",bold=True); TH=Side(style="thin",color="D9D9D9"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
wb=openpyxl.Workbook()

def import_sheet(name, key, imp):
    ws=wb.create_sheet(name)
    # Matrixify product import columns + one metafield list column
    cols=["ID","Handle","Title","Metafield: %s [list.metaobject_reference]"%key,
          "(ref) Mapped to","(ref) Raw option values","(ref) Unmapped — left blank"]
    for i,h in enumerate(cols,start=1):
        c=ws.cell(1,i,h); c.font=WHITE; c.fill=H; c.alignment=Alignment(wrap_text=True)
    for ri,row in enumerate(imp,start=2):
        ws.cell(ri,1,row["id"]); ws.cell(ri,2,row["handle"]); ws.cell(ri,3,row["title"][:70])
        ws.cell(ri,4,"\n".join(row["gids"]))            # list items separated by newline
        ws.cell(ri,5,row["names"]); ws.cell(ri,6,row["raw"]); ws.cell(ri,7,row["unmapped"])
        for ci in range(1,8): ws.cell(ri,ci).alignment=Alignment(wrap_text=True,vertical="top")
    for col,w in zip("ABCDEFG",[26,26,40,44,30,34,28]): ws.column_dimensions[col].width=w
    ws.freeze_panes="A2"

import_sheet("Size Import","shopify.size",size_imp)
import_sheet("Color Import","shopify.color-pattern",color_imp)

def dict_sheet(name, learned):
    ws=wb.create_sheet(name)
    for i,h in enumerate(["Raw value","Maps to (taxonomy)","Metaobject GID","Learned from # products"],start=1):
        c=ws.cell(1,i,h); c.font=WHITE; c.fill=SUB
    for ri,(raw,(g,cnt,tot)) in enumerate(sorted(learned.items()),start=2):
        ws.cell(ri,1,raw).border=BD; ws.cell(ri,2,names.get(g,"?")).border=BD
        ws.cell(ri,3,g).border=BD; ws.cell(ri,4,cnt if cnt else "name-match").border=BD
    for col,w in zip("ABCD",[26,26,40,18]): ws.column_dimensions[col].width=w
    ws.freeze_panes="A2"
dict_sheet("Size Mapping", size_map)
dict_sheet("Color Mapping", color_map)

def review_sheet(name, rev, amb):
    ws=wb.create_sheet(name)
    ws.cell(1,1,"Raw values that could NOT be mapped confidently — assign a taxonomy value by hand").font=Font(bold=True,color="C55A11")
    for i,h in enumerate(["Raw value","# products affected","Candidate guesses (top votes)"],start=1):
        c=ws.cell(3,i,h); c.font=WHITE; c.fill=SUB
    ri=4
    for raw,n in rev.most_common():
        ws.cell(ri,1,raw).border=BD; ws.cell(ri,2,n).border=BD
        cand=amb.get(raw)
        if cand: ws.cell(ri,3,", ".join(f"{names.get(g,g)}({c})" for g,c in cand[0])).border=BD
        ri+=1
    for col,w in zip("ABC",[26,18,50]): ws.column_dimensions[col].width=w
    ws.freeze_panes="A4"
review_sheet("Size Needs Review", size_rev, size_amb)
review_sheet("Color Needs Review", color_rev, color_amb)

# README
rm=wb.active; rm.title="README"; rm.sheet_view.showGridLines=False
rm.column_dimensions['A'].width=2; rm.column_dimensions['B'].width=108
def L(r,t,b=False,sz=11,col="000000"):
    c=rm.cell(r,2,t); c.font=Font(bold=b,size=sz,color=col); c.alignment=Alignment(wrap_text=True,vertical="top"); return r+1
r=1
r=L(r,"Metafield bulk import — README",16,True,"1F3864")
r=L(r,f"Populates shopify.size ({len(size_imp)} products) and shopify.color-pattern ({len(color_imp)} products) for "
      "items that already have a variant option but lack the taxonomy metafield. In-stock scoped.",11); r+=1
r=L(r,"How the values were chosen",12,True,"1F3864")
r=L(r,"Each raw option value was mapped to a taxonomy metaobject LEARNED from products your team already mapped "
      "(single-value -> single-metaobject cases, >=70% agreement, >=2 products). Nothing was guessed. Raw values "
      "we hadn't seen mapped are NOT in the import — they're on the 'Needs Review' tabs for a human to assign.")
r=L(r,"See the 'Size Mapping' / 'Color Mapping' tabs to audit every raw->taxonomy decision before importing.")
r+=1
r=L(r,"How to import (Matrixify)",12,True,"1F3864")
r=L(r,"1. REVIEW the Mapping tabs. 2. Import 'Size Import' and 'Color Import' with Matrixify (Products). The metafield "
      "column header is already in Matrixify format; list items are newline-separated metaobject GIDs. "
      "3. TEST FIRST: import 5-10 rows, confirm the metafield + storefront filter look right, then run the rest.")
r=L(r,"After import: in Search & Discovery add/confirm the Color and Size filters, and (per your note) uncheck "
      "'Show filter color swatch' in the theme so colors render as clean text.")
r+=1
r=L(r,"Not included (by design)",12,True,"C55A11")
r=L(r,"Products with NO option (real gaps / no-color items) are NOT here — those need data entered at the source, not "
      "a metafield backfill. They're in the 'Real gap' rows of the corrected readiness workbook.")

# order sheets
order=["README","Size Import","Color Import","Size Mapping","Color Mapping","Size Needs Review","Color Needs Review"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
out="The Clothing Cove - Metafield Import (size + color) June 2026.xlsx"
wb.save(out)
print("saved:", out)
