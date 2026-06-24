# Builds the projected-impact workbook for The Clothing Cove.
# Two levers modeled separately, then combined:
#   Track A  SEO / AEO / GEO  -> organic SESSIONS lift (drives traffic)
#   Track B  Filter & Search  -> revenue-per-session lift (drives conversion)
# No orders API access (403), so the revenue story is an INDEX (baseline = 100),
# not dollars. Client multiplies by their real monthly revenue to dollarize.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

BASELINE_SESSIONS = 3000

months = ["M0 (Jun 26)","M1 (Jul)","M2 (Aug)","M3 (Sep)","M4 (Oct)","M5 (Nov)",
          "M6 (Dec)","M7 (Jan 27)","M8 (Feb)","M9 (Mar)","M10 (Apr)","M11 (May)","M12 (Jun 27)"]

# Cumulative organic-sessions lift %, S-curve (slow crawl/trust start -> plateau)
seo_cons = [0,0,1,1,2,2,3,3,4,4,4,5,5]
seo_exp  = [0,1,2,3,4,6,7,8,9,10,11,11,12]
seo_str  = [0,1,3,5,8,12,16,19,22,25,27,29,30]

# Filter & search: revenue-per-session uplift %, steps up once filters ship, then flattens
flt_cons = [0,0,1,1,2,2,2,3,3,3,3,3,3]
flt_exp  = [0,0,1,2,4,5,6,7,7,8,8,8,8]
flt_str  = [0,1,3,5,8,10,12,13,14,15,16,16,16]

def sessions(lift): return [round(BASELINE_SESSIONS*(1+x/100)) for x in lift]
def idx(lift):      return [round(100*(1+x/100),1) for x in lift]
def combined(s, f): return [round(100*(1+s[i]/100)*(1+f[i]/100),1) for i in range(len(s))]

# ---------- workbook ----------
wb = openpyxl.Workbook()

H_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="2E5496")
BASE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

# ===== Read Me sheet =====
rm = wb.active
rm.title = "Read Me"
rm.sheet_view.showGridLines = False
rm.column_dimensions['A'].width = 3
rm.column_dimensions['B'].width = 104
def line(ws, r, text, size=11, bold=False, color="000000", fill=None):
    c = ws.cell(r, 2, text)
    c.font = Font(size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if fill: c.fill = fill
    return r+1

r = 1
r = line(rm, r, "The Clothing Cove — Projected Impact", 18, True, "1F3864")
r = line(rm, r, "SEO / AEO / GEO  +  Filter & Search Readiness   |   12-month outlook from a 3,000 organic sessions/mo baseline", 11, False, "595959")
r += 1
r = line(rm, r, "TWO LEVERS, MODELED SEPARATELY", 12, True, "1F3864")
r = line(rm, r, "Track A — SEO / AEO / GEO drives TRAFFIC. Optimized collection + product content, schema, and AI-answer "
                "readiness pull more organic sessions. Modeled as a cumulative lift on the 3,000/mo baseline.")
r = line(rm, r, "Track B — Filter & Search Readiness drives CONVERSION. Color/size/material filters + on-site search help the "
                "people already arriving actually find and buy. Modeled as revenue-per-session uplift, NOT extra sessions.")
r += 1
r = line(rm, r, "WHY REVENUE IS AN INDEX, NOT DOLLARS", 12, True, "1F3864")
r = line(rm, r, "The Orders API was not granted (403) at audit time, so we have no baseline conversion rate or AOV. The revenue "
                "columns are indexed to 100 at today's baseline. To dollarize: multiply the index by (current monthly revenue / 100). "
                "Example: at $40,000/mo, a revenue index of 140 = ~$56,000/mo.")
r += 1
r = line(rm, r, "WHERE AEO & GEO SIT", 12, True, "1F3864")
r = line(rm, r, "AEO (answer engines / AI Overviews) and GEO (generative engines: ChatGPT, Perplexity, Gemini, Claude) mostly "
                "affect CITATIONS and click-through, not raw session counts — direct AI-referral traffic is still small but "
                "growing. Its measurable session contribution is folded into the upper half of the Track A band; treat raw "
                "AI-citation share as a separate KPI to watch, not a line on this chart.")
r += 1
r = line(rm, r, "SCENARIO DEFINITIONS", 12, True, "1F3864")
r = line(rm, r, "Conservative — partial execution. For a store that has DECLINED for years, this case is mostly STABILIZATION "
                "(arresting the slide), not net growth. SEO +5% / Filters +3% by M12.")
r = line(rm, r, "Expected — solid execution of the category-page content + the core filters. SEO +12% / Filters +8% by M12.")
r = line(rm, r, "Strong — category + brand/'near me' long-tail rank and the filters convert well. SEO +30% / Filters +16% by M12.")
r += 1
r = line(rm, r, "THE HEADLINE (Expected, M12): traffic ≈ +12%, revenue-per-session ≈ +8%, COMBINED revenue ≈ +21% — range +8% "
                "(conservative/stabilization) to +51% (strong). The conversion (filter) lever is the more controllable half; "
                "the traffic lever depends on rankings we have not yet baselined.", 11, True, "C55A11", BASE_FILL)
r += 1
r = line(rm, r, "ILLUSTRATIVE — NOT A PROMISE. There is no Search Console / GA4 baseline yet, so these are scenario shapes, not "
                "committed targets. Phase 0 (connect GSC + GA4, pull a 16-month baseline) sets the real numbers; this model is "
                "re-calibrated against that data before any target is committed. SEO compounds: expect noise for 8–12 weeks.", 10, False, "808080")

# ===== Projection sheet =====
ps = wb.create_sheet("Projection")
ps.sheet_view.showGridLines = False

# Title
ps.merge_cells("A1:K1")
t = ps.cell(1,1,"Projection — Sessions (Track A) & Revenue Index (Track B + Combined)")
t.font = Font(size=14, bold=True, color="1F3864")

# Two-row header
top = ["", "SEO / AEO / GEO — Organic sessions/mo", None, None,
       "Filter & Search — Revenue/session (index, base 100)", None, None,
       "Combined revenue index (base 100)", None, None]
sub = ["Month", "Conserv.", "Expected", "Strong",
       "Conserv.", "Expected", "Strong",
       "Conserv.", "Expected", "Strong"]
hr1, hr2 = 3, 4
for i, v in enumerate(top, start=1):
    if v is None: continue
    c = ps.cell(hr1, i, v); c.font = WHITE; c.fill = H_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
ps.merge_cells(start_row=hr1,start_column=2,end_row=hr1,end_column=4)
ps.merge_cells(start_row=hr1,start_column=5,end_row=hr1,end_column=7)
ps.merge_cells(start_row=hr1,start_column=8,end_row=hr1,end_column=10)
for i, v in enumerate(sub, start=1):
    c = ps.cell(hr2, i, v); c.font = WHITE; c.fill = SUB_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

sc, se, ss = sessions(seo_cons), sessions(seo_exp), sessions(seo_str)
fc, fe, fs = idx(flt_cons), idx(flt_exp), idx(flt_str)
cc, ce, cs = combined(seo_cons, flt_cons), combined(seo_exp, flt_exp), combined(seo_str, flt_str)

start = hr2 + 1
for i, m in enumerate(months):
    row = start + i
    vals = [m, sc[i], se[i], ss[i], fc[i], fe[i], fs[i], cc[i], ce[i], cs[i]]
    for j, v in enumerate(vals, start=1):
        c = ps.cell(row, j, v); c.border = BORDER
        if j == 1:
            c.font = Font(bold=(i in (0,12)))
        else:
            c.alignment = Alignment(horizontal="center")
            if j in (2,3,4): c.number_format = "#,##0"
            else: c.number_format = "0.0"
        if i == 0: c.fill = BASE_FILL
ps.column_dimensions['A'].width = 13
for col in "BCDEFGHIJ": ps.column_dimensions[col].width = 10.5

# Footnote
fn = start + len(months) + 1
ps.cell(fn,1,"Combined index = (sessions ÷ 3,000) × (1 + filter uplift). Multiply index by (monthly revenue ÷ 100) for dollars.")
ps.cell(fn,1).font = Font(italic=True, size=9, color="808080")

# ---- Chart 1: organic sessions band ----
ch1 = LineChart(); ch1.title = "Track A — Organic sessions/mo (SEO/AEO/GEO)"
ch1.style = 2; ch1.height = 7.5; ch1.width = 16
ch1.y_axis.title = "Sessions / month"; ch1.x_axis.title = None
data = Reference(ps, min_col=2, max_col=4, min_row=hr2, max_row=start+len(months)-1)
cats = Reference(ps, min_col=1, min_row=start, max_row=start+len(months)-1)
ch1.add_data(data, titles_from_data=True); ch1.set_categories(cats)
ch1.y_axis.scaling.min = 2800
ps.add_chart(ch1, "A20")

# ---- Chart 2: SEO-only vs Filters-only vs Both (Expected) ----
# build a small helper block for these three Expected series
hb = fn + 3
ps.cell(hb,1,"Expected scenario — revenue index comparison").font = Font(bold=True, color="1F3864")
labels = ["Month","SEO only","Filters only","Both"]
for j,v in enumerate(labels, start=1):
    c=ps.cell(hb+1,j,v); c.font=WHITE; c.fill=SUB_FILL; c.alignment=Alignment(horizontal="center")
seo_only = idx(seo_exp)
for i,m in enumerate(months):
    row=hb+2+i
    for j,v in enumerate([m, seo_only[i], fe[i], ce[i]], start=1):
        c=ps.cell(row,j,v); c.border=BORDER
        if j>1: c.number_format="0.0"; c.alignment=Alignment(horizontal="center")
ch2 = LineChart(); ch2.title = "Revenue index — SEO only vs Filters only vs Both (Expected)"
ch2.style = 2; ch2.height = 7.5; ch2.width = 16
ch2.y_axis.title = "Revenue index (base 100)"
d2 = Reference(ps, min_col=2, max_col=4, min_row=hb+1, max_row=hb+1+len(months))
c2 = Reference(ps, min_col=1, min_row=hb+2, max_row=hb+1+len(months))
ch2.add_data(d2, titles_from_data=True); ch2.set_categories(c2)
ch2.y_axis.scaling.min = 95
ps.add_chart(ch2, "A36")

# ---- Chart 3: combined revenue band ----
ch3 = LineChart(); ch3.title = "Combined revenue index — Conservative / Expected / Strong"
ch3.style = 2; ch3.height = 7.5; ch3.width = 16
ch3.y_axis.title = "Revenue index (base 100)"
d3 = Reference(ps, min_col=8, max_col=10, min_row=hr2, max_row=start+len(months)-1)
ch3.add_data(d3, titles_from_data=True); ch3.set_categories(cats)
ch3.y_axis.scaling.min = 95
ps.add_chart(ch3, "A52")

out = "The Clothing Cove - Projected Impact.xlsx"
wb.save(out)
print("saved:", out)
print("M12 Expected -> sessions:", se[12], "| filter idx:", fe[12], "| combined idx:", ce[12])
print("M12 Conservative combined:", cc[12], "| Strong combined:", cs[12])
