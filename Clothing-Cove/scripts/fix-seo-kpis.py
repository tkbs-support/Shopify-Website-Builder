# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font
F="The Clothing Cove - SEO & AEO Plan.xlsx"
ORANGE="ED7D31"
wb=openpyxl.load_workbook(F); ex=wb["Executive Summary"]
# Card 1 (title) -> hygiene: relabel + recolor red->orange (A4:B6)
ex["A4"]="NO CUSTOM SEO TITLE (LIVE)"
for r in range(4,7):
    for c in (1,2): ex.cell(r,c).fill=PatternFill("solid",start_color=ORANGE)
# Card 2 (was tags) -> live collections without description (keep its red fill = critical gap)
ex["D4"]="COLLECTIONS W/O DESCRIPTION (LIVE)"
ex["D5"]=218/254
ex["D6"]="218 of 254 navigable collections — the real content gap"
# Card 3 -> relabel to distinguish (all collections incl. non-navigable junk)
ex["G4"]="COLLECTIONS W/O DESCRIPTION (ALL 544)"
ex["G6"]="423 of 544 (incl. non-navigable/event)"
# kill any leftover all-caps IN-STOCK across the SEO plan
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and "IN-STOCK" in c.value:
                c.value=c.value.replace("IN-STOCK","LIVE")
wb.save(F)
print("KPIs updated. charts:", len(ex._charts))
for cell in ("A4","D4","D5","D6","G4","G6"): print(f"  {cell}: {ex[cell].value}")
