"""Add 'Progress (June 11)' notes to the SEO & AEO Plan for items already
addressed via the TKBS theme work (May SEO edits + June 11 live sync + publish).

IMPORTANT: build-reports.py regenerates the SEO workbook from scratch and wipes
these overlay notes. Always run this script AFTER build-reports.py.

Reflects final state as of June 11: TKBS theme PUBLISHED; hours verified
(Thu 10-8, Fri 10-6) and synced across theme + SEO Manager app schema.
"""
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

FILE = "The Clothing Cove - SEO & AEO Plan (June 2026).xlsx"

HEADER_FILL = PatternFill("solid", start_color="FF1F3864")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
DONE_FONT = Font(name="Arial", size=10, color="FF1E7B33")
PARTIAL_FONT = Font(name="Arial", size=10, color="FFB45F06")
WRAP = Alignment(wrap_text=True, vertical="top")

# sheet -> (note column, header row, [(row, done?, text)])
NOTES = {
    "Fix Roadmap": ("F", 4, [
        (5, True, "Done — updated theme synced with all recent live-site changes — theme PUBLISHED June 11"),
        (6, True, "Done in updated theme — keyworded title, meta description & H1 in place — LIVE as of June 11"),
        (16, False, "Nearly done — hours verified & consistent across theme + SEO app schema (Thu 10–8, Fri 10–6, confirmed live June 11); remaining: GBP check"),
    ]),
    "Technical SEO": ("E", 4, [
        (5, True, "Done — updated theme synced with all recent live-site changes — theme PUBLISHED June 11"),
        (6, True, "Done in updated theme — \"The Clothing Cove | Women's Clothing, Dresses & Brighton Jewelry | Milford, MI\" — LIVE as of June 11"),
        (8, True, "Done in updated theme — keyworded H1 added — LIVE as of June 11"),
        (15, False, "Partly done — theme-level ClothingStore + FAQ JSON-LD added in updated theme; og:/twitter tags still app-only"),
    ]),
    "Local SEO & Entity": ("E", 4, [
        (8, True, "Done — hours verified (Thu 10–8, Fri 10–6); theme and SEO Manager app schema both match, confirmed on live site June 11"),
        (11, False, "Partly done — updated theme adds ClothingStore schema with full name/address/phone; app's Organization block still has no NAP"),
        (13, True, "Done in updated theme — Milford in homepage title, H1 and a new content section — LIVE as of June 11"),
    ]),
    "AEO Readiness": ("F", 4, [
        (10, True, "FAQPage JSON-LD added in updated theme — LIVE as of June 11"),
        (11, False, "Partly done — keyworded homepage title + ClothingStore schema in updated theme — LIVE as of June 11"),
        (12, False, "Hours verified & all site schema consistent (Thu 10–8, Fri 10–6, June 11); GBP still to confirm"),
    ]),
}

wb = load_workbook(FILE)
for sheet, (col, hdr_row, notes) in NOTES.items():
    ws = wb[sheet]
    h = ws[f"{col}{hdr_row}"]
    h.value = "Progress (June 11)"
    h.fill = HEADER_FILL
    h.font = HEADER_FONT
    h.alignment = WRAP
    ws.column_dimensions[col].width = 48
    for row, done, text in notes:
        c = ws[f"{col}{row}"]
        c.value = ("✓ " if done else "◐ ") + text
        c.font = DONE_FONT if done else PARTIAL_FONT
        c.alignment = WRAP
    print(f"{sheet}: {len(notes)} notes in column {col}")

wb.save(FILE)
print("Saved.")
